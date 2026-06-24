"""
BVX Auswertung + Verladeplanung - Streamlit Version

Installation:
    pip install streamlit pandas plotly openpyxl reportlab pillow

Ausführen:
    streamlit run bvx_auswertung_streamlit_verladung_varianteA_pdf_logo_ansichten_v3.py

Hinweis:
    Hauptausgabe ist der A3-PDF-Pritschenplan.
    Excel bleibt als Begleitdatei mit den kleinen BSD-/Pritschenzetteln erhalten.
    Die Transportabmessungen sind Beispiel-/Stammdaten und müssen intern geprüft und angepasst werden.
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import re
import math
import io
import base64
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Any, Tuple
from collections import Counter
from datetime import datetime

DEFAULT_LOGO_B64 = '''/9j/4AAQSkZJRgABAQEBLAEsAAD/4QBWRXhpZgAATU0AKgAAAAgABAEaAAUAAAABAAAAPgEbAAUAAAABAAAARgEoAAMAAAABAAIAAAITAAMAAAABAAEAAAAAAAAAAAEsAAAAAQAAASwAAAAB/+0ALFBob3Rvc2hvcCAzLjAAOEJJTQQEAAAAAAAPHAFaAAMbJUccAQAAAgAEAP/hDIFodHRwOi8vbnMuYWRvYmUuY29tL3hhcC8xLjAvADw/eHBhY2tldCBiZWdpbj0n77u/JyBpZD0nVzVNME1wQ2VoaUh6cmVTek5UY3prYzlkJz8+Cjx4OnhtcG1ldGEgeG1sbnM6eD0nYWRvYmU6bnM6bWV0YS8nIHg6eG1wdGs9J0ltYWdlOjpFeGlmVG9vbCAxMC4xMCc+CjxyZGY6UkRGIHhtbG5zOnJkZj0naHR0cDovL3d3dy53My5vcmcvMTk5OS8wMi8yMi1yZGYtc3ludGF4LW5zIyc+CgogPHJkZjpEZXNjcmlwdGlvbiByZGY6YWJvdXQ9JycKICB4bWxuczp0aWZmPSdodHRwOi8vbnMuYWRvYmUuY29tL3RpZmYvMS4wLyc+CiAgPHRpZmY6UmVzb2x1dGlvblVuaXQ+MjwvdGlmZjpSZXNvbHV0aW9uVW5pdD4KICA8dGlmZjpYUmVzb2x1dGlvbj4zMDAvMTwvdGlmZjpYUmVzb2x1dGlvbj4KICA8dGlmZjpZUmVzb2x1dGlvbj4zMDAvMTwvdGlmZjpZUmVzb2x1dGlvbj4KIDwvcmRmOkRlc2NyaXB0aW9uPgoKIDxyZGY6RGVzY3JpcHRpb24gcmRmOmFib3V0PScnCiAgeG1sbnM6eG1wTU09J2h0dHA6Ly9ucy5hZG9iZS5jb20veGFwLzEuMC9tbS8nPgogIDx4bXBNTTpEb2N1bWVudElEPmFkb2JlOmRvY2lkOnN0b2NrOjc5MDJhZWNhLWMyZWYtNDY1My1hNTBhLWMxM2NmYzRiYmE4MjwveG1wTU06RG9jdW1lbnRJRD4KICA8eG1wTU06SW5zdGFuY2VJRD54bXAuaWlkOjY5ZGNlMzk1LTZhYWYtNGI0Zi1iNTg2LWE2NTVhMDgzOTI0MDwveG1wTU06SW5zdGFuY2VJRD4KIDwvcmRmOkRlc2NyaXB0aW9uPgo8L3JkZjpSREY+CjwveDp4bXBtZXRhPgogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAKICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIAogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAKICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIAogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAKICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIAogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAKICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIAogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAKICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIAogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAKICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIAogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAKICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIAogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAKICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIAo8P3hwYWNrZXQgZW5kPSd3Jz8+/9sAQwAFAwQEBAMFBAQEBQUFBgcMCAcHBwcPCwsJDBEPEhIRDxERExYcFxMUGhURERghGBodHR8fHxMXIiQiHiQcHh8e/9sAQwEFBQUHBgcOCAgOHhQRFB4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4e/8AAEQgBaAFxAwERAAIRAQMRAf/EAB0AAQACAgMBAQAAAAAAAAAAAAAHCAUGAwQJAgH/xABEEAABAwMCAwUFBQcCBAYDAAABAAIDBAURBgcSIUEIEzFRYSIycYGRFEJSYqEVIzNygpKiFrEkQ1PBFyU0c5OywuHw/8QAGAEBAQEBAQAAAAAAAAAAAAAAAAMCAQT/xAAgEQEBAQEAAwEBAQEBAQAAAAAAAQIREiExQVEiAzJx/9oADAMBAAIRAxEAPwC5aAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAg4qiop6dvFPNHE3ze4NH6oPymq6Wpyaephmx/03h3+yDmQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQaNuxujpjbm3Nlu87p6+ZpNNb4CDNL6+TW5+8eXlk8l2Trl1Iqfr/ALQO4Op5pI6Gv/09QE+zBbziTH5pT7RP8vCPRUmZEru1FldWVdfM6auq6irlccl88rpHH5uJXWXHTTTU0gkppZIHjmHRPLCPmEEj6G3x3F0pLG2O+S3ajbjNJciZmkeQefbb8j8ly5lamrFq9m96NM7isFEzNrvjWcT7fO8EvA8XRO++Po4dR1WLnik1Kk5ZaEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQR7vtuXRbbaSNZwsqLtV8UVupXHk9+Ob3fkbkE+eQOq7J1nV4obqC8XPUF5qbxea2Wtr6p/HNNIebj0A8gPAAcgOQVUbeuggICAg5aOpqKOrhq6SeWnqIHiSKWJxa+Nw5hzSOYIQXd7NW7bdwLI+1Xl7Gajt8YM+AGiqi8BM0dDnAcB4Eg+BGJ6nFs66mFZaEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEHzNIyKJ8sj2sYwFznOOAAPElB5673a4n1/uFX3rvHGgjcae3RnwZA0nhOPNxy4/zeirJyIava0hdcEBAQEBBm9CamuGj9W27UlscftFFKHlmcCVh5PjPo5uR9D0SzrsvK9GdO3aiv1hob1bpO9pK6nZPC7za4ZGfXofVRXl676AgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICCLu1JqR+m9m7s6CQx1Ny4bfCQeY733z/YHrWZ7Z1eRQ34DAVERAQEBAQEBBcrsU6kfdNuKzT88hdLZqstjBPhDLl7R8nd4FPc9q4vpPKy2ICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgrD27rkRTaVs7XHD5Kiqe3+UNY3/AO7lvCf/AEVaW0xAQEBAQEBBPvYguRptybrbC4hlbay/Hm6KRuP0e5Z38bx9XGU1RAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEFQ+3Q9x1zp6M+622SEfEy8/9gqY+Jb+q8LTAgICAgICAgl7sgPc3fG3geD6Kqa74cAP/AGCzr43j6vKpqiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIKo9u2icy96VuOPZkpqmDPq1zHf/AJFbwntWpbTEBAQEBAQEE2di+idU7xvqQMtpLXPIT5cTmMH+5Wd/G8fV11NUQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBBBXbVsTrjtdTXiJhc+017JHkdI5AY3f5Fi1j6xuelMVRIQEBAQEBAQWl7Ctic2m1JqWRhDZHxUMLvPhBkf+rmfRY3VMLPLCggICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICDEazsVLqfSl00/Wj9xcKV8DjjPDxDAcPUHB+SRyzrzfvlrrbJea2z3KIxVlFO+Cdp6PacH5HxHoQrIV00BAQEBAQfTGve9rI2Oe9xDWtaMlxPIAepKD0N2S0j/AKI20tFhka0VbIu+rCOs8h4n/Qnh+DQpW9q+ZyN0XHRAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBBUftu6atFv1HaNSUsrYrhdGPiqqcD+IIg3hm9DhwYfP2fIqmKnuK6LSYgICAgIJi7IumrRqHdZs10la51pp/t1NTOHKaUODQ4+jOIOx548is6vpvE7V4VNUQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEFbO27pC419ttOsKKN81NbWPp61rRnumPc0tk+HEME9MhbxU9z9VPW0xAQEBAQWH7FGkLjU6uq9ZyRvittHTyUkTyMCaZ/DxAeYa0cz5kDzWd38UxP1bxTUEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEDI80BAQEBAQEGEvertK2RzmXjUdot72+LKisjY76E5TjnYwjN2ts3ycA1zYM+tawfqu8p5Rn7TqfTd3x+yr/aq8npT1kch+gK5w6x+42o9K2LSlwl1TXUkdDJTvjkhkeC6cOaQWNZ4uJzjAXZ0tjzj5fdBA6AnmAqoCAgICAgvp2b9RaUuG1VhttjraVk9FRsiq6QyASxzAfvC5vj7TiXcXgcqep7WzZxvt1v9jtTeK6Xm3ULfOpqmR//AGIWWutcm3Y21hfwP1zp/i9K1h/2K7yueUZSz650ZeHiO16qslZI44DIa6Nzj8s5TlOxsOVx0QEBAQEDI80BAQEBAQEBAQEBAQEBAQEBAQda619Ha7bUXG4VMdLSU0bpZppHYaxjRkklBUndLtM6guVbLRaFY2025pLW1ksQfUzD8Qa7LYwegwT8PBUmf6ld/wARizdnc2Oo78a5vvHnOHVGW/2kcP6LvI55VKu2PaevVFVRUWvKWO5UTiGmvpYgyeL1cwey8fDhPx8Fy4/jU3/VrLLdLferVTXW1VkNZRVUYkhnidlr2nqD/wD2FNR3EBBqe6Ov7Bt5p512vk5LnkspaWPBlqX491o9OrjyA8ei7J1y3im25e+GutazyxC4yWW1uJDaGgkLMt/PIMOefoPRUmZErq1GJALy8jLjzLj4n5rrJk+aD8AaHcQaA7zA5oPpxLncTiXO8ycn6oPxAQEBAQEAcnBw5OHgR4j5oDwHu4ngOd5u5lAQfhDScloJ9Qg3zbzdvXOh54xa7zNU0LT7VBWuM0Dh5AE5Z8WkfNcslamrFx9mN1rDuVanvo/+CutM0Gst8jwXx55cbT99hP3ungQCp2cUmupBXGhB8VE0NNBJUVErIYY2l8kj3BrWNAySSeQAHVBWHdntOyR1U1r29poZGMJa661TOJrz5xR9R+Z3j+HqtzP9Tu/4hWs3c3OrKk1EmuL01xOcQzCJg/paAP0WvGM+Vbxt12kta2Ksih1O5uorbkCTja2OpYPNrwAHH0cOfmFy5js3VvdIaitGq9PUl+sdU2poapnFG8DBB8C1w6OByCD4EKdnFJessjogICAgICAgICAgICAgIK6duDU1RQ6Xs+lqWUsbdJnz1QB96KLh4Wn0L3A/0reIxutM7Oey9jv+mzrnXUn/AJQXO+y0zpu6jexhw6WV+QeHIIAyPDJ6Bd1r8jmc/tS/LtLsnrO0zU9iobKTGOEVdlqm95CehJY4g/BwIKz2xrxlVN3d29u+3Oqn2e4nv6aUGSirGtw2ojz446OHg5vQ46ELcvUrONz7Mm7EmhtQNsV5qXf6buEuHlx5Ucx5CUeTTyDh8HdDnmp1rOuLuscHNDmkEEZBBU1XUvdyo7NZ6u7XGdsFHRwvnnkd4NY0ZJ+gQeem6+uLluBrKqv9e57IiTHRUxORTQA+ywevVx6kn0VpOIW9atBFLPPHBBFJLLI4NZHG0uc9x8AAOZPoEcSjpzs/bo3mmZU/sOG2RP5t/aFS2J/9g4nD5gLnlGpmshcezXuhSxl8NLaK0j7sFeA4/wB7Wj9VzyjvhUZar0zf9KXMW3UdpqbZVuZ3jY5gPbbkjiaQSCMgjIPRal6zZxiEcEBAQEBAQEHesNnut+usNqstvqLhXTZ7uCBvE92Bk/IDmUPqUbV2cN0q6Nr5rbbreCM4qq5vF9GByz5Rrwr4vfZ03RtsD5orXQ3JrBkto6xrn/Jrw3PyXfKHhUWXOgrrZXy0Fyo6ijq4XcMsE8ZY9h9WnmF1l3dH6iuulNSUV/stQYK2kk4mHPsvH3mOHVrhyI//AEnOuy8eh23uqaDWejrdqS3coayIOMZOXRPHJ7D6tcCPkpWcWl7GfXHVRu1tuxJdrjPoHT9URbaV/DdJo3f+olB/g5/A0+95u5eDedMz9T3r8QnoLSl41rqil09ZIRJVVBy57vchYPekeejR+pwBzIWreMSdW/sWym0miLBHJqaC3VsuAJrheJmtY52OYa1xDGDyA5+ZKn5WqeMjQ97tkdI1Wi6rWu2r6draON089NSVHfU1RE3m8xnJ4XNGTgHBwRgFdmvfK5rM52Md2HdTzw6ivGkZZS6lqaf7fA0nk2Rha1+P5mub/am4Yv4tmsKCAgICAgICAgICAgICAgrB26rLUOj01qKNjnU8TpqOZwHJjncL2fXhePkt4T3G89m6bTes9k7HbKylpa51ld3FTSTND2slaXcDnMPIgtcHAkEZ8OYXNeq1n3Gp9oOG0bb7i6L1bpGmp7XdaqsMFdTUjRGysp8sB42N5H3uHOPEjqAu59xnXq9iVN9tAQbhaBqrW1jBc6fNRbpnfcmA5NJ/C4eyfjnoFmXla1Ox5+zxSwTSQTxOiljcWSRvGC1wOC0jzByFVFbfsibpi8Wxmg77U5uVFHm3SyO51EDR/DyfFzB9W/ylY1P1XGvx2+2vql9s0FQ6appOGW81OZgDz7iLDiPgXlg+RXMT2bvpT2CKWeeOCCN8ssjwyNjBlz3E4AA6kk4VEl5ez5s/btAWaG53OniqdT1MeZ5yA4UwI/hR+WPAuHNxz0wFPV6tnPEtrLTB651TZ9G6Zq7/AHypEFJTt8BzfK8+7GwdXE8gP+wK7J1y3jz53G1bcdb6wrtSXP2Zal2I4QctgiHJkY9AOvUknqqycRt7WvI4ICAgICAgIO/p6719gvlFerXOYK6imbNBJ5OHmOoPMEdQSEHoFtDr+07iaShvNvc2KpZiOtpC7LqaXHNp82nxaeo9cgSs4vL2NyXHWg7zbX2PcewPp6uOOmu0LD9huDW+3E7o134mE+LT8RgrsvGdTqhGoLTcLDe62zXWnNPXUUzoZ4z0cPI9QeRB6ggqqN9LHdhvVL21N80bPITG5ouNK0nwOQyUD6xn6rG5+qYv4krtNbnt0FpI2+1zgahujHMpMHnTx+Dpz8PBvm74FczOta1xRlziSXOcSfEknJ9SSqIrwdlbbtujdCMu9wp+C93ljZ5+Ie1DD4xxenI8R9TjoFPV7Vszkanp79nbidqbUtJq6OKuptOwuitFuqPahBa5rXycB5OPPi5/iH4Rjt9Rz7r2lDch2ldA7e6kvUVBRW5tRRvidHAwRtqZnNc2NvAMAuJdjOM48eQ5Znuu31FdexFY6mq3Gr7zwO+y222mFz8cjJK5oa3+1jit7vpjE9rkqaogICAgICAgICAgICAgIMRrHTlp1ZpyssF7phUUNWzhe3OHNPiHNPRwOCD5hJeOWdVeq9hd1tE6gkrtu7+2eJ/stmiqxSzlmeTZGu9h31I64Cp5S/WPGz42fbPY3V1fral1nuteft9RSSNkhpDUGd73sOWcbscLWNPPgb4ny555dT5HZm97VkVhtTftibfGw6uZrG3QYtt5fip4RyiqgOZ9OMDi+Id5qmb+Jbn6g2119Za7lTXK3VMlLWUsrZYJozh0bwcghaYS/rmtuO+Vvt98tUrJdTWih+z11iaMPmYHFxqKb/qA59pnvNwMZ5ZzP8t3/Tk7IOjxfN1JLnX07u409H37mSMIxUuJbGCD4EYe7B6tCavoxPa6/gFNVpW6G52lNvbeZr1XNfWubmCggIdUTHphv3R+Z2AuyWuWyKT7ubl6g3Ivgrbq8U9FASKOgicTHAD1/M89XH4DA5KknEbetJXXBB2ZqCtht1PcZaWWOjqZHxwTObhsrmY4w09ccQz6nCDrINt20261TuFcZqTTlHG9lOAaipnfwQw58AXYJJPPAAJ5LlvHZLXf3R2m1jt3HDU3ymp5qCZ/dsrKSQyRceM8LsgFpPPGRg45FJZXbmxoa6y7VFb62uhq5qOllnZRw/aKkxtz3UXEGl5/KC4ZPTPNB1UBBse3mtL/AKE1HHe9P1XdTAcM0L+cVRHnmx7eo9fEHmEs67LxdbaHeTSu4VNHTxTttt74f3ttqHgPJ6mN3hI34cx1AUrOKzUqSlxpU7twaSjpLxaNaU0Ya2tBoawgcjIwcUTj6lvE3+kLeKnufrS9q6ap2sq6bcfVL5Le77PK21WdwxVXPjaW8RaecUIyCXuGTgYHn2+/Tk9e0da11Nd9X6lq9QXuo76sqnZOOTI2j3WMHRrRyA+Z5krUnGbet67M+37tdbhwyVkHHZrSW1VaSPZkcD+7i/qcMkfhafNc1eR3M7V8AOWFJZAm9+yV5vWrf9d7f3Ztsvx4XTxGZ0PePaOESRyN91xAAIPI46c87mvysaz+xH1Tsrvfrm5U7Nb32NlNAfZlq64TiMdSyKPkXY6nHxXfKT4542/Vk9sNDWXb/S0NiszHOaHd5UVEmO8qJSOb3Y+AAA5AABYt63JxtK46ICAgICAgICAgICAgICAgINAv2822ljvb7NcdV0kdZG/glbGx8jYneT3taWtI65PLqu+NZ8o3qjqaespYquknjnp5mCSKWNwc17SMggjkQR1XGmJ1zpm2aw0rX6du8XHS1kfASPejd4te09HNIBHwSXjlnXnzuNo68aF1XVafvMWJYjxQzNGGVERPsyM9D5dDkHwVpeo2cYKjqaijqoqukqJaeohcHxSxPLHscPAtcOYPwRxKml+0FuHYu8JktNyklDRLPV0I76ThGG8ckZaXkDq7JXPGNTdfmpe0NufeoHwR3amtMT+R/Z1MI3/3uLnD5EJ4wu6iyrqaisqpKqrqJaioldxSSyvL3vPmXHmfmusuJAPIZJwEEgae0TQ2i2Q6o3Fknt1rkbx0NrZ7NddfIMaecUXnI7HLw8QVzv8AGuftZDfW8y3izaBldQ0lvifY5KiGkpWcMUEclS8MY0ejI2jPiTk9Uhr8Reusrxdj6hpaXZKgqIGtEtZV1M07h4l4kLBn4NY0Kevq2PjaN/aGluGzWq4KtrSxltlmaT918Y42H4hzQuZ+mvjz1VUUj9nS4VVBuLL9i7ozz2a4RxiVnGwuFO6Roc0+8MxjI6rmvjWfpW6XtOu6F9+28gEFyDO9uGmOLMsXLLpKTP8AFi/J7zfDBGAnefTnfiOXtcx7mPaWuaS1zSMEEeII6FdZfiD9Y5zHtexxa5p4muacEHzB6FBJmk999zdOwMpor/8AtKnZybHcohOQPLj5P+riuXMamrGQ1J2iNxL3SsgcLHR8DxIySG3h72PHg5plLw1wycEDITxhd1Fl2uNwu1wluN0rqmurJjmWeokL3vPqSusufTVkumpL7SWSzUj6uvq5BHDG3z6kno0DmT0AT47J16A7PaDoNvNFU1ipC2aoP72tqQMGeYgcTvgMAAdAB6qVvatmcjcSQASTgDquOo8/8bNrv24bP/rChFQJO74y1/c8WcY73HB8849V3xrPlEhMc17Q5pDmkZBB5ELjT9QEBAQEBAQEBAQEBAQEBAQEGlb63evsO0WprrbHujrIaFwikb70ZcQ3iHqA4n5Ls+ua9REHZY0BoPUW09TXXey0N0r6irngqpJ2cckIaRwtYfFnskOyMEl2crWresZksSp2fbBddNbX0FouomY6KeodTxTfxIqd0zjE1w6O4SCR0zjos6va3mciQFx1WvtTa429r71SaJvVDLXSU/E6puVE4Ge1SOA4QwHlIer4yRyx97GN5l+p6s+K96q0JeLLQi80bob5p6Q/ubvbsyQH0ePehf5teBj1W+sWNUHMZByPMI4YQfrQXSNjaC57jhrQMk/AdUElaC2O3D1a6OWOzutNC7mau5Awtx5tZjjd9APVcupGpm13tTy6Y2l1DU2KwW+PUOqKEtbPebmxrqellLckQU/Npc3I9p5OCOQT6eojO9Xa53y6zXS8V9RX1s7syzzvLnu9M9AOgHIdF1nrcd1z32nNuqtg/dv0tHCD04o6iZrh9SFyfrV/GgrrK23Ye1RFU6Yu2kZpAKihqPtkDSeboZcB2Pg8f5hY3FMX8bR2vtTxWPaSptTZAKu9yto425592CHSu+HCMf1BczPbW76UhVEW/wDZ+xHuXDWP/hUdsuFTIfJraSUc/m4LmvjWfrQ7dU1NFNTVdJUS01TBwvilieWPjcPAtcOYPwXWUlWrVFh1/caO0bh28RXOpkZTxalt4bFUNc48LTUx44Jm5Iy7k4Bc5z413v139f8AZ83A0w6Segom6hoG8xNQAmUD80J9rP8ALxJNSlxYiaphmpah1NUxSQTsOHRSsLHtPq08wusvjB8kA8hk8h6oNl0fofUOqI5KuipWU1rh9qpula/uKOnb1LpXcj8G5PolvHZOpr7O+sttdGa4GnKNrqh9dH3D9TVQ7sTTZGImMPOKA9CTlzscXLGM6lsbzZKtmpqNa3St12u+3GobZYn8FzqrdNFS+1w5eWEAZ6Z8M+q7Prl+Ic3J212/sPZwnqnWCnobhSW2OWKqki4Kv7UQ0BryeZLnHhLTy54A5BalvWbJxtfZGu9ddtmKFtdI+Q0NTNRwveckxMcOAZ9A7h+DQua+u4+JdWWhAQEBAQEBAQEBAQEBAQEBB0r/AGujvdkrbPcIu9pK2B8EzPNjgQfnzQUjqpNwuz3ryopaKcmkqHZjdNGXUtxiafZJHR4B54Ic056EZr60j7zUkW3tZgUrRctEPdUAc3U1wAYT8HMyPqVnwa82n7g9pbWWoKWWhsNLBpymkBDpYZDLUkeQkIAZ8WjPquzMcu6jDQejdSa91ALXYaOSqnc7iqJ5Ce7gBPN8r+nn5noCVq3jMlq9W0G3Nn260mLNRYqaic95X1b24dUyYxkjo0DkG9B5kkqVvVpOPm/7Rba32d1RcdG2p0zvekhjMDifMmMtynaeMYiDYHaWJ/GNIxPPlJVzuH0L8LvlXPGNw03ovSWmwP2Fpu1W5w+/BSta/wDuxn9VztdkkZmtnipKOaqmPDHDG6R58g0ZP+y468z77cprze6+71Li6auqZKl5Pm9xd/3VnnrrU8M1TURU9PE6WaV7Y442+L3OOAB8SQEFld/9qnae2E0u+lBnqdNZbXPHPLaggyuH5RLjHkCVjN9qazyKzLabZ9rtYVmhdcW7UlGHSCnfw1EIP8aB3J7PiRzHqAlnY7Lyt07VOoLlqLcSnrJXNdZHW+KWyPjdlk1PIOIyfzF+Q4dOEDos5+O6vaiRaZWL7IG38d/tmqr1dI3toayiks0Lm8ie8AMzmn0HAPmVnVUxOoH1VZK3TepLjYLi3FVb6h9PIccncJ5OHoRgj0K1GLOMZ7X3SQ7oR0PQo49IdtL1/qLb6wXtzuJ9bb4ZZD+csHF/llSv1ee47t/01p6/xd1fLJbrk3GAKqmZJj4FwyFzpZ1pdTsTtPUPL36Mo2E/9KaWMfRrwF3yrnjHesuz22NonbPRaLtXeN8HTxmcj/5C5O13xjM660ZY9Y6QqNMXWlb9ilaO77oBroHt9x7OgLT4fTwKS8LOqJ7s7aaj26vLqS7wGaglcRSXCNh7moHl+V/mw8/LI5qkvUrmxtW2faC1ro6kitlX3N/tsQDY4qx7hNE0fdbKMnHo4Ox6LlzKTdiSJ+1pS/Zj3Ghqnv8AHISXFvAD8QzP6Lng15ov1Jq/cXffVFHYIIGd13nHDQUoIp4OhmlceZwD7x+DRk89cmWbbpcrbbSlHonRNs01QuMjKOLD5SMGWQnie8/FxJ9OQU7eqycjYlx0QEBAQEBAQEBAQEBAQEBAQEGM1Lp+y6ltUlqv1spbjRSc3RTsDhnzHUH1GChZ1C9/7LWh6yodLarperUHHPdNkZMxvoOMcX1JWvOseEfWn+y5oShnbNdbjebtwnPdPlbDG70PAA7/ACTzp4RM2mrBZdN2tlssNspbdRs8IqeMNGfM+Z9TkrLcnHFqTVGm9Nxsff77bbWH82faqlsZd8ATk/JOOd45NPajsGoqd1RYrzb7nE33nUtQ2Th+ODy+acO9ZRHRBqW8tf8AszajVVaDh0dpqA0+pjLR+pXZ9cvx50gcIDfIYVUEtdk7TLdRbw0NRPHx01nidXyZ8ONuGxD+9wd/SuavprE7V55vb6O62qqtlwgbUUlVC6GaJ3g9jhgj6FSWeeu7uha/b3WtVYasPkps97Q1DhyngJ9l38w91w8x6hWl6hZytQRxIuiXjWujpdvqgh13oTJXaakcebn44p6PPlIBxtH42+q5fXtqe5xqei9N3TVuqKHTtohL6yrk4BxA4iaPee/ya0ZJ+GPFdt45J16H6E01btH6St2nLW0imooQwOI9qR3i57vVziSfipW9Wk4qz23NMtt+t7ZqeCPhju1MYZyOs0OME/Fjm/2LeKnue+q+rTC9vZNrvtmxdjYTl1K6enP9Mz8foQp6+rZ+JWWWhB0b3eLTZKI1t4uVHb6YcjLUzNjZnyy4jmh1jtNa10lqWZ0Fg1JarlM0ZdFT1TXvA8+EHOPVd5Y5LKyt1t1BdrfLQXOip62kmbwyQzxh7Hj1B5FcdQ1qfsybe3Sd89skulje457ummD4h8GyA4+AK1N1i4jGWjsq6Op6gSXK/wB8rowc920xwg/EtaT9CE86eETJovR+mtG239n6atFNb4XYLzGMvkI6vecucfiVy3rUnGeXHRAQEBAQEBAQEBAQEBAQEBAQEBAQEHUvFWaC01lc2MymngfKGDxdwtJx88IKj7U7T1u9dLcNwNXamq4ZKyqfHEIGNe8luM83cmsbkNa0DwHRUuuekpny91o2srPqLY7dZsdrupfU07GVNLUsbwCphcT7EjM8wS1zXN5jqOi7P9Ry/wCavTpS8Q6g0zbL5Tt4YrhSRVLG5zwh7Q7HyzhSWntk0EXdqquFFsXqAZw6pbDTt/rmYD+mVrP1nXxQ0+Koitn2F7IIdM6g1C9g4qusZSRuPjwxM4j/AJSfosbUxFkFhRG3aI0hpjVOgJzqO4UtofROD6O5zcm08riGgOPVjiQ0j4HxAXc3lZ1JYoxqiw3XTV9qbLeaV1NWU7sPbnLXA82va7wc1w5hw8Qqz2lZx0qGqqaGtgraOd9PU08jZYZWHDo3tOWuHqCAUcXa7M9NpG9Wu4bg2iiZBfLvLwXaMEFtPM0AyMjH3WPd+9x14x5ACeu/Fs8+pjWWkMdseyC57OT17WAyWqshqgeoaT3bv0fn5LWfrO56UjVEVyuxFXNn2tuFDn26S7S8vIPjjcP1yp7+q4+J5WW3zK9scbpHkNa0EknoAgoPfbnqDe7eKCj+1lorqt0NvjkJMVHTjJyG+fA0uJHNx+WK/Ij/AOqkPczs+v0DpCXWem9WV0lfZw2ok442xOwCMviczm1wznBzkZGVya76auOe1itnNQ1uq9sbBqC4ta2srKQOnwMBzwS0uA6AlufmsWcrcvY21cdEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQCAQQRkFBp2hNBUeiblcjYK6eGzV8rqh1rkaHRwTnGXRO8WtIHNhyPDGF23rknFO+1DqeLUu8N1lpXd5TW1rbfE4c+IxZ4yP63OHyVMz0lq9q6G1drlsu22m7VUAtmpbZTxyg+IeIxxD5HKnfqs+NlXHVfu3Ddm023lptAOJK+5iQjzZExxP8Ak5i3j6xv4p6FtJfXsuWn9k7Iafa5hZLVxvrH5Hj3r3OH+PCp6+rZ+JOWWlU+3Bq6SW5WjRNNLiGKP7fWNB957iWxNPwAe75hbxP1Pd/EY6TqWbhadg0JdJGC/wBDG7/TVdI7BkAHEaCRx8WuwTGT7ruXgcLV9e2Z79I4ljkilfFLG+ORji17HjDmuBwQR0IPJdZTV2PdYSWDcwWCeXFBfmdyWk8m1DAXRu+JHE35jyWdT03i+111NVre6VoF+241FZ+DjdVW2eNg/PwEt/yAXZ9cvx5vtJLQSMEjJVUFm+wldmsuGqLG4+1JHBWRj+UuY7/dixtT/mtSsKOKthFRSTU7jgSscwnyyMf90HnjtzdpNv8Adq119ex0f7JuRgrGkc2sBdFJ9ASfkq33EJ6q824ekIteWeC0Vl4qaeySuElZBSYa6saCHNYZOZazIBPCMnlzHWcvFrOtktdBR2u201ut9PHTUlNE2KGKMYaxjRgAfALjrsoCAgICAgICAgICAgICAgICAgICAgICAgICDXNzr67TG3t+v7DiWhoJZYv/AHA0hn+RC7Prl9RRHZWys1Lu3py1Vv76KavbLUcX/MbGDK4H48H6qlvpGe69ER4KS4gpL2wtVsv+6X7IppQ+lsUH2Y4OQZ3Hil+nsN+LSqZnpLd9ojsNrqr5fKGy0TS6pr6hlNEB+J7g3Pyzn5LTE9vSuy0EFqs9Ha6UYgo4GQRDyaxoaP0Ci9DtnwQUE7TVbJXb5ame857maOBvo1kLB/vn6qufiOvrq9nzTQ1Vu7YrdI97IIZvtszmOLXcMPt4BHMZcGjPqlvIZna2Ptd6ag0/u7NV0rGsgvNM2uLW+AlyWSfUtDvi4rmb6NzlRno2tltur7LcYXFslNcaeVpHmJWlark+vS4KK4fBB5z7vacfpPcu/WIsLIoKx76fPWF/txn+1wHyVp8Qs5WW7PGq2aP3as9xqZRHRVLzRVbicARy4AcfQPDD8iuanY7m8r0CHgpLCCiva0scNm3puMkDA2K5wRVxaPxuBa/6uYT81XPxHf1aDsy3+XUOzFiqKiQyVFJG6hlcfEmFxY0/2hqnqcqmb2JKXGhAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQajvNZKjUW1mpLNSNLqmpt8ghaPFzwOJrfmWgfNdn1zU7FAtFahrNK6ttmpKBodUUFQ2Zsb+QePBzD5AtJHzVbOoy8q8Gkd8ttr/AGuOqfqSjtU5YDLS3CQQyRnqOfJ3xaSFPxqs1Gmbvdo7TlptM9BoeqZeLvK0sZVMafs1MT9/J/iOHQDlnxPQ9mf65dz8U9qJpaiokqKiV8s0ry+SR5y57ickk9SSSVtJYnsZbezV19k19coCKKh4oLdxD+LMRh8g9GAlufxOP4VnV/FMT9W3U1AoKG9qm2Ptu+N9c5pDKwQ1cZx4h0bWn/JjlXPxHf13eyFXRUW91BFK4N+2UdRTsJ/Fwh4H+BTXwx9Znts3WCt3SoLdC4OdbrY1s3o+R7n4/t4T81zHx3f1FG21qkvm4WnrTE0l1TcoGnHRoeHOPyaCVq/GZ9ekY8FFcQVx7Z23stztVPry1wF89uj7i4tYObqfOWyf0EnP5XZ6Leb+Mbn6qURyIK2ktRsR2ibXBZaXTuv6iSmnpWCKC6FpeyVgGGiXGS1wHLiwQfE4Pji5/imd/wBS1ed6NsbXbnVsmsLZUgNy2Kkl7+V58gxuTn44WfGteUUs3g1tPuBr2t1HJAaaB4bDSwOOTFCzPCCfM5Lj6lVk4lb2rfdlCy1Fl2UtP2pjmSV75a4McMEMkd7H1aGn5qevquJyJWWWhAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEFV+0F2fblLd6vVOg6ZtTHUvM1Xa2ENex55ufDnkQTklnIgnlnOBua/qesfxW26W+vtVU6lulDVUE7Dh0VTC6JwPwcAtpvy3UVbcqltPbqOprZ3nDY6eJ0rifg0EoJ02i7N+ob5Vw3HW0clktQIcaTiH2qcfhwP4Q8yfa8gPFZuv55Mf1YfVNVq7T1BT2DbPQ9DUw0cDWtkrKptNSxNA5Rsb7z3Y8TyA8yc4x/wDW72fGlbc7619Rrh2h9x7BHpy8ulEUUkbj3TpD7rHBxPDxfdcCWuyPDIXbn9jk175U6rLat/bb0ZJW2W263o4i59u/4WuwOYhe7LHn0a8kf1reL+J7n6q5pu71mn9QW++W84q6CpZURZ8C5pzg+h5g+hW056bZvpE6q11LqqCeSqtmpY23OgnfzPA4Broj5Oic0sI6AN81yfGtfUo9irQ0lbf6vXdbCRS0LXUtASPfmcMSOHo1p4fi8+S5u/jWJ+rbKaiG95t7W6S1BBpDSto/1BqaZzWmAE8ELne6whvNzz48IxgcyQtTPWbrnpnNC37cysljpNwNCUFLS1Y4PtFvrGzCLI92aIkkNPhlpcB1GOY5ZPwlv6hrejs2V0NZPetvI2T00hL32lzw18R8T3LjyLfykgjwBPgNTX9Z1j+K73qz3ayVbqW82utts7Tgx1UDoj/kOfyW0+OrSQT1k7YKOGWplecNjhYZHE/BuSUE7bJ9nm+325U921tRS2myxuDzRzezUVf5S3xjYepOHEcgBnKzdfxvOP6uNBFHBCyGFjY42NDWMaMBoHIADoFNV9oCAgICAgICAgICAgICAgICAgICAgICAgICAgICDhqqWmqmcFTTxTN8pGBw/VApaSlpWcFNTxQN8o2Bo/RBzICCv3bL0I27aTh1rb4P+Ps/s1TmD2n0pPjy/A4hw8gXLeb+MbnrrN9mjdul1rp+Gw3qsYzU1DHwPEjsGsjb4St83Y94efPwK5qcM66ly70NFc7XVW64wR1FHUxOinikHsvY4YIPyWW3n1utolmkr3JJaa+G76dqKiSOguMEgkYSwkOhe4chIzwI6jDhyPKsvULOPrQt6sldajonWdTNS2SapFRR3GKPjktc5wHPDfvRPHJ7fMBw5gpf7HZfyr56Gt1itOkrbbtMmB1ogga2lfC8Pa9v4uIcnEnJJ6klSqsaJ2gN3bdt5Y5KOhmhqdS1LMUtNni7kH/myDo0dAfePLwyRrOeua1xGvY50VV3C43Dc6/95UTzvkioJZubpHuP76fPmT7AP867q/jOJ+rPrCgg46iCCoj7ueKOVh+69ocPoUHHSUNHSZ+y0kEGfHu4w3/YIOwgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIOKspoKykmpKqJk0EzHRyRvGWvaRggjqCCgo1vltFfdt7/JdbPFVz6edL3lHWwlxfSHPJkhHNpHgH+BHXOQqy9R1njS7ruHre7Ws225ayvNXREcLoZK1xa4eTuftD45XeRztbn2fdPa0vtwltlvsDLnpSvc1l2jrw6Ojc0eD2vHMTN+66PLh4HkuasdzK2PdLs16mslRLXaNc6/W3Jc2nc5rauIeWDhsnxGD6Lk07cX8RnbRubpeSS222PV9nc8kPp6eOpiDj/K0Yz6hd9Vn3G/7U7B6v1jeWXTWEFbZ7S5/eVElWSKuq9Gtd7Qz+N3h0BXLqRqZt+rl2i3UVptlNbLbTR0tHSxNighjGGsY0YACmq7SAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICD8e1r2lr2hzXDBBGQQgwJ0To01X2o6TsPf5z3n7Oi4s+eeFd7XORnYo44o2xxsaxjRhrWjAA9AuOvpAx8fqgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICD/2Q=='''


def get_embedded_default_logo() -> Optional[bytes]:
    try:
        return base64.b64decode(DEFAULT_LOGO_B64)
    except Exception:
        return None



# =============================================================================
# Datenmodelle
# =============================================================================

@dataclass
class Operation:
    op_type: str
    diameter: Optional[float] = None
    length: Optional[float] = None
    depth: Optional[float] = None
    volume: float = 0.0
    count: int = 1
    x: Optional[float] = None
    y: Optional[float] = None
    z: Optional[float] = None
    production_state: Optional[str] = None
    plunge_type: Optional[str] = None


@dataclass
class Part:
    name: str
    length: float = 0.0
    width: float = 0.0
    height: float = 0.0
    part_no: str = ""
    part_id: str = ""
    unit: str = ""
    profile: str = ""
    surface: str = ""
    grade: str = ""
    user_attribute_2: str = ""
    volume_m3: float = 0.0
    attributes: Dict[str, str] = field(default_factory=dict)


@dataclass
class AnalysisResult:
    file_name: str
    part_count: int
    operation_count: int
    total_operation_count: int
    total_volume: float
    machined_volume: float
    operations: List[Operation]
    parts: List[Part]
    part_dimensions: Optional[Dict[str, float]] = None


# =============================================================================
# BVX Parser
# =============================================================================

class BVXParser:
    """Parser für BVX-Dateien (XML und Text-Format)."""

    OPERATION_TYPES = [
        'Drilling', 'SawCut', 'Slot', 'Step', 'Pocket', 'BlindSlot', 'BlindStep',
        'Mill', 'Countersink', 'Thread', 'FrameBox', 'BirdsMouth', 'Mortise',
        'Tenon', 'Notch', 'Rabbet', 'Chamfer', 'Groove', 'Dado', 'LapJoint',
        'DovetailJoint', 'FingerJoint', 'ScarfJoint', 'HalfLap', 'CrossLap',
        'LowerBoomSingleStepJoint', 'CADPosition', 'CADPositionList', 'Kappen'
    ]

    def parse(self, file_content: str, file_name: str = "uploaded.bvx") -> AnalysisResult:
        """Analysiert BVX-Dateiinhalt und gibt Analyseergebnis zurück."""
        if file_content.strip().startswith('<?xml') or '<RectangularPart' in file_content:
            return self._parse_xml_format(file_content, file_name)
        return self._parse_text_format(file_content, file_name)

    def _extract_xml_attribute(self, tag: str, attr_name: str) -> Optional[str]:
        """Extrahiert XML-Attributwert aus einem Tag."""
        pattern = rf'{re.escape(attr_name)}="([^"]*)"'
        match = re.search(pattern, tag, re.IGNORECASE)
        return match.group(1) if match else None

    def _extract_all_xml_attributes(self, tag: str) -> Dict[str, str]:
        """Extrahiert alle XML-Attribute aus einem Tag."""
        return {key: value for key, value in re.findall(r'([\w:.-]+)="([^"]*)"', tag)}

    def _get_attr_first(self, attrs: Dict[str, str], keys: List[str], default: str = "") -> str:
        """Sucht ein Attribut case-insensitive über mehrere mögliche Namen."""
        lower_map = {k.lower(): v for k, v in attrs.items()}
        for key in keys:
            value = lower_map.get(key.lower())
            if value is not None:
                return value
        return default

    def _safe_float(self, value: Optional[str], default: float = 0.0) -> float:
        """Sicher einen String zu Float konvertieren."""
        if value is None or value == "":
            return default
        try:
            return float(str(value).replace(',', '.'))
        except (ValueError, TypeError):
            return default

    def _make_part_from_tag(self, tag: str, fallback_name: str) -> Part:
        attrs = self._extract_all_xml_attributes(tag)

        name = self._get_attr_first(attrs, ['Name', 'PartName'], fallback_name)
        dim_x = self._safe_float(self._get_attr_first(attrs, ['DimensionX', 'Length', 'Laenge']))
        dim_y = self._safe_float(self._get_attr_first(attrs, ['DimensionY', 'Width', 'Breite']))
        dim_z = self._safe_float(self._get_attr_first(attrs, ['DimensionZ', 'Height', 'Hoehe', 'Höhe']))

        # BVX-Grundannahme für Verladung:
        # X = Länge, Y = Breite, Z = Höhe
        volume_m3 = (dim_x * dim_y * dim_z) / 1_000_000_000 if dim_x and dim_y and dim_z else 0.0

        part_no = self._get_attr_first(attrs, ['PartNo', 'PartNumber', 'Number', 'Bauteilnummer', 'PieceNo'])
        part_id = self._get_attr_first(attrs, ['PartId', 'PartID', 'Id', 'ID', 'Guid'])
        unit = self._get_attr_first(attrs, ['Unit', 'Pak', 'Package', 'Paket'])
        profile = self._get_attr_first(attrs, ['Profile', 'Profil'])
        surface = self._get_attr_first(attrs, ['Surface', 'Oberflaeche', 'Oberfläche'])
        grade = self._get_attr_first(attrs, ['Grade', 'Quality', 'Qualitaet', 'Qualität'])
        user_attribute_2 = self._get_attr_first(attrs, ['User_Attribut_2', 'User_Attribute_2', 'UserAttribut2'])

        return Part(
            name=name,
            length=dim_x,
            width=dim_y,
            height=dim_z,
            part_no=part_no,
            part_id=part_id,
            unit=unit,
            profile=profile,
            surface=surface,
            grade=grade,
            user_attribute_2=user_attribute_2,
            volume_m3=volume_m3,
            attributes=attrs,
        )

    def _parse_xml_format(self, file_content: str, file_name: str) -> AnalysisResult:
        """Parst XML-Format BVX-Dateien."""
        parts: List[Part] = []
        operations: List[Dict[str, Any]] = []

        # RectangularPart mit Inhalt
        part_matches = re.finditer(r'<RectangularPart\b([^>]*)>([\s\S]*?)</RectangularPart>', file_content, re.IGNORECASE)
        for i, match in enumerate(part_matches):
            attrs_text = match.group(1)
            content = match.group(2)
            full_tag = f'<RectangularPart{attrs_text}>'
            part = self._make_part_from_tag(full_tag, f'Bauteil_{len(parts) + 1}')
            parts.append(part)

            op_pattern = '|'.join(self.OPERATION_TYPES)
            self_closing = re.findall(rf'<({op_pattern})\s+([^>]*)/>', content, re.IGNORECASE)
            content_tags = re.findall(rf'<({op_pattern})\s+([^>]*)>[\s\S]*?</\1>', content, re.IGNORECASE)

            for op_type, attrs in self_closing + content_tags:
                op = self._parse_operation(op_type, f'<{op_type} {attrs}>')
                operations.append(op)

            # Kappen grob als zwei Sägeschnitte je Bauteil erfassen
            blade_thickness = 6
            if part.width and part.height:
                operations.append({
                    'type': 'Kappen',
                    'diameter': blade_thickness,
                    'length': part.width,
                    'depth': part.height,
                })
                operations.append({
                    'type': 'Kappen',
                    'diameter': blade_thickness,
                    'length': part.width,
                    'depth': part.height,
                })

        # Self-closing RectangularPart tags
        self_closing_parts = re.finditer(r'<RectangularPart\b([^>]*)/>', file_content, re.IGNORECASE)
        for match in self_closing_parts:
            full_tag = f'<RectangularPart{match.group(1)}/>'
            parts.append(self._make_part_from_tag(full_tag, f'Bauteil_{len(parts) + 1}'))

        # Globale Operations-Sektion
        global_ops_match = re.search(r'<Operations>([\s\S]*?)</Operations>', file_content, re.IGNORECASE)
        if global_ops_match:
            ops_content = global_ops_match.group(1)
            global_ops = re.findall(r'<([A-Z][a-zA-Z]+)\s+([^>]*)/>', ops_content)
            for op_type, attrs in global_ops:
                op = self._parse_operation(op_type, f'<{op_type} {attrs}/>')
                operations.append(op)

        return self._build_result(file_name, parts, operations)

    def _parse_operation(self, op_type: str, tag: str) -> Dict[str, Any]:
        """Parst einzelne Operation aus XML-Tag."""
        x = self._safe_float(self._extract_xml_attribute(tag, 'X'))
        y = self._safe_float(self._extract_xml_attribute(tag, 'Y'))
        z = self._safe_float(self._extract_xml_attribute(tag, 'Z'))

        drill_diam = self._safe_float(self._extract_xml_attribute(tag, 'DrillDiam'))
        hole_depth = self._safe_float(self._extract_xml_attribute(tag, 'HoleDepth'))
        dim_x = self._safe_float(self._extract_xml_attribute(tag, 'DimensionX'))
        dim_y = self._safe_float(self._extract_xml_attribute(tag, 'DimensionY'))
        dim_z = self._safe_float(self._extract_xml_attribute(tag, 'DimensionZ'))
        depth_attr = self._safe_float(self._extract_xml_attribute(tag, 'Depth'))
        prod_state = self._extract_xml_attribute(tag, 'ProductionState')
        plunge = self._extract_xml_attribute(tag, 'PlungeType') or self._extract_xml_attribute(tag, 'PocketPlungeType')

        diameter = None
        depth = None
        length = None

        if op_type == 'Drilling':
            diameter = drill_diam if drill_diam > 0 else None
            depth = hole_depth if hole_depth > 0 else None
        elif op_type == 'Slot':
            diameter = dim_x if dim_x > 0 else None
            length = dim_y if dim_y > 0 else None
            depth = z if z > 0 else (depth_attr if depth_attr > 0 else None)
        elif op_type == 'Step':
            diameter = y if y > 0 else None
            length = z if z > 0 else None
            depth = 20
        elif op_type == 'Pocket':
            diameter = dim_x if dim_x > 0 else None
            length = dim_y if dim_y > 0 else None
            depth = dim_z if dim_z > 0 else (depth_attr if depth_attr > 0 else None)
        elif op_type == 'SawCut':
            diameter = dim_x if dim_x > 0 else None
            length = dim_y if dim_y > 0 else None
            depth = dim_z if dim_z > 0 else (depth_attr if depth_attr > 0 else None)
        else:
            diameter = dim_x if dim_x > 0 else (drill_diam if drill_diam > 0 else None)
            depth = hole_depth if hole_depth > 0 else (depth_attr if depth_attr > 0 else (z if z > 0 else None))
            length = dim_x if dim_x > 0 else None

        return {
            'type': op_type,
            'diameter': diameter,
            'depth': depth,
            'length': length,
            'x': x,
            'y': y,
            'z': z,
            'production_state': prod_state,
            'plunge_type': plunge,
        }

    def _parse_text_format(self, file_content: str, file_name: str) -> AnalysisResult:
        """Parst Text-Format BVX-Dateien."""
        lines = [line.strip() for line in file_content.split('\n')]
        parts: List[Part] = []
        operations: List[Dict[str, Any]] = []

        for i, line in enumerate(lines):
            if 'BEGIN PART' in line or 'PART_DEF' in line:
                name = self._extract_text_value(line, 'NAME') or f'Part_{len(parts) + 1}'
                dims = self._extract_dimensions(lines, i)
                volume_m3 = dims['length'] * dims['width'] * dims['height'] / 1_000_000_000
                parts.append(Part(name=name, volume_m3=volume_m3, **dims))

            if 'OPERATION' in line:
                if 'DRILL' in line or 'BOHR' in line:
                    diameter = self._extract_number(line, ['DIA', 'DIAMETER', 'D'])
                    depth = self._extract_number(line, ['DEPTH', 'TIEFE'])
                    if diameter and depth:
                        operations.append({
                            'type': 'Bohrung',
                            'diameter': diameter,
                            'depth': depth,
                            'x': self._extract_number(line, ['X']),
                            'y': self._extract_number(line, ['Y']),
                            'z': self._extract_number(line, ['Z']),
                        })

                if 'MILL' in line or 'FRÄS' in line or 'FRAS' in line:
                    diameter = self._extract_number(line, ['DIA', 'DIAMETER', 'D', 'WIDTH'])
                    depth = self._extract_number(line, ['DEPTH', 'TIEFE', 'Z'])
                    length = self._extract_number(line, ['LENGTH', 'LÄNGE', 'L'])
                    if diameter and depth:
                        operations.append({
                            'type': 'Fräsung',
                            'diameter': diameter,
                            'depth': depth,
                            'length': length or diameter,
                            'x': self._extract_number(line, ['X']),
                            'y': self._extract_number(line, ['Y']),
                        })

                if 'COUNTERSINK' in line or 'SENK' in line:
                    diameter = self._extract_number(line, ['DIA', 'DIAMETER', 'D'])
                    depth = self._extract_number(line, ['DEPTH', 'TIEFE'])
                    if diameter and depth:
                        operations.append({
                            'type': 'Senkung',
                            'diameter': diameter,
                            'depth': depth,
                            'x': self._extract_number(line, ['X']),
                            'y': self._extract_number(line, ['Y']),
                        })

                if 'THREAD' in line or 'GEWINDE' in line:
                    diameter = self._extract_number(line, ['DIA', 'DIAMETER', 'D', 'M'])
                    length = self._extract_number(line, ['LENGTH', 'LÄNGE', 'L'])
                    if diameter and length:
                        operations.append({
                            'type': 'Gewinde',
                            'diameter': diameter,
                            'length': length,
                            'x': self._extract_number(line, ['X']),
                            'y': self._extract_number(line, ['Y']),
                        })

        if not parts:
            parts.append(Part(name='Bauteil', length=1000, width=500, height=200, volume_m3=0.1))

        return self._build_result(file_name, parts, operations)

    def _extract_text_value(self, line: str, key: str) -> Optional[str]:
        """Extrahiert Wert aus Text-Zeile."""
        pattern = rf'{key}[=:\s]+([^\s,;]+)'
        match = re.search(pattern, line, re.IGNORECASE)
        return match.group(1) if match else None

    def _extract_number(self, line: str, keys: List[str]) -> Optional[float]:
        """Extrahiert Zahlenwert aus Text-Zeile."""
        for key in keys:
            pattern = rf'{key}[=:\s]+([\d.]+)'
            match = re.search(pattern, line, re.IGNORECASE)
            if match:
                return float(match.group(1))
        return None

    def _extract_dimensions(self, lines: List[str], start_idx: int) -> Dict[str, float]:
        """Extrahiert Dimensionen aus Textzeilen."""
        dims = {'length': 1000, 'width': 500, 'height': 200}
        for i in range(start_idx, min(start_idx + 20, len(lines))):
            line = lines[i]
            length = self._extract_number(line, ['LENGTH', 'LÄNGE', 'LAENGE', 'L'])
            width = self._extract_number(line, ['WIDTH', 'BREITE', 'W'])
            height = self._extract_number(line, ['HEIGHT', 'HÖHE', 'HOEHE', 'H'])
            if length:
                dims['length'] = length
            if width:
                dims['width'] = width
            if height:
                dims['height'] = height
            if 'END PART' in line or line == '':
                break
        return dims

    def _calculate_volume(self, op: Dict[str, Any]) -> float:
        """Berechnet Volumen einer Operation in m³."""
        op_type = op.get('type', '')
        diameter = op.get('diameter') or 0
        depth = op.get('depth') or 0
        length = op.get('length') or 0

        volume_mm3 = 0

        if op_type == 'Drilling':
            radius = diameter / 2
            volume_mm3 = math.pi * radius ** 2 * depth
        elif op_type in ['Slot', 'Step', 'Pocket']:
            width = diameter or length
            vol_length = length or width
            vol_depth = depth or 10
            volume_mm3 = width * vol_length * vol_depth
        elif op_type in ['Kappen', 'SawCut']:
            blade_thickness = diameter or 6
            width = length or 100
            height = depth or 100
            volume_mm3 = width * height * blade_thickness
        else:
            if diameter and depth:
                radius = diameter / 2
                volume_mm3 = math.pi * radius ** 2 * depth
            elif diameter and length and depth:
                volume_mm3 = diameter * length * depth

        return volume_mm3 / 1_000_000_000

    def _group_operations(self, operations: List[Dict[str, Any]]) -> List[Operation]:
        """Gruppiert identische Operationen."""
        groups: Dict[str, List[Dict[str, Any]]] = {}

        for op in operations:
            diameter = f"{op.get('diameter', 0):.2f}" if op.get('diameter') else 'none'
            length = f"{op.get('length', 0):.2f}" if op.get('length') else 'none'
            depth = f"{op.get('depth', 0):.2f}" if op.get('depth') else 'none'
            prod_state = op.get('production_state') or 'active'
            plunge = op.get('plunge_type') or 'none'
            key = f"{op['type']}_{diameter}_{length}_{depth}_{prod_state}_{plunge}"

            if key not in groups:
                groups[key] = []
            groups[key].append(op)

        result: List[Operation] = []
        for ops in groups.values():
            first = ops[0]
            total_volume = sum(self._calculate_volume(op) for op in ops)
            result.append(Operation(
                op_type=first['type'],
                diameter=first.get('diameter'),
                length=first.get('length'),
                depth=first.get('depth'),
                volume=total_volume,
                count=len(ops),
                x=first.get('x'),
                y=first.get('y'),
                z=first.get('z'),
                production_state=first.get('production_state'),
                plunge_type=first.get('plunge_type'),
            ))

        return result

    def _build_result(self, file_name: str, parts: List[Part], operations: List[Dict[str, Any]]) -> AnalysisResult:
        """Erstellt Analyseergebnis."""
        total_operation_count = len(operations)
        grouped_ops = self._group_operations(operations)

        total_volume = sum((p.volume_m3 or (p.length * p.width * p.height / 1_000_000_000)) for p in parts)
        machined_volume = sum(op.volume for op in grouped_ops)

        part_dims = None
        if parts:
            part_dims = {
                'length': parts[0].length,
                'width': parts[0].width,
                'height': parts[0].height,
            }

        return AnalysisResult(
            file_name=file_name,
            part_count=len(parts),
            operation_count=len(grouped_ops),
            total_operation_count=total_operation_count,
            total_volume=total_volume,
            machined_volume=machined_volume,
            operations=grouped_ops,
            parts=parts,
            part_dimensions=part_dims,
        )


# =============================================================================
# Allgemeine Hilfsfunktionen
# =============================================================================

def read_uploaded_text(uploaded_file) -> str:
    raw = uploaded_file.read()
    for encoding in ('utf-8', 'latin-1', 'cp1252'):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode('utf-8', errors='ignore')


def format_volume(volume: float) -> str:
    """Formatiert Volumen für Anzeige."""
    if volume < 0.001:
        return f"{volume * 1_000_000:.2f} mm³"
    if volume < 1:
        return f"{volume * 1000:.4f} dm³"
    return f"{volume:.6f} m³"


def parts_to_dataframe(parts: List[Part], density_kg_m3: float = 500.0) -> pd.DataFrame:
    rows = []
    for idx, part in enumerate(parts, start=1):
        volume_m3 = part.volume_m3 or (part.length * part.width * part.height / 1_000_000_000)
        row = {
            'Index': idx,
            'Name': part.name,
            'Bauteilnummer': part.part_no or part.name,
            'PartId': part.part_id,
            'Pak/Unit': part.unit,
            'Profil': part.profile,
            'Oberfläche': part.surface,
            'Qualität': part.grade,
            'User_Attribut_2': part.user_attribute_2,
            'Länge_mm': float(part.length or 0),
            'Breite_mm': float(part.width or 0),
            'Höhe_mm': float(part.height or 0),
            'Volumen_m3': float(volume_m3 or 0),
            'Gewicht_kg': float((volume_m3 or 0) * density_kg_m3),
        }
        # Alle BVX-Attribute zusätzlich verfügbar machen, damit sie z. B.
        # für Decke/Bauabschnitt oder spätere Kopf-/Sortierfelder ausgewählt werden können.
        for attr_key, attr_value in (part.attributes or {}).items():
            col = f'BVX_{attr_key}'
            if col not in row:
                row[col] = attr_value
        rows.append(row)
    return pd.DataFrame(rows)



def read_parts_excel_to_dataframe(uploaded_excel, density_kg_m3: float = 500.0) -> Tuple[pd.DataFrame, List[str]]:
    """Liest eine Bauteile-Excel als Ersatz für BVX ein.

    Erwartet bevorzugt ein Blatt "Bauteile". Falls es nicht existiert, wird das
    erste Blatt gelesen. Mindestfelder sind Länge_mm, Breite_mm und Höhe_mm.
    Weitere Attribute werden unverändert für Sortierung/Kopfbereiche übernommen.
    """
    messages: List[str] = []
    if uploaded_excel is None:
        return pd.DataFrame(), ['Keine Bauteile-Excel geladen.']

    try:
        xls = pd.ExcelFile(uploaded_excel)
        sheet_name = 'Bauteile' if 'Bauteile' in xls.sheet_names else xls.sheet_names[0]
        df = normalize_columns(pd.read_excel(xls, sheet_name=sheet_name))
    except Exception as exc:
        return pd.DataFrame(), [f'Bauteile-Excel konnte nicht gelesen werden: {exc}']

    if df.empty:
        return pd.DataFrame(), ['Bauteile-Excel enthält keine Zeilen.']

    # Häufige alternative Spaltennamen normalisieren.
    alias_map = {
        'Länge': 'Länge_mm', 'Laenge': 'Länge_mm', 'Laenge_mm': 'Länge_mm', 'Length': 'Länge_mm', 'X': 'Länge_mm', 'DimensionX': 'Länge_mm',
        'Breite': 'Breite_mm', 'Width': 'Breite_mm', 'Y': 'Breite_mm', 'DimensionY': 'Breite_mm',
        'Höhe': 'Höhe_mm', 'Hoehe': 'Höhe_mm', 'Hoehe_mm': 'Höhe_mm', 'Height': 'Höhe_mm', 'Z': 'Höhe_mm', 'DimensionZ': 'Höhe_mm',
        'Bauteilnr': 'Bauteilnummer', 'Bauteil_Nr': 'Bauteilnummer', 'Bauteil-Nr': 'Bauteilnummer', 'PartNo': 'Bauteilnummer', 'Part_No': 'Bauteilnummer',
        'Paket': 'Pak/Unit', 'Pak': 'Pak/Unit', 'Unit': 'Pak/Unit',
        'Profile': 'Profil', 'Surface': 'Oberfläche', 'Oberflaeche': 'Oberfläche', 'Grade': 'Qualität', 'Qualitaet': 'Qualität',
        'Volumen': 'Volumen_m3', 'Volume_m3': 'Volumen_m3', 'Gewicht': 'Gewicht_kg', 'Weight_kg': 'Gewicht_kg',
    }
    df = df.rename(columns={c: alias_map.get(str(c).strip(), c) for c in df.columns})

    required_order = [
        'Index', 'Name', 'Bauteilnummer', 'PartId', 'Pak/Unit', 'Profil', 'Oberfläche', 'Qualität',
        'User_Attribut_2', 'Länge_mm', 'Breite_mm', 'Höhe_mm', 'Volumen_m3', 'Gewicht_kg'
    ]
    for col in required_order:
        if col not in df.columns:
            df[col] = '' if col not in {'Index', 'Länge_mm', 'Breite_mm', 'Höhe_mm', 'Volumen_m3', 'Gewicht_kg'} else 0

    df = df.dropna(how='all').copy().reset_index(drop=True)
    if df.empty:
        return pd.DataFrame(), ['Bauteile-Excel enthält keine verwertbaren Zeilen.']

    if (pd.to_numeric(df['Index'], errors='coerce').fillna(0) == 0).all():
        df['Index'] = range(1, len(df) + 1)

    for col in ['Länge_mm', 'Breite_mm', 'Höhe_mm', 'Volumen_m3', 'Gewicht_kg']:
        df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', '.', regex=False), errors='coerce').fillna(0.0)

    missing_dims = df[(df['Länge_mm'] <= 0) | (df['Breite_mm'] <= 0) | (df['Höhe_mm'] <= 0)]
    if not missing_dims.empty:
        messages.append(f'{len(missing_dims)} Zeile(n) haben fehlende Länge/Breite/Höhe und werden trotzdem angezeigt, können aber evtl. nicht verladen werden.')

    # Volumen/Gewicht bei Bedarf berechnen.
    calc_volume = (df['Länge_mm'] * df['Breite_mm'] * df['Höhe_mm']) / 1_000_000_000
    df.loc[df['Volumen_m3'] <= 0, 'Volumen_m3'] = calc_volume[df['Volumen_m3'] <= 0]
    df.loc[df['Gewicht_kg'] <= 0, 'Gewicht_kg'] = df.loc[df['Gewicht_kg'] <= 0, 'Volumen_m3'] * float(density_kg_m3)

    # Pflichttexte auffüllen.
    df['Name'] = df['Name'].astype(str).replace({'nan': ''})
    df['Bauteilnummer'] = df['Bauteilnummer'].astype(str).replace({'nan': ''})
    for idx in df.index:
        if not str(df.at[idx, 'Bauteilnummer']).strip():
            fallback = str(df.at[idx, 'Name']).strip() or str(int(safe_number(df.at[idx, 'Index'], idx + 1)))
            df.at[idx, 'Bauteilnummer'] = fallback
        if not str(df.at[idx, 'Name']).strip():
            df.at[idx, 'Name'] = str(df.at[idx, 'Bauteilnummer']).strip()

    # Gewünschte Standardspalten zuerst, Zusatzattribute danach.
    extra_cols = [c for c in df.columns if c not in required_order]
    df = df[required_order + extra_cols]
    messages.append('Bauteile wurden aus Excel geladen.')
    return df, messages

def _available_bvx_meta_fields(parts_df: pd.DataFrame) -> List[str]:
    """Felder, die sinnvoll für Kopf-/Projektdaten auswählbar sind."""
    if parts_df is None or parts_df.empty:
        return []
    fields: List[str] = []
    for col in parts_df.columns:
        if col in {'Index', 'Länge_mm', 'Breite_mm', 'Höhe_mm', 'Volumen_m3', 'Gewicht_kg'}:
            continue
        try:
            values = parts_df[col].dropna().astype(str).str.strip()
            values = values[(values != '') & (values.str.lower() != 'nan')]
            if len(values) > 0:
                fields.append(col)
        except Exception:
            pass
    return fields


def _default_attr_index(options: List[str], preferred: List[str]) -> int:
    lower_map = {str(opt).lower(): i for i, opt in enumerate(options)}
    for pref in preferred:
        idx = lower_map.get(pref.lower())
        if idx is not None:
            return idx
    for pref in preferred:
        for i, opt in enumerate(options):
            if pref.lower() in str(opt).lower():
                return i
    return 0


def _unique_display_values_from_field(parts_df: pd.DataFrame, field: str, max_values: int = 6) -> str:
    """Nimmt die eindeutigen Werte eines BVX-Feldes für Kopfbereiche."""
    if parts_df is None or parts_df.empty or not field or field == 'Manuell' or field not in parts_df.columns:
        return ''
    values: List[str] = []
    for value in parts_df[field].tolist():
        text = _format_label_value(value)
        if text and text not in values:
            values.append(text)
        if len(values) >= max_values:
            break
    return ', '.join(values)


def sort_parts_dataframe(
    df: pd.DataFrame,
    main_attr: str,
    main_asc: bool,
    second_attr: str,
    second_asc: bool,
) -> pd.DataFrame:
    sort_cols = []
    sort_ascending = []

    if main_attr and main_attr in df.columns:
        sort_cols.append(main_attr)
        sort_ascending.append(main_asc)
    if second_attr and second_attr != 'Keine' and second_attr in df.columns and second_attr not in sort_cols:
        sort_cols.append(second_attr)
        sort_ascending.append(second_asc)

    if not sort_cols:
        return df.copy()

    result = df.copy()
    for col in sort_cols:
        # leere Werte sauber ans Ende bringen
        if result[col].dtype == object:
            result[col] = result[col].fillna('').astype(str)
    return result.sort_values(by=sort_cols, ascending=sort_ascending, kind='stable').reset_index(drop=True)


# =============================================================================
# Verladeplanung
# =============================================================================

def yes_no_to_bool(value: Any) -> bool:
    """Wandelt JA/NEIN, True/False, 1/0 robust in Bool."""
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)) and not pd.isna(value):
        return value != 0
    text = str(value).strip().lower()
    return text in {'ja', 'j', 'yes', 'y', 'true', 'wahr', '1', 'x'}


def safe_number(value: Any, default: float = 0.0) -> float:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return default
    try:
        return float(str(value).replace(',', '.'))
    except (TypeError, ValueError):
        return default


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Entfernt unsichtbare Leerzeichen aus Excel-Spaltennamen."""
    result = df.copy()
    result.columns = [str(col).strip().replace('\n', ' ') for col in result.columns]
    return result


def get_transport_presets() -> Dict[str, List[Dict[str, Any]]]:
    """Fallback-Stammdaten, falls keine Excel-Vorlage geladen wird."""
    return {
        'LKW solo': [
            {'Freigabe': True, 'Pritsche': 'LKW', 'Pritschenname': 'LKW', 'Fuhrenoption': 'LKW solo', 'Pritschen_Reihenfolge': 1,
             'Länge_mm': 9300, 'Breite_mm': 2450, 'Max_Höhe_mm': 2600, 'Überhang_vorne_mm': 0,
             'Überhang_hinten_mm': 300, 'Max_Gewicht_kg': 15000, 'Eigengewicht_Pritsche_kg': 0, 'Kantholz_erste_Lage_mm': 80,
             'Einlage_zwischen_Lagen_mm': 40, 'Einlage_allgemein_mm': 0, 'Drehen_90_erlaubt': False},
        ],
        'LKW mit Anhänger': [
            {'Freigabe': True, 'Pritsche': 'LKW', 'Pritschenname': 'LKW', 'Fuhrenoption': 'LKW mit Anhänger', 'Pritschen_Reihenfolge': 1,
             'Länge_mm': 9300, 'Breite_mm': 2450, 'Max_Höhe_mm': 2600, 'Überhang_vorne_mm': 0,
             'Überhang_hinten_mm': 300, 'Max_Gewicht_kg': 15000, 'Eigengewicht_Pritsche_kg': 0, 'Kantholz_erste_Lage_mm': 80,
             'Einlage_zwischen_Lagen_mm': 40, 'Einlage_allgemein_mm': 0, 'Drehen_90_erlaubt': False},
            {'Freigabe': True, 'Pritsche': 'Anhänger', 'Pritschenname': 'Anhänger', 'Fuhrenoption': 'LKW mit Anhänger', 'Pritschen_Reihenfolge': 2,
             'Länge_mm': 8200, 'Breite_mm': 2450, 'Max_Höhe_mm': 2600, 'Überhang_vorne_mm': 0,
             'Überhang_hinten_mm': 300, 'Max_Gewicht_kg': 12000, 'Eigengewicht_Pritsche_kg': 0, 'Kantholz_erste_Lage_mm': 80,
             'Einlage_zwischen_Lagen_mm': 40, 'Einlage_allgemein_mm': 0, 'Drehen_90_erlaubt': False},
        ],
        'Anhängerzug': [
            {'Freigabe': True, 'Pritsche': 'LKW', 'Pritschenname': 'LKW', 'Fuhrenoption': 'Anhängerzug', 'Pritschen_Reihenfolge': 1,
             'Länge_mm': 7600, 'Breite_mm': 2450, 'Max_Höhe_mm': 2600, 'Überhang_vorne_mm': 0,
             'Überhang_hinten_mm': 300, 'Max_Gewicht_kg': 12000, 'Eigengewicht_Pritsche_kg': 0, 'Kantholz_erste_Lage_mm': 80,
             'Einlage_zwischen_Lagen_mm': 40, 'Einlage_allgemein_mm': 0, 'Drehen_90_erlaubt': False},
            {'Freigabe': True, 'Pritsche': 'Anhänger', 'Pritschenname': 'Anhänger', 'Fuhrenoption': 'Anhängerzug', 'Pritschen_Reihenfolge': 2,
             'Länge_mm': 8200, 'Breite_mm': 2450, 'Max_Höhe_mm': 2600, 'Überhang_vorne_mm': 0,
             'Überhang_hinten_mm': 300, 'Max_Gewicht_kg': 12000, 'Eigengewicht_Pritsche_kg': 0, 'Kantholz_erste_Lage_mm': 80,
             'Einlage_zwischen_Lagen_mm': 40, 'Einlage_allgemein_mm': 0, 'Drehen_90_erlaubt': False},
        ],
        'Tiefbettauflieger': [
            {'Freigabe': True, 'Pritsche': 'Tiefbett', 'Pritschenname': 'Tiefbett', 'Fuhrenoption': 'Tiefbettauflieger', 'Pritschen_Reihenfolge': 1,
             'Länge_mm': 13000, 'Breite_mm': 2550, 'Max_Höhe_mm': 3200, 'Überhang_vorne_mm': 500,
             'Überhang_hinten_mm': 1000, 'Max_Gewicht_kg': 24000, 'Eigengewicht_Pritsche_kg': 0, 'Kantholz_erste_Lage_mm': 100,
             'Einlage_zwischen_Lagen_mm': 40, 'Einlage_allgemein_mm': 0, 'Drehen_90_erlaubt': True},
        ],
    }


def read_transport_config_excel(uploaded_excel) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, Any], List[str]]:
    """Liest die Excel-Stammdaten für Fuhrenoptionen, Pritschen und Standards."""
    messages: List[str] = []
    standards: Dict[str, Any] = {
        'Holzdichte': 500.0,
        'Max_Bundgewicht': 1000.0,
        'Standard_Kantholz_erste_Lage': 80.0,
        'Standard_Einlage_zwischen_Lagen': 40.0,
        'Standard_Einlage_allgemein': 0.0,
        'Längenversatz_je_Lage': 100.0,
        'Abstand_zwischen_Einheiten': 0.0,
        'Sichtseite_nach_unten': 'JA',
        'Plane_Folie': 'JA',
    }

    if uploaded_excel is None:
        options = pd.DataFrame([
            {'Freigegeben': True, 'Priorität': 1, 'Fuhrenoption': 'LKW solo', 'Wiederholen_bis_alles_verladen': True, 'Strategie': 'Variante A'},
        ])
        preset = get_transport_presets()['LKW solo']
        pritschen = pd.DataFrame(preset)
        return options, pritschen, standards, ['Keine Excel-Stammdaten geladen. Fallback LKW solo wird verwendet.']

    try:
        xls = pd.ExcelFile(uploaded_excel)
        options_raw = normalize_columns(pd.read_excel(xls, sheet_name='Fuhrenoptionen'))
        pritschen_raw = normalize_columns(pd.read_excel(xls, sheet_name='Pritschen'))
        standards_raw = normalize_columns(pd.read_excel(xls, sheet_name='Standards'))
    except Exception as exc:
        preset = get_transport_presets()['LKW solo']
        options = pd.DataFrame([
            {'Freigegeben': True, 'Priorität': 1, 'Fuhrenoption': 'LKW solo', 'Wiederholen_bis_alles_verladen': True, 'Strategie': 'Variante A'},
        ])
        pritschen = pd.DataFrame(preset)
        return options, pritschen, standards, [f'Excel konnte nicht gelesen werden: {exc}. Fallback LKW solo wird verwendet.']

    if {'Parameter', 'Wert'}.issubset(set(standards_raw.columns)):
        for _, row in standards_raw.dropna(how='all').iterrows():
            key = str(row.get('Parameter', '')).strip()
            if not key:
                continue
            value = row.get('Wert')
            if key in ['Sichtseite_nach_unten', 'Plane_Folie']:
                standards[key] = value
            else:
                standards[key] = safe_number(value, standards.get(key, 0.0))

    options = options_raw.dropna(how='all').copy()
    needed_options = ['Freigegeben', 'Priorität', 'Fuhrenoption']
    for col in needed_options:
        if col not in options.columns:
            options[col] = ''
    options = options[options['Fuhrenoption'].notna() & (options['Fuhrenoption'].astype(str).str.strip() != '')].copy()
    options['Freigegeben'] = options['Freigegeben'].apply(yes_no_to_bool)
    options['Priorität'] = options['Priorität'].apply(lambda v: int(safe_number(v, 999)))
    if 'Wiederholen_bis_alles_verladen' not in options.columns:
        options['Wiederholen_bis_alles_verladen'] = True
    options['Wiederholen_bis_alles_verladen'] = options['Wiederholen_bis_alles_verladen'].apply(yes_no_to_bool)
    if 'Strategie' not in options.columns:
        options['Strategie'] = 'Variante A'

    p = pritschen_raw.dropna(how='all').copy()
    rename_map = {
        'Pritschenname': 'Pritschenname',
        'Max_Ladehöhe_mm': 'Max_Höhe_mm',
        'Max_Höhe_mm': 'Max_Höhe_mm',
        'Aktiv': 'Freigabe',
        'Drehen_90_erlaubt': 'Drehen_90_erlaubt',
    }
    p = p.rename(columns={k: v for k, v in rename_map.items() if k in p.columns})

    required_cols = [
        'Fuhrenoption', 'Pritschen_Reihenfolge', 'Pritschenname', 'Freigabe', 'Länge_mm', 'Breite_mm',
        'Max_Höhe_mm', 'Max_Gewicht_kg', 'Eigengewicht_Pritsche_kg', 'Überhang_vorne_mm', 'Überhang_hinten_mm',
        'Kantholz_erste_Lage_mm', 'Einlage_zwischen_Lagen_mm', 'Einlage_allgemein_mm', 'Drehen_90_erlaubt'
    ]
    for col in required_cols:
        if col not in p.columns:
            p[col] = ''

    p = p[p['Fuhrenoption'].notna() & (p['Fuhrenoption'].astype(str).str.strip() != '')].copy()
    p['Freigabe'] = p['Freigabe'].apply(yes_no_to_bool)
    p['Drehen_90_erlaubt'] = p['Drehen_90_erlaubt'].apply(yes_no_to_bool)
    p['Pritschen_Reihenfolge'] = p['Pritschen_Reihenfolge'].apply(lambda v: int(safe_number(v, 999)))
    for col in ['Länge_mm', 'Breite_mm', 'Max_Höhe_mm', 'Max_Gewicht_kg', 'Eigengewicht_Pritsche_kg', 'Überhang_vorne_mm', 'Überhang_hinten_mm', 'Kantholz_erste_Lage_mm', 'Einlage_zwischen_Lagen_mm', 'Einlage_allgemein_mm']:
        p[col] = p[col].apply(lambda v: safe_number(v, 0.0))
    p['Pritsche'] = p['Pritschenname'].astype(str)

    if options.empty:
        messages.append('Keine gültigen Fuhrenoptionen in der Excel gefunden.')
    if p.empty:
        messages.append('Keine gültigen Pritschen in der Excel gefunden.')
    return options, p, standards, messages


def make_bundle_signature(row: pd.Series, same_height: bool, same_width: bool, same_quality: bool, same_profile: bool) -> Tuple[Any, ...]:
    signature: List[Any] = []
    if same_height:
        signature.append(row.get('Höhe_mm'))
    if same_width:
        signature.append(row.get('Breite_mm'))
    if same_quality:
        signature.append(row.get('Qualität'))
    if same_profile:
        signature.append(row.get('Profil'))
    return tuple(signature)


def _format_label_value(value: Any) -> str:
    """Formatiert Werte für die Anzeige in Ansichten und Pritschenplan."""
    if value is None:
        return ''
    try:
        if pd.isna(value):
            return ''
    except Exception:
        pass
    if isinstance(value, (int, float)):
        val = float(value)
        if abs(val - round(val)) < 0.001:
            return str(int(round(val)))
        return f'{val:.2f}'.rstrip('0').rstrip('.')
    text = str(value).strip()
    if text.lower() == 'nan':
        return ''
    return text


def _part_display_label(row: pd.Series, label_attr: str) -> str:
    """Liefert die separat gewählte Beschriftung für Plan, Ansicht und BSD."""
    attr = str(label_attr or '').strip()
    if attr in ['Bauteilnummer + Pak/Unit', 'Bauteilnummer + Pak']:
        a = _format_label_value(row.get('Bauteilnummer'))
        b = _format_label_value(row.get('Pak/Unit', row.get('Pak', row.get('Unit'))))
        label = ' / '.join([v for v in [a, b] if v])
    elif attr == 'Bauteilnummer + Name':
        a = _format_label_value(row.get('Bauteilnummer'))
        b = _format_label_value(row.get('Name'))
        label = ' / '.join([v for v in [a, b] if v])
    elif attr == 'Name + Pak/Unit':
        a = _format_label_value(row.get('Name'))
        b = _format_label_value(row.get('Pak/Unit', row.get('Pak', row.get('Unit'))))
        label = ' / '.join([v for v in [a, b] if v])
    else:
        label = _format_label_value(row.get(attr)) if attr else ''

    if not label:
        label = _format_label_value(row.get('Bauteilnummer'))
    if not label:
        label = _format_label_value(row.get('Name'))
    return label or 'Bauteil'


def _unique_options(values: List[str]) -> List[str]:
    seen = set()
    out: List[str] = []
    for value in values:
        value = str(value or '').strip()
        if not value or value in seen:
            continue
        seen.add(value)
        out.append(value)
    return out



def build_loading_units(
    sorted_parts: pd.DataFrame,
    max_bundle_weight: float,
    use_bundles: bool,
    bundle_spacer_height: float,
    general_spacer_height: float,
    same_height: bool,
    same_width: bool,
    same_quality: bool,
    same_profile: bool,
    label_attr: str = 'Bauteilnummer',
) -> pd.DataFrame:
    """Erzeugt Verladeeinheiten: einzelnes Bauteil oder Bund.

    V55:
    - Wenn ein gleichartiger Bund am Gewichtsende "unschön" abbrechen würde,
      wird die letzte Bundgruppe lokal neu aufgeteilt.
    - Ziel: weniger kleine Restbunde und weniger Löcher auf der Pritsche.
    """
    if sorted_parts.empty:
        return pd.DataFrame()

    units: List[Dict[str, Any]] = []

    def _rows_to_unit(bundle_rows: List[pd.Series]) -> Dict[str, Any]:
        count = len(bundle_rows)
        length = max(float(r['Länge_mm']) for r in bundle_rows)
        width = max(float(r['Breite_mm']) for r in bundle_rows)
        internal_spacer = general_spacer_height if general_spacer_height > 0 else 0.0
        height = sum(float(r['Höhe_mm']) for r in bundle_rows) + max(0, count - 1) * internal_spacer
        volume = sum(float(r['Volumen_m3']) for r in bundle_rows)
        weight = sum(float(r['Gewicht_kg']) for r in bundle_rows)
        part_labels = [str(r.get('Bauteilnummer') or r.get('Name')) for r in bundle_rows]
        view_labels = [_part_display_label(r, label_attr) for r in bundle_rows]
        view_list = '|'.join(view_labels)
        if len(view_labels) <= 3:
            view_label = ', '.join(view_labels)
        else:
            view_label = ', '.join(view_labels[:3]) + ' ...'
        part_lengths = [str(float(r['Länge_mm'])) for r in bundle_rows]
        part_widths = [str(float(r['Breite_mm'])) for r in bundle_rows]
        part_heights = [str(float(r['Höhe_mm'])) for r in bundle_rows]
        return {
            'Einheit_ID': f'B{len(units) + 1:03d}',
            'Typ': 'Bund' if count > 1 else 'Bauteil',
            'Anzahl_Bauteile': count,
            'Bauteile': ', '.join(part_labels),
            'Bauteile_Liste': '|'.join(part_labels),
            'Ansicht_Attribut': label_attr,
            'Ansicht_Label': view_label,
            'Ansicht_Liste': view_list,
            'Einzellängen_mm': '|'.join(part_lengths),
            'Einzelbreiten_mm': '|'.join(part_widths),
            'Einzelhöhen_mm': '|'.join(part_heights),
            'Einlage_allgemein_mm': float(general_spacer_height),
            'Bundeinlage_mm': float(bundle_spacer_height),
            'Länge_mm': length,
            'Breite_mm': width,
            'Höhe_mm': height,
            'Volumen_m3': volume,
            'Gewicht_kg': weight,
            'Warnung': 'über max. Bundgewicht' if weight > max_bundle_weight else '',
        }

    def _flush_rows(bundle_rows: List[pd.Series]) -> None:
        if not bundle_rows:
            return
        units.append(_rows_to_unit(bundle_rows))

    def _best_local_split(rows: List[pd.Series]) -> Optional[Tuple[List[pd.Series], List[pd.Series]]]:
        if len(rows) < 2:
            return None
        weights = [float(r['Gewicht_kg']) for r in rows]
        best = None
        total = len(rows)
        for cut in range(1, total):
            left = rows[:cut]
            right = rows[cut:]
            wl = sum(float(r['Gewicht_kg']) for r in left)
            wr = sum(float(r['Gewicht_kg']) for r in right)
            if wl > max_bundle_weight or wr > max_bundle_weight:
                continue
            # Möglichst ähnlich viele Bauteile und ähnliche Gewichte.
            score = abs(len(left) - len(right)) * 100000 + abs(wl - wr)
            # Kleine Restbunde vermeiden.
            min_count = min(len(left), len(right))
            if min_count == 1:
                score += 50000
            if best is None or score < best[0]:
                best = (score, left, right)
        return (best[1], best[2]) if best is not None else None

    if not use_bundles:
        for idx, row in sorted_parts.iterrows():
            default_label = str(row.get('Bauteilnummer') or row.get('Name'))
            view_label = _part_display_label(row, label_attr)
            units.append({
                'Einheit_ID': f'E{idx + 1:03d}',
                'Typ': 'Bauteil',
                'Anzahl_Bauteile': 1,
                'Bauteile': default_label,
                'Bauteile_Liste': default_label,
                'Ansicht_Attribut': label_attr,
                'Ansicht_Label': view_label,
                'Ansicht_Liste': view_label,
                'Einzellängen_mm': str(row['Länge_mm']),
                'Einzelbreiten_mm': str(row['Breite_mm']),
                'Einzelhöhen_mm': str(row['Höhe_mm']),
                'Einlage_allgemein_mm': float(general_spacer_height),
                'Bundeinlage_mm': float(bundle_spacer_height),
                'Länge_mm': row['Länge_mm'],
                'Breite_mm': row['Breite_mm'],
                'Höhe_mm': row['Höhe_mm'],
                'Volumen_m3': row['Volumen_m3'],
                'Gewicht_kg': row['Gewicht_kg'],
                'Warnung': 'über max. Bundgewicht' if row['Gewicht_kg'] > max_bundle_weight else '',
            })
        return pd.DataFrame(units)

    current_rows: List[pd.Series] = []
    current_weight = 0.0
    current_signature: Optional[Tuple[Any, ...]] = None

    for _, row in sorted_parts.iterrows():
        row_weight = float(row['Gewicht_kg'])
        row_signature = make_bundle_signature(row, same_height, same_width, same_quality, same_profile)
        signature_break = current_signature is not None and row_signature != current_signature
        weight_break = current_rows and (current_weight + row_weight > max_bundle_weight)

        if signature_break:
            _flush_rows(current_rows)
            current_rows = []
            current_weight = 0.0
            current_signature = None

        elif weight_break:
            # Gleichartige Gruppe lokal besser aufteilen, statt einen sehr kleinen Restbund zu erzeugen.
            candidate_rows = list(current_rows) + [row]
            split = _best_local_split(candidate_rows)
            if split is not None:
                left_rows, right_rows = split
                _flush_rows(left_rows)
                current_rows = list(right_rows)
                current_weight = sum(float(r['Gewicht_kg']) for r in current_rows)
                current_signature = row_signature if current_rows else None
                continue
            else:
                _flush_rows(current_rows)
                current_rows = []
                current_weight = 0.0
                current_signature = None

        current_rows.append(row)
        current_weight += row_weight
        current_signature = row_signature

    _flush_rows(current_rows)
    return pd.DataFrame(units)


def init_platform_state(row: pd.Series, base_wood_height: float, layer_spacer_height: float, gap_length: float) -> Dict[str, Any]:
    base_height = safe_number(row.get('Kantholz_erste_Lage_mm'), base_wood_height)
    layer_height = safe_number(row.get('Einlage_zwischen_Lagen_mm'), layer_spacer_height)
    return {
        'Fuhre_Nr': int(row.get('Fuhre_Nr', 1)) if not pd.isna(row.get('Fuhre_Nr', 1)) else 1,
        'Fuhrenoption': str(row.get('Fuhrenoption', '')),
        'Pritschenname': str(row.get('Pritschenname', row.get('Pritsche', 'Pritsche'))),
        'Pritsche': str(row['Pritsche']),
        'Länge_mm': float(row['Länge_mm']),
        'Breite_mm': float(row['Breite_mm']),
        'Max_Höhe_mm': float(row['Max_Höhe_mm']),
        'Überhang_vorne_mm': float(row['Überhang_vorne_mm']),
        'Überhang_hinten_mm': float(row['Überhang_hinten_mm']),
        'Max_Gewicht_kg': float(row['Max_Gewicht_kg']),
        'Eigengewicht_Pritsche_kg': float(row.get('Eigengewicht_Pritsche_kg', 0.0) or 0.0),
        'Eff_Länge_mm': float(row['Länge_mm']) + float(row['Überhang_vorne_mm']) + float(row['Überhang_hinten_mm']),
        'base_wood_height': float(base_height),
        'layer_spacer_height': float(layer_height),
        'general_spacer_height': float(row.get('Einlage_allgemein_mm', 0.0) or 0.0),
        'gap_length': float(gap_length),
        'allow_rotation_platform': yes_no_to_bool(row.get('Drehen_90_erlaubt', False)),
        'current_x': 0.0,
        'current_y': 0.0,
        'current_z': float(base_height),
        'row_max_width': 0.0,
        'layer_max_height': 0.0,
        'used_length': 0.0,
        'used_width': 0.0,
        'used_height': float(base_height),
        'total_weight': 0.0,
        'current_layer_has_bundle': False,
        'placements': [],
    }


def can_place(state: Dict[str, Any], x: float, y: float, z: float, length: float, width: float, height: float, weight: float) -> bool:
    if state['total_weight'] + weight + state.get('Eigengewicht_Pritsche_kg', 0.0) > state['Max_Gewicht_kg']:
        return False
    if x + length > state['Eff_Länge_mm']:
        return False
    if y + width > state['Breite_mm']:
        return False
    if z + height > state['Max_Höhe_mm']:
        return False
    return True


def commit_place(
    state: Dict[str, Any],
    unit: pd.Series,
    x: float,
    y: float,
    z: float,
    length: float,
    width: float,
    height: float,
    rotation: int,
    mode: str,
) -> Dict[str, Any]:
    placement = {
        'Fuhre_Nr': state['Fuhre_Nr'],
        'Fuhrenoption': state['Fuhrenoption'],
        'Pritschenname': state['Pritschenname'],
        'Pritsche': state['Pritsche'],
        'Einheit_ID': unit['Einheit_ID'],
        'Typ': unit['Typ'],
        'Anzahl_Bauteile': int(unit['Anzahl_Bauteile']),
        'Bauteile': unit['Bauteile'],
        'Bauteile_Liste': unit.get('Bauteile_Liste', unit.get('Bauteile', '')),
        'Ansicht_Attribut': unit.get('Ansicht_Attribut', ''),
        'Ansicht_Label': unit.get('Ansicht_Label', unit.get('Einheit_ID', '')),
        'Ansicht_Liste': unit.get('Ansicht_Liste', unit.get('Ansicht_Label', '')),
        'Einzellängen_mm': unit.get('Einzellängen_mm', ''),
        'Einzelbreiten_mm': unit.get('Einzelbreiten_mm', ''),
        'Einzelhöhen_mm': unit.get('Einzelhöhen_mm', ''),
        'Einlage_allgemein_mm': safe_number(unit.get('Einlage_allgemein_mm'), 0.0),
        'Bundeinlage_mm': safe_number(unit.get('Bundeinlage_mm'), 0.0),
        'X_mm': round(x, 1),
        'Y_mm': round(y, 1),
        'Z_mm': round(z, 1),
        'Länge_mm': round(length, 1),
        'Breite_mm': round(width, 1),
        'Höhe_mm': round(height, 1),
        'Drehung': rotation,
        'Ebene': mode,
        'Gewicht_kg': round(float(unit['Gewicht_kg']), 2),
    }
    state['placements'].append(placement)

    state['current_x'] = x + length + state['gap_length']
    state['current_y'] = y
    state['row_max_width'] = max(state['row_max_width'], width)
    state['layer_max_height'] = max(state['layer_max_height'], height)
    state['used_length'] = max(state['used_length'], x + length)
    state['used_width'] = max(state['used_width'], y + width)
    state['used_height'] = max(state['used_height'], z + height)
    state['total_weight'] += float(unit['Gewicht_kg'])
    state['current_layer_has_bundle'] = bool(state.get('current_layer_has_bundle', False) or str(unit.get('Typ', '')).strip() == 'Bund')
    return placement


def _unit_orientations_for_state(state: Dict[str, Any], unit: pd.Series, allow_rotation: bool) -> List[Tuple[float, float, int]]:
    """Mögliche Drehungen einer Verladeeinheit."""
    length = float(unit['Länge_mm'])
    width = float(unit['Breite_mm'])
    rotation_allowed = bool(allow_rotation and state.get('allow_rotation_platform', False))
    orientations = [(length, width, 0)]
    if rotation_allowed and abs(length - width) > 0.001:
        orientations.append((width, length, 90))
    return orientations


def _current_gap_fit(state: Dict[str, Any], unit: pd.Series, allow_rotation: bool) -> Optional[Tuple[float, float, float, float, float, float, int, str]]:
    """Prüft nur die aktuelle offene Stelle in der aktuellen Reihe/Lage.

    Diese Funktion bewegt nicht in eine neue Reihe oder neue Lage. Sie wird genutzt,
    um kleinere spätere Bunde/Bauteile in vorhandene Löcher zu setzen, bevor ein
    Unterbau erzeugt wird.
    """
    height = float(unit['Höhe_mm'])
    weight = float(unit['Gewicht_kg'])
    x = float(state['current_x'])
    y = float(state['current_y'])
    z = float(state['current_z'])
    for use_length, use_width, rotation in _unit_orientations_for_state(state, unit, allow_rotation):
        if can_place(state, x, y, z, use_length, use_width, height, weight):
            return (x, y, z, use_length, use_width, height, rotation, 'Lücke gefüllt')
    return None


def _gap_candidate_score(state: Dict[str, Any], fit: Tuple[float, float, float, float, float, float, int, str]) -> float:
    """Bewertet, wie gut ein Kandidat die offene Stelle füllt."""
    _, _, _, length, width, height, _, _ = fit
    rest_l = max(1.0, float(state['Eff_Länge_mm']) - float(state['current_x']))
    rest_w = max(1.0, float(state['Breite_mm']) - float(state['current_y']))
    fill_l = min(1.0, length / rest_l)
    fill_w = min(1.0, width / rest_w)
    area = length * width
    return fill_l * 1000000.0 + fill_w * 10000.0 + area / 1000.0 + height


def _find_later_gap_candidate(pending: List[pd.Series], state: Dict[str, Any], allow_rotation: bool, max_search: int = 40) -> Optional[Tuple[int, Tuple[float, float, float, float, float, float, int, str]]]:
    """Sucht in den nächsten Verladeeinheiten nach einem Teil, das die offene Lücke füllt."""
    if len(pending) <= 1:
        return None
    # Nur suchen, wenn überhaupt schon eine Reihe/Lage begonnen wurde.
    if float(state.get('current_x', 0.0)) <= 0 and float(state.get('current_y', 0.0)) <= 0 and float(state.get('used_height', 0.0)) <= float(state.get('base_wood_height', 0.0)) + 0.1:
        return None
    best: Optional[Tuple[int, Tuple[float, float, float, float, float, float, int, str], float]] = None
    limit = min(len(pending), max_search + 1)
    for idx in range(1, limit):
        fit = _current_gap_fit(state, pending[idx], allow_rotation)
        if fit is None:
            continue
        score = _gap_candidate_score(state, fit)
        if best is None or score > best[2]:
            best = (idx, fit, score)
    if best is None:
        return None
    return best[0], best[1]


def try_place_unit(
    state: Dict[str, Any],
    unit: pd.Series,
    allow_beside: bool,
    allow_stack: bool,
    allow_rotation: bool,
) -> Optional[Dict[str, Any]]:
    height = float(unit['Höhe_mm'])
    weight = float(unit['Gewicht_kg'])

    for use_length, use_width, rotation in _unit_orientations_for_state(state, unit, allow_rotation):
        # 1. hintereinander in aktueller Reihe
        x = state['current_x']
        y = state['current_y']
        z = state['current_z']
        if can_place(state, x, y, z, use_length, use_width, height, weight):
            return commit_place(state, unit, x, y, z, use_length, use_width, height, rotation, 'hintereinander')

        # 2. neue Reihe daneben
        if allow_beside and state['row_max_width'] > 0:
            x = 0.0
            y = state['current_y'] + state['row_max_width']
            z = state['current_z']
            if can_place(state, x, y, z, use_length, use_width, height, weight):
                state['current_x'] = 0.0
                state['current_y'] = y
                state['row_max_width'] = 0.0
                return commit_place(state, unit, x, y, z, use_length, use_width, height, rotation, 'nebeneinander')

        # 3. neue Lage darüber
        if allow_stack and state['layer_max_height'] > 0:
            x = 0.0
            y = 0.0
            # Bundeinlage/Lagenholz nur zwischen separaten Bunden/Lagen verwenden.
            # Innerhalb eines Bundes kommt diese Einlage nicht zum Einsatz.
            # Wenn keine Bundlage betroffen ist, kann optional die allgemeine Einlage
            # zwischen einzelnen Bauteilen/Lagen gerechnet werden.
            current_layer_has_bundle = bool(state.get('current_layer_has_bundle', False))
            next_is_bundle = str(unit.get('Typ', '')).strip() == 'Bund'
            if current_layer_has_bundle or next_is_bundle:
                effective_layer_spacer = float(state.get('layer_spacer_height', 0.0))
            else:
                effective_layer_spacer = float(state.get('general_spacer_height', 0.0))
            z = state['current_z'] + state['layer_max_height'] + max(0.0, effective_layer_spacer)
            if can_place(state, x, y, z, use_length, use_width, height, weight):
                state['current_x'] = 0.0
                state['current_y'] = 0.0
                state['current_z'] = z
                state['row_max_width'] = 0.0
                state['layer_max_height'] = 0.0
                state['current_layer_has_bundle'] = False
                return commit_place(state, unit, x, y, z, use_length, use_width, height, rotation, 'übereinander')

    return None


def create_loading_plan(
    units: pd.DataFrame,
    platforms: pd.DataFrame,
    base_wood_height: float,
    layer_spacer_height: float,
    gap_length: float,
    allow_beside: bool,
    allow_stack: bool,
    allow_rotation: bool,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Greedy-Verladevorschlag für eine einzelne Fuhre."""
    if units.empty or platforms.empty:
        return pd.DataFrame(), pd.DataFrame()

    active_platforms = platforms[platforms['Freigabe'] == True].copy()
    states = [init_platform_state(row, base_wood_height, layer_spacer_height, gap_length) for _, row in active_platforms.iterrows()]
    not_loaded: List[Dict[str, Any]] = []

    pending: List[pd.Series] = [row for _, row in units.iterrows()]

    def append_not_loaded(unit: pd.Series) -> None:
        not_loaded.append({
            'Fuhre_Nr': None,
            'Fuhrenoption': '',
            'Pritschenname': '',
            'Pritsche': 'NICHT VERLADEN',
            'Einheit_ID': unit['Einheit_ID'],
            'Typ': unit['Typ'],
            'Anzahl_Bauteile': int(unit['Anzahl_Bauteile']),
            'Bauteile': unit['Bauteile'],
            'Bauteile_Liste': unit.get('Bauteile_Liste', unit.get('Bauteile', '')),
            'Ansicht_Attribut': unit.get('Ansicht_Attribut', ''),
            'Ansicht_Label': unit.get('Ansicht_Label', unit.get('Einheit_ID', '')),
            'Ansicht_Liste': unit.get('Ansicht_Liste', unit.get('Ansicht_Label', '')),
            'Einzellängen_mm': unit.get('Einzellängen_mm', ''),
            'Einzelbreiten_mm': unit.get('Einzelbreiten_mm', ''),
            'Einzelhöhen_mm': unit.get('Einzelhöhen_mm', ''),
            'Einlage_allgemein_mm': safe_number(unit.get('Einlage_allgemein_mm'), 0.0),
            'Bundeinlage_mm': safe_number(unit.get('Bundeinlage_mm'), 0.0),
            'X_mm': None,
            'Y_mm': None,
            'Z_mm': None,
            'Länge_mm': unit['Länge_mm'],
            'Breite_mm': unit['Breite_mm'],
            'Höhe_mm': unit['Höhe_mm'],
            'Drehung': None,
            'Ebene': 'nicht passend',
            'Gewicht_kg': round(float(unit['Gewicht_kg']), 2),
        })

    while pending:
        placed = False

        # Ruhige Praxislogik:
        # Keine aggressive Lückenfüllung mit späteren Einzelteilen/Bunden.
        # Die Reihenfolge bleibt stabil; dadurch entstehen weniger chaotische Stapel
        # und deutlich weniger künstlicher Unterbau.
        unit = pending.pop(0)
        for state in states:
            result = try_place_unit(
                state,
                unit,
                allow_beside=allow_beside,
                allow_stack=allow_stack,
                allow_rotation=allow_rotation,
            )
            if result is not None:
                placed = True
                break
        if not placed:
            append_not_loaded(unit)

    placements = []
    summary = []
    for state in states:
        placements.extend(state['placements'])
        summary.append({
            'Fuhre_Nr': state['Fuhre_Nr'],
            'Fuhrenoption': state['Fuhrenoption'],
            'Pritschenname': state['Pritschenname'],
            'Pritsche': state['Pritsche'],
            'Länge genutzt_mm': round(state['used_length'], 1),
            'Breite genutzt_mm': round(state['used_width'], 1),
            'Höhe genutzt_mm': round(state['used_height'], 1),
            'Gewicht genutzt_kg': round(state['total_weight'], 2),
            'Eigengewicht Pritsche_kg': round(state.get('Eigengewicht_Pritsche_kg', 0.0), 2),
            'Gesamtgewicht inkl. Pritsche_kg': round(state['total_weight'] + state.get('Eigengewicht_Pritsche_kg', 0.0), 2),
            'Max Länge effektiv_mm': round(state['Eff_Länge_mm'], 1),
            'Max Breite_mm': round(state['Breite_mm'], 1),
            'Max Höhe_mm': round(state['Max_Höhe_mm'], 1),
            'Max Gewicht_kg': round(state['Max_Gewicht_kg'], 1),
        })

    placements.extend(not_loaded)
    return pd.DataFrame(placements), pd.DataFrame(summary)



# -----------------------------------------------------------------------------
# V17: ruhiger Neuaufbau der Platzierlogik
# -----------------------------------------------------------------------------
def _clean_init_runtime_state(state: Dict[str, Any]) -> None:
    """Initialisiert die einfachen Lagen-/Spurwerte für die ruhige Verladelogik.

    V28: links und rechts bekommen eigene Höhenentwicklung. Damit wird nicht mehr
    künstlich versucht, die ganze Pritsche auf gleicher Lagehöhe zu halten.
    """
    base_z = float(state.get('base_wood_height', 0.0))
    state['_clean_center_y'] = float(state['Breite_mm']) / 2.0
    state['_clean_lane_x'] = {'left': 0.0, 'right': 0.0}
    state['_clean_side_z'] = {'left': base_z, 'right': base_z}
    state['_clean_side_layer_height'] = {'left': 0.0, 'right': 0.0}
    state['_clean_side_layer_has_bundle'] = {'left': False, 'right': False}
    state['_clean_layer_z'] = base_z
    state['_clean_layer_height'] = 0.0
    state['_clean_layer_has_bundle'] = False
    state['_clean_layer_units'] = 0
    state['_clean_side_toggle'] = 'left'


def _clean_effective_spacer_for_side(state: Dict[str, Any], side: str, unit: pd.Series) -> float:
    """Bestimmt die Einlage für einen einzelnen Seitenstapel."""
    side_height = float(state.get('_clean_side_layer_height', {}).get(side, 0.0))
    if side_height <= 0:
        return 0.0
    current_has_bundle = bool(state.get('_clean_side_layer_has_bundle', {}).get(side, False))
    next_is_bundle = str(unit.get('Typ', '')).strip() == 'Bund'
    if current_has_bundle or next_is_bundle:
        return float(state.get('layer_spacer_height', 0.0))
    return float(state.get('general_spacer_height', 0.0))


def _clean_start_new_side_layer(state: Dict[str, Any], side: str, unit: pd.Series) -> Optional[Tuple[float, float]]:
    """Startet nur auf einer Seite eine neue Lage.

    Rückgabe: (new_x, new_z) oder None wenn Höhe nicht mehr passt.
    """
    base_z = float(state.get('_clean_side_z', {}).get(side, state.get('base_wood_height', 0.0)))
    base_h = float(state.get('_clean_side_layer_height', {}).get(side, 0.0))
    spacer = _clean_effective_spacer_for_side(state, side, unit)
    new_z = base_z + base_h + max(0.0, spacer)
    if new_z + float(unit['Höhe_mm']) > float(state['Max_Höhe_mm']):
        return None
    return (0.0, new_z)


def _clean_start_new_common_layer(state: Dict[str, Any], unit: pd.Series) -> Optional[Tuple[float, float]]:
    """Startet für breite/mittige Einheiten eine gemeinsame neue Lage über beiden Seiten."""
    left_top = float(state.get('_clean_side_z', {}).get('left', 0.0)) + float(state.get('_clean_side_layer_height', {}).get('left', 0.0))
    right_top = float(state.get('_clean_side_z', {}).get('right', 0.0)) + float(state.get('_clean_side_layer_height', {}).get('right', 0.0))
    next_is_bundle = str(unit.get('Typ', '')).strip() == 'Bund'
    has_bundle = bool(state.get('_clean_side_layer_has_bundle', {}).get('left', False) or state.get('_clean_side_layer_has_bundle', {}).get('right', False) or next_is_bundle)
    spacer = float(state.get('layer_spacer_height' if has_bundle else 'general_spacer_height', 0.0))
    new_z = max(left_top, right_top) + max(0.0, spacer)
    if new_z + float(unit['Höhe_mm']) > float(state['Max_Höhe_mm']):
        return None
    return (0.0, new_z)


def _clean_sync_global_layer_fields(state: Dict[str, Any]) -> None:
    """Hält alte Summary-Felder mit der neuen Seitenlogik synchron."""
    left_z = float(state.get('_clean_side_z', {}).get('left', 0.0))
    right_z = float(state.get('_clean_side_z', {}).get('right', 0.0))
    left_h = float(state.get('_clean_side_layer_height', {}).get('left', 0.0))
    right_h = float(state.get('_clean_side_layer_height', {}).get('right', 0.0))
    state['_clean_layer_z'] = min(left_z, right_z)
    state['_clean_layer_height'] = max(left_z + left_h, right_z + right_h) - state['_clean_layer_z']
    state['_clean_layer_has_bundle'] = bool(state.get('_clean_side_layer_has_bundle', {}).get('left', False) or state.get('_clean_side_layer_has_bundle', {}).get('right', False))
    state['layer_max_height'] = max(left_h, right_h)
    state['current_layer_has_bundle'] = bool(state['_clean_layer_has_bundle'])


def _clean_y_for_side(platform_width: float, item_width: float, side: str) -> float:
    center = platform_width / 2.0
    if side == 'left':
        return max(0.0, center - item_width)
    return min(center, max(0.0, platform_width - item_width))


def _clean_try_place_on_current_layer(
    state: Dict[str, Any],
    unit: pd.Series,
    allow_beside: bool,
    allow_rotation: bool,
    allow_stack: bool = True,
) -> Optional[Dict[str, Any]]:
    """Platziert eine Einheit mit unabhängigen Höhen links/rechts.

    V28:
    - links und rechts dürfen eigene Lagenhöhen haben
    - breite Einheiten werden als gemeinsame mittige Lage behandelt
    - die Ladung wird in X später als ganzer Block mittig auf die Pritsche zentriert
    """
    height = float(unit['Höhe_mm'])
    weight = float(unit['Gewicht_kg'])
    platform_width = float(state['Breite_mm'])
    center = platform_width / 2.0

    best: Optional[Tuple[float, Dict[str, Any], Tuple[str, ...], float, Dict[str, Any]]] = None

    for use_length, use_width, rotation in _unit_orientations_for_state(state, unit, allow_rotation):
        if use_width > platform_width + 0.001:
            continue

        wide = bool(use_width >= platform_width * 0.75 or use_width > center + 1.0)
        if wide:
            cur_x = max(float(state['_clean_lane_x']['left']), float(state['_clean_lane_x']['right']))
            cur_z = max(float(state['_clean_side_z']['left']), float(state['_clean_side_z']['right']))
            y = max(0.0, (platform_width - use_width) / 2.0)
            candidates = []
            if can_place(state, cur_x, y, cur_z, use_length, use_width, height, weight):
                candidates.append((cur_x, cur_z, False))
            if allow_stack:
                new_common = _clean_start_new_common_layer(state, unit)
                if new_common is not None:
                    nx, nz = new_common
                    if can_place(state, nx, y, nz, use_length, use_width, height, weight):
                        candidates.append((nx, nz, True))
            for x, z, new_layer in candidates:
                score = z * 1000000.0 + x * 1000.0 - use_width
                candidate = {
                    'x': x, 'y': y, 'z': z, 'length': use_length, 'width': use_width,
                    'height': height, 'rotation': rotation, 'mode': 'breit/mittig',
                    'new_layer': new_layer, 'wide': True
                }
                updates = {'left': (x + use_length + float(state['gap_length']), z, height), 'right': (x + use_length + float(state['gap_length']), z, height)}
                if best is None or score < best[0]:
                    best = (score, candidate, ('left', 'right'), x + use_length + float(state['gap_length']), updates)
            continue

        sides = ['left']
        if allow_beside:
            lx = float(state['_clean_lane_x']['left'])
            rx = float(state['_clean_lane_x']['right'])
            if abs(lx - rx) < 1.0:
                first = str(state.get('_clean_side_toggle', 'left'))
                sides = [first, 'right' if first == 'left' else 'left']
            elif lx < rx:
                sides = ['left', 'right']
            else:
                sides = ['right', 'left']

        for side in sides:
            y = _clean_y_for_side(platform_width, use_width, side)
            side_x = float(state['_clean_lane_x'][side])
            side_z = float(state['_clean_side_z'][side])
            candidates = []
            if can_place(state, side_x, y, side_z, use_length, use_width, height, weight):
                candidates.append((side_x, side_z, False))
            if allow_stack:
                new_side = _clean_start_new_side_layer(state, side, unit)
                if new_side is not None:
                    nx, nz = new_side
                    if can_place(state, nx, y, nz, use_length, use_width, height, weight):
                        candidates.append((nx, nz, True))
            for x, z, new_layer in candidates:
                other = 'right' if side == 'left' else 'left'
                after_x = x + use_length + float(state['gap_length'])
                balance = abs(after_x - float(state['_clean_lane_x'][other]))
                # Niedrige Z zuerst, dann kompakte X-Position.
                score = z * 1000000.0 + x * 1000.0 + balance
                candidate = {
                    'x': x, 'y': y, 'z': z, 'length': use_length, 'width': use_width,
                    'height': height, 'rotation': rotation, 'mode': f'{side} / unabhängige Stapelhöhe',
                    'new_layer': new_layer, 'wide': False
                }
                updates = {side: (after_x, z, height)}
                if best is None or score < best[0]:
                    best = (score, candidate, (side,), after_x, updates)

    if best is None:
        return None

    _score, cand, affected_sides, new_lane_x, side_updates = best
    state['current_z'] = cand['z']
    placement = commit_place(
        state,
        unit,
        cand['x'], cand['y'], cand['z'],
        cand['length'], cand['width'], cand['height'],
        cand['rotation'], cand['mode'],
    )

    is_bundle = str(unit.get('Typ', '')).strip() == 'Bund'
    for side, (after_x, z, h) in side_updates.items():
        if cand.get('wide') and cand.get('new_layer'):
            state['_clean_lane_x'][side] = after_x
            state['_clean_side_z'][side] = z
            state['_clean_side_layer_height'][side] = h
            state['_clean_side_layer_has_bundle'][side] = is_bundle
        elif cand.get('new_layer'):
            state['_clean_lane_x'][side] = after_x
            state['_clean_side_z'][side] = z
            state['_clean_side_layer_height'][side] = h
            state['_clean_side_layer_has_bundle'][side] = is_bundle
        else:
            state['_clean_lane_x'][side] = max(float(state['_clean_lane_x'][side]), after_x)
            state['_clean_side_layer_height'][side] = max(float(state['_clean_side_layer_height'][side]), h)
            state['_clean_side_layer_has_bundle'][side] = bool(state['_clean_side_layer_has_bundle'][side] or is_bundle)
            # side_z bleibt auf derselben Lage

    if len(affected_sides) == 1:
        state['_clean_side_toggle'] = 'right' if affected_sides[0] == 'left' else 'left'

    state['_clean_layer_units'] = int(state.get('_clean_layer_units', 0)) + 1
    _clean_sync_global_layer_fields(state)
    return placement


def _clean_place_unit(
    state: Dict[str, Any],
    unit: pd.Series,
    allow_beside: bool,
    allow_stack: bool,
    allow_rotation: bool,
) -> Optional[Dict[str, Any]]:
    """Platziert eine Einheit direkt mit der neuen Seiten-/Stapel-Logik."""
    return _clean_try_place_on_current_layer(state, unit, allow_beside, allow_rotation, allow_stack=allow_stack)


def create_loading_plan(
    units: pd.DataFrame,
    platforms: pd.DataFrame,
    base_wood_height: float,
    layer_spacer_height: float,
    gap_length: float,
    allow_beside: bool,
    allow_stack: bool,
    allow_rotation: bool,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Greedy-Verladevorschlag für eine einzelne Fuhre."""
    if units.empty or platforms.empty:
        return pd.DataFrame(), pd.DataFrame()

    active_platforms = platforms[platforms['Freigabe'] == True].copy()
    states = [init_platform_state(row, base_wood_height, layer_spacer_height, gap_length) for _, row in active_platforms.iterrows()]
    not_loaded: List[Dict[str, Any]] = []

    pending: List[pd.Series] = [row for _, row in units.iterrows()]

    def append_not_loaded(unit: pd.Series) -> None:
        not_loaded.append({
            'Fuhre_Nr': None,
            'Fuhrenoption': '',
            'Pritschenname': '',
            'Pritsche': 'NICHT VERLADEN',
            'Einheit_ID': unit['Einheit_ID'],
            'Typ': unit['Typ'],
            'Anzahl_Bauteile': int(unit['Anzahl_Bauteile']),
            'Bauteile': unit['Bauteile'],
            'Bauteile_Liste': unit.get('Bauteile_Liste', unit.get('Bauteile', '')),
            'Ansicht_Attribut': unit.get('Ansicht_Attribut', ''),
            'Ansicht_Label': unit.get('Ansicht_Label', unit.get('Einheit_ID', '')),
            'Ansicht_Liste': unit.get('Ansicht_Liste', unit.get('Ansicht_Label', '')),
            'Einzellängen_mm': unit.get('Einzellängen_mm', ''),
            'Einzelbreiten_mm': unit.get('Einzelbreiten_mm', ''),
            'Einzelhöhen_mm': unit.get('Einzelhöhen_mm', ''),
            'Einlage_allgemein_mm': safe_number(unit.get('Einlage_allgemein_mm'), 0.0),
            'Bundeinlage_mm': safe_number(unit.get('Bundeinlage_mm'), 0.0),
            'X_mm': None,
            'Y_mm': None,
            'Z_mm': None,
            'Länge_mm': unit['Länge_mm'],
            'Breite_mm': unit['Breite_mm'],
            'Höhe_mm': unit['Höhe_mm'],
            'Drehung': None,
            'Ebene': 'nicht passend',
            'Gewicht_kg': round(float(unit['Gewicht_kg']), 2),
        })

    while pending:
        placed = False

        # Ruhige Praxislogik:
        # Keine aggressive Lückenfüllung mit späteren Einzelteilen/Bunden.
        # Die Reihenfolge bleibt stabil; dadurch entstehen weniger chaotische Stapel
        # und deutlich weniger künstlicher Unterbau.
        unit = pending.pop(0)
        for state in states:
            result = try_place_unit(
                state,
                unit,
                allow_beside=allow_beside,
                allow_stack=allow_stack,
                allow_rotation=allow_rotation,
            )
            if result is not None:
                placed = True
                break
        if not placed:
            append_not_loaded(unit)

    placements = []
    summary = []
    for state in states:
        placements.extend(state['placements'])
        summary.append({
            'Fuhre_Nr': state['Fuhre_Nr'],
            'Fuhrenoption': state['Fuhrenoption'],
            'Pritschenname': state['Pritschenname'],
            'Pritsche': state['Pritsche'],
            'Länge genutzt_mm': round(state['used_length'], 1),
            'Breite genutzt_mm': round(state['used_width'], 1),
            'Höhe genutzt_mm': round(state['used_height'], 1),
            'Gewicht genutzt_kg': round(state['total_weight'], 2),
            'Eigengewicht Pritsche_kg': round(state.get('Eigengewicht_Pritsche_kg', 0.0), 2),
            'Gesamtgewicht inkl. Pritsche_kg': round(state['total_weight'] + state.get('Eigengewicht_Pritsche_kg', 0.0), 2),
            'Max Länge effektiv_mm': round(state['Eff_Länge_mm'], 1),
            'Max Breite_mm': round(state['Breite_mm'], 1),
            'Max Höhe_mm': round(state['Max_Höhe_mm'], 1),
            'Max Gewicht_kg': round(state['Max_Gewicht_kg'], 1),
        })

    placements.extend(not_loaded)
    return pd.DataFrame(placements), pd.DataFrame(summary)



# -----------------------------------------------------------------------------
# V17: ruhiger Neuaufbau der Platzierlogik
# -----------------------------------------------------------------------------
def _clean_init_runtime_state(state: Dict[str, Any]) -> None:
    """Initialisiert die einfachen Lagen-/Spurwerte für die ruhige Verladelogik.

    V28: links und rechts bekommen eigene Höhenentwicklung. Damit wird nicht mehr
    künstlich versucht, die ganze Pritsche auf gleicher Lagehöhe zu halten.
    """
    base_z = float(state.get('base_wood_height', 0.0))
    state['_clean_center_y'] = float(state['Breite_mm']) / 2.0
    state['_clean_lane_x'] = {'left': 0.0, 'right': 0.0}
    state['_clean_side_z'] = {'left': base_z, 'right': base_z}
    state['_clean_side_layer_height'] = {'left': 0.0, 'right': 0.0}
    state['_clean_side_layer_has_bundle'] = {'left': False, 'right': False}
    state['_clean_layer_z'] = base_z
    state['_clean_layer_height'] = 0.0
    state['_clean_layer_has_bundle'] = False
    state['_clean_layer_units'] = 0
    state['_clean_side_toggle'] = 'left'


def _clean_effective_spacer_for_side(state: Dict[str, Any], side: str, unit: pd.Series) -> float:
    """Bestimmt die Einlage für einen einzelnen Seitenstapel."""
    side_height = float(state.get('_clean_side_layer_height', {}).get(side, 0.0))
    if side_height <= 0:
        return 0.0
    current_has_bundle = bool(state.get('_clean_side_layer_has_bundle', {}).get(side, False))
    next_is_bundle = str(unit.get('Typ', '')).strip() == 'Bund'
    if current_has_bundle or next_is_bundle:
        return float(state.get('layer_spacer_height', 0.0))
    return float(state.get('general_spacer_height', 0.0))


def _clean_start_new_side_layer(state: Dict[str, Any], side: str, unit: pd.Series) -> Optional[Tuple[float, float]]:
    """Startet nur auf einer Seite eine neue Lage.

    Rückgabe: (new_x, new_z) oder None wenn Höhe nicht mehr passt.
    """
    base_z = float(state.get('_clean_side_z', {}).get(side, state.get('base_wood_height', 0.0)))
    base_h = float(state.get('_clean_side_layer_height', {}).get(side, 0.0))
    spacer = _clean_effective_spacer_for_side(state, side, unit)
    new_z = base_z + base_h + max(0.0, spacer)
    if new_z + float(unit['Höhe_mm']) > float(state['Max_Höhe_mm']):
        return None
    return (0.0, new_z)


def _clean_start_new_common_layer(state: Dict[str, Any], unit: pd.Series) -> Optional[Tuple[float, float]]:
    """Startet für breite/mittige Einheiten eine gemeinsame neue Lage über beiden Seiten."""
    left_top = float(state.get('_clean_side_z', {}).get('left', 0.0)) + float(state.get('_clean_side_layer_height', {}).get('left', 0.0))
    right_top = float(state.get('_clean_side_z', {}).get('right', 0.0)) + float(state.get('_clean_side_layer_height', {}).get('right', 0.0))
    next_is_bundle = str(unit.get('Typ', '')).strip() == 'Bund'
    has_bundle = bool(state.get('_clean_side_layer_has_bundle', {}).get('left', False) or state.get('_clean_side_layer_has_bundle', {}).get('right', False) or next_is_bundle)
    spacer = float(state.get('layer_spacer_height' if has_bundle else 'general_spacer_height', 0.0))
    new_z = max(left_top, right_top) + max(0.0, spacer)
    if new_z + float(unit['Höhe_mm']) > float(state['Max_Höhe_mm']):
        return None
    return (0.0, new_z)


def _clean_sync_global_layer_fields(state: Dict[str, Any]) -> None:
    """Hält alte Summary-Felder mit der neuen Seitenlogik synchron."""
    left_z = float(state.get('_clean_side_z', {}).get('left', 0.0))
    right_z = float(state.get('_clean_side_z', {}).get('right', 0.0))
    left_h = float(state.get('_clean_side_layer_height', {}).get('left', 0.0))
    right_h = float(state.get('_clean_side_layer_height', {}).get('right', 0.0))
    state['_clean_layer_z'] = min(left_z, right_z)
    state['_clean_layer_height'] = max(left_z + left_h, right_z + right_h) - state['_clean_layer_z']
    state['_clean_layer_has_bundle'] = bool(state.get('_clean_side_layer_has_bundle', {}).get('left', False) or state.get('_clean_side_layer_has_bundle', {}).get('right', False))
    state['layer_max_height'] = max(left_h, right_h)
    state['current_layer_has_bundle'] = bool(state['_clean_layer_has_bundle'])


def _clean_y_for_side(platform_width: float, item_width: float, side: str) -> float:
    center = platform_width / 2.0
    if side == 'left':
        return max(0.0, center - item_width)
    return min(center, max(0.0, platform_width - item_width))


def _clean_try_place_on_current_layer(
    state: Dict[str, Any],
    unit: pd.Series,
    allow_beside: bool,
    allow_rotation: bool,
    allow_stack: bool = True,
) -> Optional[Dict[str, Any]]:
    """Platziert eine Einheit mit unabhängigen Höhen links/rechts.

    V28:
    - links und rechts dürfen eigene Lagenhöhen haben
    - breite Einheiten werden als gemeinsame mittige Lage behandelt
    - die Ladung wird in X später als ganzer Block mittig auf die Pritsche zentriert
    """
    height = float(unit['Höhe_mm'])
    weight = float(unit['Gewicht_kg'])
    platform_width = float(state['Breite_mm'])
    center = platform_width / 2.0

    best: Optional[Tuple[float, Dict[str, Any], Tuple[str, ...], float, Dict[str, Any]]] = None

    for use_length, use_width, rotation in _unit_orientations_for_state(state, unit, allow_rotation):
        if use_width > platform_width + 0.001:
            continue

        wide = bool(use_width >= platform_width * 0.75 or use_width > center + 1.0)
        if wide:
            cur_x = max(float(state['_clean_lane_x']['left']), float(state['_clean_lane_x']['right']))
            cur_z = max(float(state['_clean_side_z']['left']), float(state['_clean_side_z']['right']))
            y = max(0.0, (platform_width - use_width) / 2.0)
            candidates = []
            if can_place(state, cur_x, y, cur_z, use_length, use_width, height, weight):
                candidates.append((cur_x, cur_z, False))
            if allow_stack:
                new_common = _clean_start_new_common_layer(state, unit)
                if new_common is not None:
                    nx, nz = new_common
                    if can_place(state, nx, y, nz, use_length, use_width, height, weight):
                        candidates.append((nx, nz, True))
            for x, z, new_layer in candidates:
                score = z * 1000000.0 + x * 1000.0 - use_width
                candidate = {
                    'x': x, 'y': y, 'z': z, 'length': use_length, 'width': use_width,
                    'height': height, 'rotation': rotation, 'mode': 'breit/mittig',
                    'new_layer': new_layer, 'wide': True
                }
                updates = {'left': (x + use_length + float(state['gap_length']), z, height), 'right': (x + use_length + float(state['gap_length']), z, height)}
                if best is None or score < best[0]:
                    best = (score, candidate, ('left', 'right'), x + use_length + float(state['gap_length']), updates)
            continue

        sides = ['left']
        if allow_beside:
            lx = float(state['_clean_lane_x']['left'])
            rx = float(state['_clean_lane_x']['right'])
            if abs(lx - rx) < 1.0:
                first = str(state.get('_clean_side_toggle', 'left'))
                sides = [first, 'right' if first == 'left' else 'left']
            elif lx < rx:
                sides = ['left', 'right']
            else:
                sides = ['right', 'left']

        for side in sides:
            y = _clean_y_for_side(platform_width, use_width, side)
            side_x = float(state['_clean_lane_x'][side])
            side_z = float(state['_clean_side_z'][side])
            candidates = []
            if can_place(state, side_x, y, side_z, use_length, use_width, height, weight):
                candidates.append((side_x, side_z, False))
            if allow_stack:
                new_side = _clean_start_new_side_layer(state, side, unit)
                if new_side is not None:
                    nx, nz = new_side
                    if can_place(state, nx, y, nz, use_length, use_width, height, weight):
                        candidates.append((nx, nz, True))
            for x, z, new_layer in candidates:
                other = 'right' if side == 'left' else 'left'
                after_x = x + use_length + float(state['gap_length'])
                balance = abs(after_x - float(state['_clean_lane_x'][other]))
                # Niedrige Z zuerst, dann kompakte X-Position.
                score = z * 1000000.0 + x * 1000.0 + balance
                candidate = {
                    'x': x, 'y': y, 'z': z, 'length': use_length, 'width': use_width,
                    'height': height, 'rotation': rotation, 'mode': f'{side} / unabhängige Stapelhöhe',
                    'new_layer': new_layer, 'wide': False
                }
                updates = {side: (after_x, z, height)}
                if best is None or score < best[0]:
                    best = (score, candidate, (side,), after_x, updates)

    if best is None:
        return None

    _score, cand, affected_sides, new_lane_x, side_updates = best
    state['current_z'] = cand['z']
    placement = commit_place(
        state,
        unit,
        cand['x'], cand['y'], cand['z'],
        cand['length'], cand['width'], cand['height'],
        cand['rotation'], cand['mode'],
    )

    is_bundle = str(unit.get('Typ', '')).strip() == 'Bund'
    for side, (after_x, z, h) in side_updates.items():
        if cand.get('wide') and cand.get('new_layer'):
            state['_clean_lane_x'][side] = after_x
            state['_clean_side_z'][side] = z
            state['_clean_side_layer_height'][side] = h
            state['_clean_side_layer_has_bundle'][side] = is_bundle
        elif cand.get('new_layer'):
            state['_clean_lane_x'][side] = after_x
            state['_clean_side_z'][side] = z
            state['_clean_side_layer_height'][side] = h
            state['_clean_side_layer_has_bundle'][side] = is_bundle
        else:
            state['_clean_lane_x'][side] = max(float(state['_clean_lane_x'][side]), after_x)
            state['_clean_side_layer_height'][side] = max(float(state['_clean_side_layer_height'][side]), h)
            state['_clean_side_layer_has_bundle'][side] = bool(state['_clean_side_layer_has_bundle'][side] or is_bundle)
            # side_z bleibt auf derselben Lage

    if len(affected_sides) == 1:
        state['_clean_side_toggle'] = 'right' if affected_sides[0] == 'left' else 'left'

    state['_clean_layer_units'] = int(state.get('_clean_layer_units', 0)) + 1
    _clean_sync_global_layer_fields(state)
    return placement


def _clean_place_unit(
    state: Dict[str, Any],
    unit: pd.Series,
    allow_beside: bool,
    allow_stack: bool,
    allow_rotation: bool,
) -> Optional[Dict[str, Any]]:
    """Platziert eine Einheit direkt mit der Seiten-/Stapel-Logik.

    Wichtig: Es gibt keine gemeinsame _clean_new_layer-Funktion mehr,
    weil links und rechts unabhängig in der Höhe aufgebaut werden dürfen.
    """
    return _clean_try_place_on_current_layer(
        state,
        unit,
        allow_beside,
        allow_rotation,
        allow_stack=allow_stack,
    )


def create_loading_plan(
    units: pd.DataFrame,
    platforms: pd.DataFrame,
    base_wood_height: float,
    layer_spacer_height: float,
    gap_length: float,
    allow_beside: bool,
    allow_stack: bool,
    allow_rotation: bool,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """V17: ruhiger Lagen-Verladevorschlag für eine einzelne Fuhre.

    Diese Version ersetzt die alte Cursor-Logik bewusst:
    - keine aggressive Lochfüllung,
    - keine nachträgliche Z-Spiegelung,
    - keine Einzelteile aus Bund herauslösen,
    - normale 1200er von der Y-Mitte links/rechts,
    - Restbunde nicht automatisch mittig, sondern links/rechts auf Stapel,
    - Pritschen werden fortlaufend und kompakt gefüllt.
    """
    if units.empty or platforms.empty:
        return pd.DataFrame(), pd.DataFrame()

    active_platforms = platforms[platforms['Freigabe'] == True].copy()
    states = [init_platform_state(row, base_wood_height, layer_spacer_height, gap_length) for _, row in active_platforms.iterrows()]
    for state in states:
        _clean_init_runtime_state(state)

    not_loaded: List[Dict[str, Any]] = []

    def append_not_loaded(unit: pd.Series) -> None:
        not_loaded.append({
            'Fuhre_Nr': None,
            'Fuhrenoption': '',
            'Pritschenname': '',
            'Pritsche': 'NICHT VERLADEN',
            'Einheit_ID': unit['Einheit_ID'],
            'Typ': unit['Typ'],
            'Anzahl_Bauteile': int(unit['Anzahl_Bauteile']),
            'Bauteile': unit['Bauteile'],
            'Bauteile_Liste': unit.get('Bauteile_Liste', unit.get('Bauteile', '')),
            'Ansicht_Attribut': unit.get('Ansicht_Attribut', ''),
            'Ansicht_Label': unit.get('Ansicht_Label', unit.get('Einheit_ID', '')),
            'Ansicht_Liste': unit.get('Ansicht_Liste', unit.get('Ansicht_Label', '')),
            'Einzellängen_mm': unit.get('Einzellängen_mm', ''),
            'Einzelbreiten_mm': unit.get('Einzelbreiten_mm', ''),
            'Einzelhöhen_mm': unit.get('Einzelhöhen_mm', ''),
            'Einlage_allgemein_mm': safe_number(unit.get('Einlage_allgemein_mm'), 0.0),
            'Bundeinlage_mm': safe_number(unit.get('Bundeinlage_mm'), 0.0),
            'X_mm': None,
            'Y_mm': None,
            'Z_mm': None,
            'Länge_mm': unit['Länge_mm'],
            'Breite_mm': unit['Breite_mm'],
            'Höhe_mm': unit['Höhe_mm'],
            'Drehung': None,
            'Ebene': 'nicht passend',
            'Gewicht_kg': round(float(unit['Gewicht_kg']), 2),
        })

    for _, unit in units.iterrows():
        placed = False
        for state in states:
            result = _clean_place_unit(
                state,
                unit,
                allow_beside=allow_beside,
                allow_stack=allow_stack,
                allow_rotation=allow_rotation,
            )
            if result is not None:
                placed = True
                break
        if not placed:
            append_not_loaded(unit)

    placements: List[Dict[str, Any]] = []
    summary: List[Dict[str, Any]] = []
    for state in states:
        placements.extend(state['placements'])
        summary.append({
            'Fuhre_Nr': state['Fuhre_Nr'],
            'Fuhrenoption': state['Fuhrenoption'],
            'Pritschenname': state['Pritschenname'],
            'Pritsche': state['Pritsche'],
            'Länge genutzt_mm': round(state['used_length'], 1),
            'Breite genutzt_mm': round(state['used_width'], 1),
            'Höhe genutzt_mm': round(state['used_height'], 1),
            'Gewicht genutzt_kg': round(state['total_weight'], 2),
            'Eigengewicht Pritsche_kg': round(state.get('Eigengewicht_Pritsche_kg', 0.0), 2),
            'Gesamtgewicht inkl. Pritsche_kg': round(state['total_weight'] + state.get('Eigengewicht_Pritsche_kg', 0.0), 2),
            'Max Länge effektiv_mm': round(state['Eff_Länge_mm'], 1),
            'Max Breite_mm': round(state['Breite_mm'], 1),
            'Max Höhe_mm': round(state['Max_Höhe_mm'], 1),
            'Max Gewicht_kg': round(state['Max_Gewicht_kg'], 1),
        })

    placements.extend(not_loaded)
    return pd.DataFrame(placements), pd.DataFrame(summary)

def center_placements_geometrically(placements_df: pd.DataFrame, platforms_df: pd.DataFrame) -> pd.DataFrame:
    """Ruhige geometrische Ausrichtung der fertigen Verladung.

    Wichtig: Nicht mehr jede Lage grundsätzlich in Y mittig setzen.
    Praxisregel:
    - Untere/normale Lagen bleiben links/rechts kompakt und werden nur als Gruppe leicht zur Mitte korrigiert, wenn sie fast die ganze Breite belegen.
    - Einzelne obere Restbunde werden NICHT automatisch mittig gesetzt.
    - Nur echte breite Elemente/Bunde, die links/rechts nicht sinnvoll aufgeteilt werden können, werden mittig gesetzt.
    - In X wird die Lage weiterhin als Gruppe mittig auf die nutzbare Pritschenlänge gelegt.
    """
    if placements_df.empty or platforms_df.empty:
        return placements_df

    result = placements_df.copy()
    if 'Pritsche' not in result.columns:
        return result

    numeric_cols = ['X_mm', 'Y_mm', 'Z_mm', 'Länge_mm', 'Breite_mm', 'Höhe_mm']
    for col in numeric_cols:
        if col in result.columns:
            result[col] = pd.to_numeric(result[col], errors='coerce')

    platform_lookup = {str(row.get('Pritsche', '')): row for _, row in platforms_df.iterrows()}
    helper_types = {'Unterbau', 'Kantholz', 'Bundeinlage', 'Einlage', 'Lagenholz'}

    for platform_name, platform_row in platform_lookup.items():
        if not platform_name or platform_name == 'NICHT VERLADEN':
            continue

        eff_length = (
            safe_number(platform_row.get('Länge_mm'))
            + safe_number(platform_row.get('Überhang_vorne_mm'))
            + safe_number(platform_row.get('Überhang_hinten_mm'))
        )
        platform_width = safe_number(platform_row.get('Breite_mm'))
        if eff_length <= 0 or platform_width <= 0:
            continue

        mask_platform = (
            result['Pritsche'].astype(str).eq(platform_name)
            & result['X_mm'].notna()
            & result['Y_mm'].notna()
            & result['Z_mm'].notna()
            & result['Länge_mm'].notna()
            & result['Breite_mm'].notna()
        )
        if not mask_platform.any():
            continue

        non_helper_mask = mask_platform & ~result.get('Typ', pd.Series(dtype=str)).astype(str).isin(helper_types)
        layer_keys = result.loc[mask_platform, 'Z_mm'].round(1).unique().tolist()
        if not layer_keys:
            continue

        # X nicht mehr je Lage separat mittig schieben. Das hatte in der Seitenansicht
        # scheinbare Löcher/Überstände erzeugt, weil obere Lagen unabhängig von den
        # tragenden unteren Lagen verschoben wurden. Die ganze Ladung wird als Block
        # auf der Pritsche zentriert; die Lagen bleiben zueinander stabil.
        global_x0 = result.loc[mask_platform, 'X_mm'].min()
        global_x1 = (result.loc[mask_platform, 'X_mm'] + result.loc[mask_platform, 'Länge_mm']).max()
        global_span_x = global_x1 - global_x0
        if 0 < global_span_x <= eff_length:
            global_shift_x = (eff_length - global_span_x) / 2 - global_x0
            result.loc[mask_platform, 'X_mm'] = (result.loc[mask_platform, 'X_mm'] + global_shift_x).round(1)

        # Oberste echte Bauteil-/Bundlage. Hilfszeilen zählen nicht als obere Restlage.
        if non_helper_mask.any():
            top_layer_key = float(result.loc[non_helper_mask, 'Z_mm'].round(1).max())
        else:
            top_layer_key = float(max(layer_keys))

        for layer_key in layer_keys:
            layer_mask = mask_platform & result['Z_mm'].round(1).eq(layer_key)
            if not layer_mask.any():
                continue

            x0 = result.loc[layer_mask, 'X_mm'].min()
            x1 = (result.loc[layer_mask, 'X_mm'] + result.loc[layer_mask, 'Länge_mm']).max()
            y0 = result.loc[layer_mask, 'Y_mm'].min()
            y1 = (result.loc[layer_mask, 'Y_mm'] + result.loc[layer_mask, 'Breite_mm']).max()
            span_x = x1 - x0
            span_y = y1 - y0

            # X wurde bereits global als ganze Ladung zentriert. Dadurch entsteht die
            # Verladung in der Länge ruhiger von der Mitte aus und nicht einseitig bündig.
            # Je-Lage-X-Verschiebung bleibt bewusst aus, damit unten keine unnötigen
            # Löcher und oben keine scheinbaren Überschneidungen entstehen.

            # Y: nicht jede Lage mittig machen. Nur vollständige Breitenlagen leicht zentrieren,
            # breite Einzelteile mittig setzen oder die oberste Restlage mittig setzen.
            real_rows = result.loc[layer_mask & ~result.get('Typ', pd.Series(dtype=str)).astype(str).isin(helper_types)]
            real_count = len(real_rows)
            is_top_real_layer = abs(float(layer_key) - top_layer_key) < 0.1
            almost_full_width = span_y >= platform_width * 0.90
            has_wide_single = False
            if real_count == 1:
                only_width = safe_number(real_rows.iloc[0].get('Breite_mm'), 0.0)
                # Erst ab echter Breite mittig. Normale 1200er/1180er Restbunde
                # bleiben auf links oder rechts und werden NICHT in die Mitte gezogen.
                has_wide_single = only_width >= platform_width * 0.75
            # Normale 1200er bleiben links/rechts an der Pritschenmitte.
            # Mittig nur: fast volle Breite oder wirklich breite Einzelteile.
            # Oberste Restbunde werden bewusst NICHT mehr automatisch mittig gesetzt.
            center_y = bool(almost_full_width or has_wide_single)

            if center_y and 0 < span_y <= platform_width:
                shift_y = (platform_width - span_y) / 2 - y0
                result.loc[layer_mask, 'Y_mm'] = (result.loc[layer_mask, 'Y_mm'] + shift_y).round(1)
            else:
                # Nicht mehr auf Y=0 zurückschieben. Sonst wird eine korrekt rechts
                # platzierte Lage fälschlich nach links gezogen.
                pass

            if 'Ebene' in result.columns:
                result.loc[layer_mask, 'Ebene'] = result.loc[layer_mask, 'Ebene'].astype(str).apply(
                    lambda v: v if 'ruhig ausgerichtet' in v.lower() else f'{v} / ruhig ausgerichtet'
                )

    return result


def center_length_groups_from_platform_center(placements_df: pd.DataFrame, platforms_df: pd.DataFrame) -> pd.DataFrame:
    """Richtet Stapelgruppen in X wieder von der Pritschenmitte aus.

    V28:
    - Die Ladung wird in der Länge nicht an einer Stirnseite bündig aufgebaut.
    - Jede echte Stapelgruppe wird in X um die Pritschenmitte gesetzt.
    - Seitenansichten werden danach seitenbezogen gefiltert, damit linke und
      rechte Stapel nicht mehr optisch ineinanderlaufen.
    """
    if placements_df is None or placements_df.empty or platforms_df is None or platforms_df.empty:
        return placements_df.copy() if placements_df is not None else pd.DataFrame()

    result = placements_df.copy()
    for col in ['X_mm', 'Y_mm', 'Z_mm', 'Länge_mm', 'Breite_mm']:
        if col in result.columns:
            result[col] = pd.to_numeric(result[col], errors='coerce')

    helper_types = {'Unterbau', 'Kantholz', 'Bundeinlage', 'Einlage', 'Lagenholz'}
    platform_lookup = {str(r.get('Pritsche', '')): r for _, r in platforms_df.iterrows()}

    for pname, prow in platform_lookup.items():
        eff_length = (
            safe_number(prow.get('Länge_mm'))
            + safe_number(prow.get('Überhang_vorne_mm'))
            + safe_number(prow.get('Überhang_hinten_mm'))
        )
        width = safe_number(prow.get('Breite_mm'), 0.0)
        if eff_length <= 0 or width <= 0:
            continue

        mask = (
            result['Pritsche'].astype(str).eq(pname)
            & result['X_mm'].notna()
            & result['Y_mm'].notna()
            & result['Z_mm'].notna()
            & result['Länge_mm'].notna()
            & ~result.get('Typ', pd.Series(dtype=str)).astype(str).isin(helper_types)
        )
        if not mask.any():
            continue

        center_y = width / 2.0
        subset = result.loc[mask].copy()
        side_keys = []
        for _, r in subset.iterrows():
            y0 = safe_number(r.get('Y_mm'))
            by = safe_number(r.get('Breite_mm'))
            y1 = y0 + by
            if by >= width * 0.75 or (y0 < center_y < y1):
                side_keys.append('mittig')
            elif y0 + by / 2.0 <= center_y:
                side_keys.append('links')
            else:
                side_keys.append('rechts')

        # Gleiche Z-Lage + gleiche Breiten-Seite = eine Stapelgruppe.
        subset['_x_center_group'] = [f"{round(float(z), 1)}|{s}" for z, s in zip(subset['Z_mm'], side_keys)]

        for _group_key, grp in subset.groupby('_x_center_group', sort=False):
            idxs = grp.index.tolist()
            x0 = float(result.loc[idxs, 'X_mm'].min())
            x1 = float((result.loc[idxs, 'X_mm'] + result.loc[idxs, 'Länge_mm']).max())
            span = x1 - x0
            if span <= 0 or span > eff_length:
                continue
            target_x0 = (eff_length - span) / 2.0
            shift = target_x0 - x0

            new_x0 = x0 + shift
            new_x1 = x1 + shift
            if new_x0 < 0:
                shift -= new_x0
            if new_x1 > eff_length:
                shift -= (new_x1 - eff_length)

            result.loc[idxs, 'X_mm'] = (result.loc[idxs, 'X_mm'] + shift).round(1)
            if 'Ebene' in result.columns:
                result.loc[idxs, 'Ebene'] = result.loc[idxs, 'Ebene'].astype(str).apply(
                    lambda v: v if 'X mittig je Stapel' in v else f'{v} / X mittig je Stapel'
                )

    return result


def normalize_y_from_platform_center(placements_df: pd.DataFrame, platforms_df: pd.DataFrame) -> pd.DataFrame:
    """Richtet die Y-Positionen praxisnah von der Pritschenmitte aus.

    Ziel:
    - 1200-mm-Elemente links/rechts treffen sich an der Pritschenmitte.
    - nur wirklich breite Elemente werden mittig geführt.
    - einzelne obere Restbunde werden NICHT mittig gezogen.
    - dadurch bleibt die reale Auflage links/rechts nachvollziehbar.
    """
    if placements_df is None or placements_df.empty or platforms_df is None or platforms_df.empty:
        return placements_df.copy() if placements_df is not None else pd.DataFrame()

    result = placements_df.copy()
    if 'Pritsche' not in result.columns:
        return result

    for col in ['Y_mm', 'Z_mm', 'Breite_mm']:
        if col in result.columns:
            result[col] = pd.to_numeric(result[col], errors='coerce')

    helper_types = {'Unterbau', 'Kantholz', 'Bundeinlage', 'Einlage', 'Lagenholz'}
    p_lookup = {str(r.get('Pritsche', '')): r for _, r in platforms_df.iterrows()}

    for pname, prow in p_lookup.items():
        width = safe_number(prow.get('Breite_mm'), 0.0)
        if width <= 0:
            continue
        center = width / 2.0
        mask = (
            result['Pritsche'].astype(str).eq(pname)
            & result['Y_mm'].notna()
            & result['Z_mm'].notna()
            & result['Breite_mm'].notna()
            & ~result.get('Typ', pd.Series(dtype=str)).astype(str).isin(helper_types)
        )
        if not mask.any():
            continue

        real = result.loc[mask].copy()
        top_z = float(real['Z_mm'].max()) if not real.empty else None

        for idx, row in real.iterrows():
            bw = safe_number(row.get('Breite_mm'), 0.0)
            if bw <= 0 or bw > width:
                continue
            z = safe_number(row.get('Z_mm'), 0.0)
            y = safe_number(row.get('Y_mm'), 0.0)
            y_mid = y + bw / 2.0

            # Breite Elemente / fast volle Pritschenbreite: sauber mittig.
            if bw >= width * 0.75:
                new_y = (width - bw) / 2.0
            else:
                # Normale 1200er/1180er Restbunde: NICHT mittig.
                # Sie werden links oder rechts an der Pritschenmitte ausgerichtet,
                # damit sie real auf einem Stapel aufliegen und nicht optisch überschneiden.
                new_y = center - bw if y_mid <= center else center
                # Sicherheit gegen negative Werte bei leicht breiteren Elementen.
                new_y = max(0.0, min(new_y, width - bw))

            result.at[idx, 'Y_mm'] = round(new_y, 1)

    return result

def invert_vertical_order_by_platform(placements_df: pd.DataFrame, platforms_df: pd.DataFrame) -> pd.DataFrame:
    """Spiegelt die Z-Reihenfolge je belegter Pritsche.

    Zweck: Bei aufsteigender Hauptattribut-Sortierung soll die niedrigste Nummer
    oben liegen. Die Verteilung auf die Pritschen bleibt trotzdem global
    fortlaufend. Es werden keine Bunde aufgelöst und keine X/Y-Positionen
    verändert; nur die Z-Lage der fertigen Verladeeinheiten wird je Pritsche
    gespiegelt.
    """
    if placements_df is None or placements_df.empty or platforms_df is None or platforms_df.empty:
        return placements_df.copy() if placements_df is not None else pd.DataFrame()

    result = placements_df.copy()
    for col in ['X_mm', 'Y_mm', 'Z_mm', 'Länge_mm', 'Breite_mm', 'Höhe_mm']:
        if col in result.columns:
            result[col] = pd.to_numeric(result[col], errors='coerce')

    platform_lookup = {str(row.get('Pritsche', '')): row for _, row in platforms_df.iterrows()}
    helper_types = {'Unterbau', 'Kantholz', 'Bundeinlage', 'Einlage', 'Lagenholz'}

    for pname, prow in platform_lookup.items():
        mask = (
            result['Pritsche'].astype(str).eq(pname)
            & result['Z_mm'].notna()
            & result['Höhe_mm'].notna()
            & ~result.get('Typ', pd.Series(dtype=str)).astype(str).isin(helper_types)
        )
        if not mask.any():
            continue
        base_height = safe_number(prow.get('Kantholz_erste_Lage_mm'), 0.0)
        top_z = float((result.loc[mask, 'Z_mm'] + result.loc[mask, 'Höhe_mm']).max())
        # Jede Einheit bleibt als Block erhalten. Nur die Höhe wird gespiegelt.
        new_z = base_height + (top_z - (result.loc[mask, 'Z_mm'] + result.loc[mask, 'Höhe_mm']))
        result.loc[mask, 'Z_mm'] = new_z.round(1)
        if 'Ebene' in result.columns:
            result.loc[mask, 'Ebene'] = result.loc[mask, 'Ebene'].astype(str).apply(
                lambda v: v if 'niedrigste oben' in v.lower() else f'{v} / niedrigste oben'
            )
    return result



def repack_loaded_platforms_for_lowest_on_top(
    loaded_try: pd.DataFrame,
    units_ordered: pd.DataFrame,
    trip_platforms: pd.DataFrame,
    base_default: float,
    layer_default: float,
    gap_default: float,
    allow_beside: bool,
    allow_stack: bool,
    allow_rotation: bool,
) -> pd.DataFrame:
    """Packt die bereits zugewiesenen Einheiten je Pritsche stabil neu.

    Ziel:
    - Die globale Pritschen-Zuordnung bleibt erhalten.
    - Innerhalb jeder Pritsche liegt bei aufsteigender Sortierung die niedrigste
      Einheit oben. Dafür wird pro Pritsche der zugewiesene Block physisch in
      umgekehrter Reihenfolge aufgebaut, statt nachträglich nur die Z-Werte zu
      spiegeln.
    - X/Y der Lagen bleiben dadurch tragfähiger; weniger scheinbare Löcher unten
      und weniger Überschneidungen oben.
    """
    if loaded_try is None or loaded_try.empty or units_ordered is None or units_ordered.empty:
        return loaded_try

    repacked_frames: List[pd.DataFrame] = []
    unit_lookup = units_ordered.copy()
    unit_lookup['Einheit_ID'] = unit_lookup['Einheit_ID'].astype(str)

    for _, platform in trip_platforms.sort_values('Pritschen_Reihenfolge', kind='stable').iterrows():
        pname = str(platform.get('Pritsche', ''))
        assigned = loaded_try[loaded_try['Pritsche'].astype(str).eq(pname)].copy()
        if assigned.empty:
            continue
        assigned_ids = set(assigned['Einheit_ID'].dropna().astype(str).tolist())
        units_for_platform = unit_lookup[unit_lookup['Einheit_ID'].isin(assigned_ids)].copy()
        if units_for_platform.empty:
            continue

        # Globale Reihenfolge bleibt in units_for_platform enthalten. Physisch wird
        # von unten nach oben umgekehrt aufgebaut, damit die niedrigste Nummer oben liegt.
        units_reversed = units_for_platform.iloc[::-1].reset_index(drop=True)
        platform_df = pd.DataFrame([platform])
        rp, _ = create_loading_plan(
            units_reversed,
            platform_df,
            base_wood_height=base_default,
            layer_spacer_height=layer_default,
            gap_length=gap_default,
            allow_beside=allow_beside,
            allow_stack=allow_stack,
            allow_rotation=allow_rotation,
        )
        rp_loaded = rp[rp['Pritsche'] != 'NICHT VERLADEN'].copy() if rp is not None and not rp.empty else pd.DataFrame()
        rp_ids = set(rp_loaded.get('Einheit_ID', pd.Series(dtype=str)).dropna().astype(str).tolist())
        if assigned_ids.issubset(rp_ids):
            if 'Ebene' in rp_loaded.columns:
                rp_loaded['Ebene'] = rp_loaded['Ebene'].astype(str).apply(
                    lambda v: v if 'stabil neu gepackt' in v.lower() else f'{v} / stabil neu gepackt'
                )
            repacked_frames.append(rp_loaded)
        else:
            # Sicherheitsfallback: Wenn die umgekehrte Packung nicht passt, keine Einheiten verlieren.
            repacked_frames.append(assigned)

    if not repacked_frames:
        return loaded_try
    return pd.concat(repacked_frames, ignore_index=True)

def build_trip_platforms(pritschen_df: pd.DataFrame, fuhrenoption: str, fuhre_nr: int) -> pd.DataFrame:
    rows = pritschen_df[
        (pritschen_df['Fuhrenoption'].astype(str) == str(fuhrenoption)) &
        (pritschen_df['Freigabe'] == True)
    ].copy()
    if rows.empty:
        return rows
    rows = rows.sort_values('Pritschen_Reihenfolge', kind='stable').reset_index(drop=True)
    rows['Fuhre_Nr'] = fuhre_nr
    rows['Pritschenname'] = rows['Pritschenname'].astype(str)
    rows['Pritsche'] = rows.apply(lambda r: f"F{fuhre_nr:02d} {r['Pritschenname']}", axis=1)
    return rows


def create_variant_a_loading_plan(
    sorted_parts: pd.DataFrame,
    options_df: pd.DataFrame,
    pritschen_df: pd.DataFrame,
    standards: Dict[str, Any],
    allow_beside: bool,
    allow_stack: bool,
    allow_rotation: bool,
    center_geometric: bool = True,
    max_fuhren: int = 50,
    use_bundles: bool = True,
    max_bundle_weight: float = 1000.0,
    bundle_spacer_height: float = 40.0,
    general_spacer_height: float = 0.0,
    same_height: bool = True,
    same_width: bool = False,
    same_quality: bool = False,
    same_profile: bool = False,
    label_attr: str = 'Bauteilnummer',
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """V22: geprüfte Block-Suche pro Pritsche mit getrennter Sortier- und Anzeige-/Stapelrichtung.

    Neue Arbeitsweise:
    1. Bauteile werden global sortiert.
    2. Für jede Pritsche wird aus dem nächsten fortlaufenden Bauteilblock die maximale Menge gesucht.
    3. Erst dieser Pritschenblock wird zu Bunden gebildet.
    4. Bunde werden logisch von oben nach unten gebildet.
    5. Für die physische Platzierung wird die Bundfolge umgedreht: hohe Nummern unten, niedrige Nummern oben.

    Dadurch werden keine späteren Bauteile vorgezogen und keine global gebildeten Bunde über Pritschengrenzen gezogen.
    """
    if sorted_parts is None or sorted_parts.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    enabled_options = options_df[options_df['Freigegeben'] == True].copy()
    enabled_options = enabled_options.sort_values('Priorität', kind='stable').reset_index(drop=True)

    remaining_parts = sorted_parts.copy().reset_index(drop=True)
    all_placements: List[pd.DataFrame] = []
    all_summary: List[pd.DataFrame] = []
    all_platforms: List[pd.DataFrame] = []
    all_units: List[pd.DataFrame] = []
    fuhren_log: List[Dict[str, Any]] = []

    base_default = safe_number(standards.get('Standard_Kantholz_erste_Lage'), 80.0)
    layer_default = safe_number(standards.get('Standard_Einlage_zwischen_Lagen'), bundle_spacer_height)
    general_default = safe_number(standards.get('Standard_Einlage_allgemein'), general_spacer_height)
    gap_default = safe_number(standards.get('Längenversatz_je_Lage'), 100.0)

    def _make_part_label(row: pd.Series) -> str:
        return str(row.get(label_attr) or row.get('Bauteilnummer') or row.get('Name') or '')

    def _make_not_loaded_rows_from_parts(parts_block: pd.DataFrame, reason: str, unit_start: int) -> Tuple[pd.DataFrame, pd.DataFrame]:
        if parts_block is None or parts_block.empty:
            return pd.DataFrame(), pd.DataFrame()
        units = _build_units_for_part_block(parts_block, unit_start, block_name='NICHT VERLADEN')
        rows: List[Dict[str, Any]] = []
        for _, unit in units.iterrows():
            rows.append({
                'Fuhre_Nr': None,
                'Fuhrenoption': '',
                'Pritschenname': '',
                'Pritsche': 'NICHT VERLADEN',
                'Einheit_ID': unit['Einheit_ID'],
                'Typ': unit['Typ'],
                'Anzahl_Bauteile': int(unit['Anzahl_Bauteile']),
                'Bauteile': unit['Bauteile'],
                'Bauteile_Liste': unit.get('Bauteile_Liste', unit.get('Bauteile', '')),
                'Ansicht_Attribut': unit.get('Ansicht_Attribut', ''),
                'Ansicht_Label': unit.get('Ansicht_Label', unit.get('Einheit_ID', '')),
                'Ansicht_Liste': unit.get('Ansicht_Liste', unit.get('Ansicht_Label', '')),
                'Einzellängen_mm': unit.get('Einzellängen_mm', ''),
                'Einzelbreiten_mm': unit.get('Einzelbreiten_mm', ''),
                'Einzelhöhen_mm': unit.get('Einzelhöhen_mm', ''),
                'Einlage_allgemein_mm': safe_number(unit.get('Einlage_allgemein_mm'), 0.0),
                'Bundeinlage_mm': safe_number(unit.get('Bundeinlage_mm'), 0.0),
                'X_mm': None,
                'Y_mm': None,
                'Z_mm': None,
                'Länge_mm': unit['Länge_mm'],
                'Breite_mm': unit['Breite_mm'],
                'Höhe_mm': unit['Höhe_mm'],
                'Drehung': None,
                'Ebene': reason,
                'Gewicht_kg': round(float(unit['Gewicht_kg']), 2),
            })
        return pd.DataFrame(rows), units

    def _prepare_trip_platforms(option_name: str, fuhre_nr: int) -> pd.DataFrame:
        trip_platforms = build_trip_platforms(pritschen_df, option_name, fuhre_nr)
        if trip_platforms.empty:
            return trip_platforms
        trip_platforms = trip_platforms.copy()
        trip_platforms['Kantholz_erste_Lage_mm'] = base_default
        trip_platforms['Einlage_zwischen_Lagen_mm'] = layer_default
        trip_platforms['Einlage_allgemein_mm'] = general_default
        return trip_platforms

    def _build_units_for_part_block(parts_block: pd.DataFrame, unit_start: int, block_name: str) -> pd.DataFrame:
        units = build_loading_units(
            parts_block,
            use_bundles=use_bundles,
            max_bundle_weight=max_bundle_weight,
            bundle_spacer_height=bundle_spacer_height,
            general_spacer_height=general_spacer_height,
            same_height=same_height,
            same_width=same_width,
            same_quality=same_quality,
            same_profile=same_profile,
            label_attr=label_attr,
        )
        if units is None or units.empty:
            return pd.DataFrame()
        units = units.copy().reset_index(drop=True)
        new_ids: List[str] = []
        for i, row in units.iterrows():
            prefix = 'B' if str(row.get('Typ', '')).strip() == 'Bund' else 'E'
            new_ids.append(f'{prefix}{unit_start + i:04d}')
        units['Einheit_ID'] = new_ids
        units['Pritschenblock'] = block_name
        units['Logische_Reihenfolge_im_Block'] = list(range(1, len(units) + 1))
        units['Bundbildung'] = 'V21 nach geprüftem Pritschenblock / von oben nach unten'
        return units

    def _pack_block_units_on_platform(block_units_top: pd.DataFrame, platform_row: pd.Series) -> Tuple[bool, pd.DataFrame, pd.DataFrame]:
        if block_units_top is None or block_units_top.empty:
            return False, pd.DataFrame(), pd.DataFrame()
        platform_df = pd.DataFrame([platform_row]).reset_index(drop=True)
        # Wichtig: logisch ist block_units_top oben -> unten. Physisch wird unten begonnen.
        physical_order = block_units_top.iloc[::-1].copy().reset_index(drop=True)
        placements_try, _summary_try = create_loading_plan(
            physical_order,
            platform_df,
            base_wood_height=base_default,
            layer_spacer_height=layer_default,
            gap_length=gap_default,
            allow_beside=allow_beside,
            allow_stack=allow_stack,
            allow_rotation=allow_rotation,
        )
        if placements_try is None or placements_try.empty:
            return False, pd.DataFrame(), pd.DataFrame()
        loaded = placements_try[placements_try['Pritsche'] != 'NICHT VERLADEN'].copy()
        wanted_ids = block_units_top['Einheit_ID'].astype(str).tolist()
        loaded_ids = loaded.get('Einheit_ID', pd.Series(dtype=str)).dropna().astype(str).tolist()
        if set(wanted_ids) != set(loaded_ids):
            return False, pd.DataFrame(), pd.DataFrame()
        loaded = loaded[loaded['Einheit_ID'].astype(str).isin(wanted_ids)].copy()
        if 'Ebene' in loaded.columns:
            loaded['Ebene'] = loaded['Ebene'].astype(str).apply(
                lambda v: v if 'V21 Pritschenblock' in v else f'{v} / V21 Pritschenblock'
            )
        summary = recompute_summary_from_placements(loaded, platform_df)
        return True, loaded, summary

    def _find_max_part_block_for_platform(parts_df: pd.DataFrame, platform_row: pd.Series, unit_start: int) -> Dict[str, Any]:
        """Sucht den grössten gültigen, fortlaufenden Bauteilblock für genau diese Pritsche.

        V21/V22-Änderung:
        Ein früher Kandidat darf fehlschlagen, ohne dass die Suche sofort abbricht.
        Grund: Durch einen zusätzlichen Bauteil kann sich die Bundbildung ändern; dadurch kann
        ein grösserer Block wieder sauberer packbar sein als ein kleiner Zwischenkandidat.
        Abgebrochen wird nur, wenn das rohe Bauteilgewicht bereits über der zulässigen
        Pritschenlast liegt, weil mehr Bauteile dann sicher nicht helfen können.
        """
        best_n = 0
        best_units = pd.DataFrame()
        best_loaded = pd.DataFrame()
        best_summary = pd.DataFrame()
        max_n = len(parts_df)
        platform_label = str(platform_row.get('Pritsche', platform_row.get('Pritschenname', 'Pritsche')))
        max_weight = safe_number(platform_row.get('Max_Gewicht_kg'), 0.0)
        tare = safe_number(platform_row.get('Eigengewicht_Pritsche_kg'), 0.0)
        tests_done = 0
        failed_candidates = 0

        # Streng fortlaufend: nur der nächste Bauteilblock wird getestet.
        # Aber: nicht beim ersten geometrischen Fehler abbrechen.
        for n in range(1, max_n + 1):
            part_block = parts_df.iloc[:n].copy().reset_index(drop=True)
            raw_weight = float(part_block['Gewicht_kg'].sum()) + tare if 'Gewicht_kg' in part_block.columns else 0.0
            if max_weight > 0 and raw_weight > max_weight + 0.001:
                break

            block_units = _build_units_for_part_block(part_block, unit_start, platform_label)
            if block_units.empty:
                break

            tests_done += 1
            ok, loaded, summary = _pack_block_units_on_platform(block_units, platform_row)
            if ok:
                best_n = n
                best_units = block_units
                best_loaded = loaded
                best_summary = summary
            else:
                failed_candidates += 1
                # weiter testen: ein späterer Bauteil kann die Bundgrenze verändern
                # und damit wieder einen gültigen Block ergeben.
                continue

        if best_n > 0:
            if 'Bundbildung' in best_units.columns:
                best_units['Bundbildung'] = best_units['Bundbildung'].astype(str) + f' / Kandidaten geprüft: {tests_done}, verworfen: {failed_candidates}'
            if 'Ebene' in best_loaded.columns:
                best_loaded['Ebene'] = best_loaded['Ebene'].astype(str) + f' / Blocktest {tests_done}'

        return {
            'parts_count': best_n,
            'units': best_units,
            'loaded': best_loaded,
            'summary': best_summary,
            'tests_done': tests_done,
            'failed_candidates': failed_candidates,
        }

    def _simulate_option_by_parts(option_row: pd.Series, remaining_df: pd.DataFrame, fuhre_nr: int, unit_start: int) -> Optional[Dict[str, Any]]:
        option_name = str(option_row['Fuhrenoption'])
        trip_platforms = _prepare_trip_platforms(option_name, fuhre_nr)
        if trip_platforms.empty:
            return None

        cursor = 0
        unit_counter = unit_start
        placement_frames: List[pd.DataFrame] = []
        summary_frames: List[pd.DataFrame] = []
        used_platform_frames: List[pd.DataFrame] = []
        unit_frames: List[pd.DataFrame] = []
        loaded_part_count = 0
        first_label = ''
        last_label = ''

        for _, platform_row in trip_platforms.sort_values('Pritschen_Reihenfolge', kind='stable').iterrows():
            if cursor >= len(remaining_df):
                break
            candidate = _find_max_part_block_for_platform(remaining_df.iloc[cursor:].reset_index(drop=True), platform_row, unit_counter)
            if int(candidate.get('parts_count', 0)) <= 0:
                continue
            n = int(candidate['parts_count'])
            block_parts = remaining_df.iloc[cursor:cursor + n].copy().reset_index(drop=True)
            if not first_label and not block_parts.empty:
                first_label = _make_part_label(block_parts.iloc[0])
            if not block_parts.empty:
                last_label = _make_part_label(block_parts.iloc[-1])

            units_for_platform = candidate['units'].copy()
            unit_counter += len(units_for_platform)
            placement_frames.append(candidate['loaded'])
            summary_frames.append(candidate['summary'])
            used_platform_frames.append(pd.DataFrame([platform_row]))
            unit_frames.append(units_for_platform)
            cursor += n
            loaded_part_count += n

        if loaded_part_count <= 0:
            return None

        loaded_try = pd.concat(placement_frames, ignore_index=True) if placement_frames else pd.DataFrame()
        summary_try = pd.concat(summary_frames, ignore_index=True) if summary_frames else pd.DataFrame()
        platforms_used = pd.concat(used_platform_frames, ignore_index=True) if used_platform_frames else pd.DataFrame()
        units_used = pd.concat(unit_frames, ignore_index=True) if unit_frames else pd.DataFrame()
        loaded_weight = float(loaded_try['Gewicht_kg'].sum()) if not loaded_try.empty else 0.0
        used_platforms = int(loaded_try['Pritsche'].nunique()) if not loaded_try.empty else 0
        priority = safe_number(option_row.get('Priorität'), 999)
        # Priorität: möglichst viele fortlaufende Bauteile, dann Gewicht, dann weniger Pritschen.
        score = (loaded_part_count * 1000000.0) + loaded_weight - (used_platforms * 100.0) - (priority * 10.0)
        return {
            'score': score,
            'option_name': option_name,
            'trip_platforms': platforms_used,
            'summary_try': summary_try,
            'loaded_try': loaded_try,
            'units_used': units_used,
            'parts_loaded_count': loaded_part_count,
            'loaded_weight': loaded_weight,
            'first_label': first_label,
            'last_label': last_label,
            'next_unit_counter': unit_counter,
        }

    if enabled_options.empty:
        not_loaded, units_left = _make_not_loaded_rows_from_parts(remaining_parts, 'keine Fuhrenoption freigegeben', 1)
        return not_loaded, pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), units_left

    fuhre_nr = 1
    unit_counter_global = 1
    while not remaining_parts.empty and fuhre_nr <= max_fuhren:
        best_try = None
        for _, option_row in enabled_options.iterrows():
            attempt = _simulate_option_by_parts(option_row, remaining_parts, fuhre_nr, unit_counter_global)
            if attempt is None:
                continue
            if best_try is None or attempt['score'] > best_try['score']:
                best_try = attempt

        if best_try is None or int(best_try.get('parts_loaded_count', 0)) <= 0:
            break

        all_placements.append(best_try['loaded_try'])
        all_summary.append(best_try['summary_try'])
        all_platforms.append(best_try['trip_platforms'])
        all_units.append(best_try['units_used'])
        unit_counter_global = int(best_try.get('next_unit_counter', unit_counter_global + len(best_try['units_used'])))

        fuhren_log.append({
            'Fuhre_Nr': fuhre_nr,
            'Fuhrenoption': best_try['option_name'],
            'Bauteile': int(best_try['parts_loaded_count']),
            'Verladeeinheiten': int(len(best_try['units_used'])) if best_try['units_used'] is not None else 0,
            'Gewicht_kg': round(float(best_try['loaded_weight']), 2),
            'Pritschen': ', '.join(best_try['trip_platforms']['Pritschenname'].astype(str).tolist()) if not best_try['trip_platforms'].empty else '',
            'Reihenfolge': 'V21 Blockprüfung: Kandidaten testen -> gültigen Max-Block fixieren -> Bund -> unten packen',
            'Erste_Bauteilnummer': best_try.get('first_label', ''),
            'Letzte_Bauteilnummer': best_try.get('last_label', ''),
        })

        remaining_parts = remaining_parts.iloc[int(best_try['parts_loaded_count']):].copy().reset_index(drop=True)
        fuhre_nr += 1

    if not remaining_parts.empty:
        not_loaded, units_left = _make_not_loaded_rows_from_parts(
            remaining_parts,
            'passt in keine freigegebene Fuhrenoption mit V21 geprüfter Blocklogik',
            unit_counter_global,
        )
        if not not_loaded.empty:
            all_placements.append(not_loaded)
        if not units_left.empty:
            all_units.append(units_left)

    placements_df = pd.concat(all_placements, ignore_index=True) if all_placements else pd.DataFrame()
    summary_df = pd.concat(all_summary, ignore_index=True) if all_summary else pd.DataFrame()
    platforms_used_df = pd.concat(all_platforms, ignore_index=True) if all_platforms else pd.DataFrame()
    plan_units_df = pd.concat(all_units, ignore_index=True) if all_units else pd.DataFrame()

    if center_geometric and not placements_df.empty and not platforms_used_df.empty:
        placements_df = center_placements_geometrically(placements_df, platforms_used_df)
        placements_df = center_length_groups_from_platform_center(placements_df, platforms_used_df)
        summary_df = recompute_summary_from_placements(placements_df, platforms_used_df)

    fuhren_log_df = pd.DataFrame(fuhren_log)
    return placements_df, summary_df, platforms_used_df, fuhren_log_df, plan_units_df

def create_loading_excel(
    parts_df: pd.DataFrame,
    units_df: pd.DataFrame,
    placements_df: pd.DataFrame,
    platforms_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    options_df: Optional[pd.DataFrame] = None,
    fuhren_log_df: Optional[pd.DataFrame] = None,
    warnings_df: Optional[pd.DataFrame] = None,
    bsd_header_df: Optional[pd.DataFrame] = None,
    bsd_matrix_df: Optional[pd.DataFrame] = None,
    project_meta: Optional[Dict[str, Any]] = None,
    logo_bytes: Optional[bytes] = None,
) -> bytes:
    """Erstellt die Excel-Begleitdatei.

    Enthalten sind die kleinen optischen BSD-/Pritschenzettel je Pritsche
    sowie die gewünschten Zusatzregister: Bauteile, Fuhrenoptionen,
    Ladeplan_BSD_Kopf und Projektkopf.
    """
    project_meta = project_meta or {}
    front_at_x_max, left_at_y_max = _orientation_flags_from_meta(project_meta)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Gewünschte Zusatzregister für Kontrolle / Weiterverarbeitung.
        wrote_any_sheet = False
        if parts_df is not None and not parts_df.empty:
            parts_df.to_excel(writer, sheet_name='Bauteile', index=False)
            wrote_any_sheet = True
        if options_df is not None and not options_df.empty:
            options_df.to_excel(writer, sheet_name='Fuhrenoptionen', index=False)
            wrote_any_sheet = True
        if bsd_header_df is not None and not bsd_header_df.empty:
            bsd_header_df.to_excel(writer, sheet_name='Ladeplan_BSD_Kopf', index=False)
            wrote_any_sheet = True
        if project_meta:
            pd.DataFrame([{'Feld': k, 'Wert': v} for k, v in project_meta.items()]).to_excel(writer, sheet_name='Projektkopf', index=False)
            wrote_any_sheet = True
        if not wrote_any_sheet:
            pd.DataFrame([{'Hinweis': 'Keine belegte Pritsche / kein Ladeplan erzeugt.'}]).to_excel(writer, sheet_name='Hinweis', index=False)

        # Optische Ladeplan-BSD-Blätter erzeugen.
        try:
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
        except Exception:
            # Falls openpyxl lokal nicht verfügbar ist, bleibt zumindest das Hinweisblatt erhalten.
            return output.getvalue()

        wb = writer.book

        def sheet_safe_name(value: str) -> str:
            cleaned = re.sub(r'[\\/*?:\[\]]', '_', str(value or 'Pritsche')).strip()
            cleaned = cleaned.replace(' ', '_')
            return cleaned[:31] or 'Pritsche'

        def unique_sheet_name(base: str) -> str:
            base = sheet_safe_name(base)
            if base not in wb.sheetnames:
                return base
            for i in range(2, 100):
                suffix = f'_{i}'
                candidate = f'{base[:31-len(suffix)]}{suffix}'
                if candidate not in wb.sheetnames:
                    return candidate
            return base[:28] + '_x'

        def val(row: pd.Series, key: str, default: Any = '') -> Any:
            try:
                v = row.get(key, default)
                if pd.isna(v):
                    return default
                return v
            except Exception:
                return default

        def fmt_mm(value: Any) -> str:
            try:
                return f'{float(value):.0f}'
            except Exception:
                return '0'

        def fmt_t(value_kg: Any) -> str:
            try:
                return f'{float(value_kg) / 1000:.2f}'
            except Exception:
                return '0.00'

        # Farben ähnlich der Vorlage.
        fill_cyan = PatternFill('solid', fgColor='CCFFFF')
        fill_yellow = PatternFill('solid', fgColor='FFFF99')
        fill_gray = PatternFill('solid', fgColor='C0C0C0')
        fill_lightgray = PatternFill('solid', fgColor='E7E6E6')
        fill_white = PatternFill('solid', fgColor='FFFFFF')
        fill_hatch = PatternFill('darkTrellis', fgColor='999999', bgColor='FFFFFF')
        thin = Side(style='thin', color='000000')
        dotted = Side(style='dotted', color='000000')
        border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
        border_dotted = Border(left=dotted, right=dotted, top=dotted, bottom=dotted)
        font_title = Font(name='Arial', size=14, bold=True)
        font_head = Font(name='Arial', size=10, bold=True)
        font_body = Font(name='Arial', size=9)
        font_small = Font(name='Arial', size=8)

        def style_range(ws, cell_range: str, fill=None, border=None, font=None, align=None):
            for row_cells in ws[cell_range]:
                for cell in row_cells:
                    if fill is not None:
                        cell.fill = fill
                    if border is not None:
                        cell.border = border
                    if font is not None:
                        cell.font = font
                    if align is not None:
                        cell.alignment = align

        def set_box(ws, cell_range: str, fill=None):
            style_range(
                ws,
                cell_range,
                fill=fill,
                border=border_thin,
                font=font_body,
                align=Alignment(horizontal='left', vertical='center', wrap_text=True),
            )

        def write_label(ws, cell: str, label: str, bold: bool = True):
            ws[cell] = label
            ws[cell].font = font_head if bold else font_body
            ws[cell].alignment = Alignment(horizontal='left', vertical='center')

        def add_styled_bsd_sheet(header: pd.Series, matrix: pd.DataFrame):
            pname = str(val(header, 'Pritsche', 'Pritsche'))
            ws = wb.create_sheet(unique_sheet_name(f'BSD_{pname}'))
            ws.sheet_view.showGridLines = True
            ws.freeze_panes = 'A13'

            # Spaltenbreiten wie breiter Excel-Ladeplan.
            widths = {
                'A': 10, 'B': 10, 'C': 10, 'D': 10,
                'E': 12, 'F': 12, 'G': 12, 'H': 12,
                'I': 11, 'J': 11, 'K': 12,
                'L': 2, 'M': 14, 'N': 14, 'O': 14, 'P': 14,
            }
            for col, width in widths.items():
                ws.column_dimensions[col].width = width
            for r in range(1, 45):
                ws.row_dimensions[r].height = 18

            # Kopf links.
            ws.merge_cells('A2:B2')
            ws.merge_cells('C2:D2')
            ws.merge_cells('C3:D3')
            set_box(ws, 'A2:D8')
            write_label(ws, 'A2', 'Pritsche:')
            ws['C2'] = pname
            ws['C2'].font = font_title
            ws['C2'].fill = fill_cyan
            ws['C3'] = str(val(header, 'Objekt_Name', '') or val(header, 'Pritschenname', '') or val(header, 'Fuhrenoption', ''))
            ws['C3'].font = font_title
            ws['C3'].fill = fill_yellow
            write_label(ws, 'A4', 'Transport:')
            ws['C4'] = str(val(header, 'Transport_Name', ''))
            ws['C4'].fill = fill_yellow
            write_label(ws, 'A6', 'Decke:')
            ws['C6'] = str(val(header, 'Decke', ''))
            ws['C6'].fill = fill_cyan
            write_label(ws, 'A7', 'Bauabschnitt:')
            ws['C7'] = str(val(header, 'Bauabschnitt', ''))
            ws['C7'].fill = fill_cyan
            write_label(ws, 'A9', 'Sachbearbeiter:')
            ws['C9'] = str(val(header, 'Sachbearbeiter', ''))
            write_label(ws, 'A10', 'Datum:')
            ws['C10'] = str(val(header, 'Datum', datetime.now().strftime('%d.%m.%Y')))
            set_box(ws, 'A9:D10')

            # Kopf Mitte.
            set_box(ws, 'F2:K12')
            write_label(ws, 'F2', 'Unternehmer:')
            ws.merge_cells('I2:K2')
            ws['I2'] = str(val(header, 'Transport_Name', ''))
            ws['I2'].fill = fill_yellow
            write_label(ws, 'F4', 'Pritschenhöhe:')
            ws['I4'] = fmt_mm(val(header, 'Pritschenhöhe_mm', 0))
            ws['I4'].fill = fill_cyan
            ws['J4'] = 'mm'
            write_label(ws, 'F5', 'Pritschenbreite:')
            ws['I5'] = fmt_mm(val(header, 'Pritschenbreite_mm', 0))
            ws['I5'].fill = fill_cyan
            ws['J5'] = 'mm'
            write_label(ws, 'F7', 'Frachthöhe:')
            ws['I7'] = fmt_mm(val(header, 'Frachthöhe_mm', 0)); ws['J7'] = 'mm'
            write_label(ws, 'F8', 'Höhe (gesamt):')
            ws['I8'] = fmt_mm(val(header, 'Höhe_gesamt_mm', 0)); ws['J8'] = 'mm'
            write_label(ws, 'F9', 'Länge (gesamt):')
            ws['I9'] = fmt_mm(val(header, 'Länge_gesamt_mm', 0)); ws['J9'] = 'mm'
            write_label(ws, 'F10', 'Breite (gesamt):')
            ws['I10'] = fmt_mm(val(header, 'Breite_gesamt_mm', 0)); ws['J10'] = 'mm'
            write_label(ws, 'F11', 'Ladegewicht:')
            ws['I11'] = fmt_t(val(header, 'Ladegewicht_kg', 0)); ws['J11'] = 'to'
            write_label(ws, 'F12', 'Gesamtgewicht inkl. Pritsche:')
            ws['I12'] = fmt_t(val(header, 'Gesamtgewicht_kg', 0)); ws['J12'] = 'to'
            set_box(ws, 'F11:K12')

            # Kontrolle rechts.
            set_box(ws, 'M2:P10')
            ws.merge_cells('M2:P2')
            ws['M2'] = 'Kontrolle Root'
            ws['M2'].font = font_title
            write_label(ws, 'M4', 'Datum / Visum:')
            write_label(ws, 'M7', 'Frachthöhe:')
            write_label(ws, 'M8', 'Überhang vorne:')
            write_label(ws, 'M9', 'Länge (gesamt):')
            write_label(ws, 'M10', 'Überhang hinten:')
            write_label(ws, 'M11', 'Überbreite:')
            set_box(ws, 'M11:P11')
            for r in range(7, 12):
                ws[f'P{r}'] = 'mm'

            # Matrix Kopf.
            start_row = 14
            ws.merge_cells(start_row=13, start_column=1, end_row=13, end_column=4)
            ws.cell(13, 1).value = 'Je 2 Elemente nebeneinander angeordnet'
            ws.cell(13, 1).font = font_head
            ws.cell(13, 1).alignment = Alignment(horizontal='center')

            headers = [
                'Vorne links', 'Vorne rechts', 'Hinten links', 'Hinten rechts',
                'Bemerkung\nVorne links', 'Bemerkung\nVorne rechts', 'Bemerkung\nHinten links', 'Bemerkung\nHinten rechts',
                'Höhe (mm)', 'Breite (mm)', 'Gesamtlänge'
            ]
            for idx, h in enumerate(headers, start=1):
                c = ws.cell(start_row, idx)
                c.value = h
                c.font = font_small if idx <= 8 else font_head
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                c.border = border_thin
                c.fill = fill_yellow if idx in [1, 2, 3, 4, 9] else fill_white
            ws.row_dimensions[start_row].height = 28

            # Etwas leere Rasterfläche oberhalb der tatsächlichen Lagen wie in Vorlage.
            data_row = start_row + 1
            max_rows = max(18, len(matrix) + 8)
            for r in range(data_row, data_row + max_rows):
                for c in range(1, 12):
                    ws.cell(r, c).border = border_dotted
                    ws.cell(r, c).font = font_small
                    ws.cell(r, c).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    if c <= 4:
                        ws.cell(r, c).fill = fill_yellow
                    elif c <= 8:
                        ws.cell(r, c).fill = fill_white
                    else:
                        ws.cell(r, c).fill = fill_white

            # Matrixdaten an den unteren Teil setzen, dadurch ähnelt es dem Beispiel optisch.
            write_start = data_row + max(0, max_rows - len(matrix) - 1)
            if matrix is not None and not matrix.empty:
                for i, (_, mrow) in enumerate(matrix.iterrows(), start=write_start):
                    row_values = [
                        mrow.get('Vorne links', ''), mrow.get('Vorne rechts', ''),
                        mrow.get('Hinten links', ''), mrow.get('Hinten rechts', ''),
                        mrow.get('Bemerkung vorne links', ''), mrow.get('Bemerkung vorne rechts', ''),
                        mrow.get('Bemerkung hinten links', ''), mrow.get('Bemerkung hinten rechts', ''),
                        fmt_mm(mrow.get('Höhe_mm', 0)), fmt_mm(mrow.get('Breite_mm', 0)), fmt_mm(mrow.get('Gesamtlänge_mm', 0)),
                    ]
                    for c, value in enumerate(row_values, start=1):
                        cell = ws.cell(i, c)
                        cell.value = value
                        cell.font = Font(name='Arial', size=8, bold=(c <= 4))
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = border_dotted
                    ws.row_dimensions[i].height = 22

            # Ladehöhe/Statuszeile.
            footer_row = data_row + max_rows
            ws.cell(footer_row, 1).value = 'Ladehöhe:'
            ws.cell(footer_row, 1).font = font_head
            for c in range(2, 5):
                ws.cell(footer_row, c).value = fmt_mm(val(header, 'Frachthöhe_mm', 0))
            ws.merge_cells(start_row=footer_row, start_column=5, end_row=footer_row, end_column=8)
            ws.cell(footer_row, 5).value = f"Gesamtgewicht ca.: {fmt_t(val(header, 'Gesamtgewicht_kg', val(header, 'Ladegewicht_kg', 0)))} Tonnen"
            ws.cell(footer_row, 5).font = font_head
            ws.cell(footer_row, 9).value = fmt_mm(val(header, 'Höhe_gesamt_mm', 0))
            ws.cell(footer_row, 10).value = fmt_mm(val(header, 'Breite_gesamt_mm', 0))
            ws.cell(footer_row, 11).value = fmt_mm(val(header, 'Länge_gesamt_mm', 0))
            style_range(ws, f'A{footer_row}:K{footer_row}', fill=fill_gray, border=border_thin, font=font_small, align=Alignment(horizontal='center'))

            # Graue Ladehöhen-Kästchen.
            box_top = footer_row + 1
            for c1, c2 in [(2, 2), (9, 9)]:
                ws.merge_cells(start_row=box_top, start_column=c1, end_row=box_top + 2, end_column=c2)
                cell = ws.cell(box_top, c1)
                cell.value = fmt_mm(val(header, 'Frachthöhe_mm', 0))
                cell.fill = fill_gray
                cell.border = border_thin
                cell.alignment = Alignment(horizontal='center', vertical='top')

            # Untere Bereiche.
            lower_top = footer_row + 4
            ws.merge_cells(start_row=lower_top, start_column=1, end_row=lower_top, end_column=4)
            ws.cell(lower_top, 1).value = 'Zusätzliches Verlade-Material:'
            ws.cell(lower_top, 1).font = font_head
            ws.merge_cells(start_row=lower_top, start_column=6, end_row=lower_top, end_column=11)
            ws.cell(lower_top, 6).value = 'Bemerkungen:'
            ws.cell(lower_top, 6).font = font_head
            for r in range(lower_top, lower_top + 7):
                for c in range(1, 12):
                    ws.cell(r, c).border = border_thin if r == lower_top else border_dotted

            # Gewichtsetikette rechts unten.
            ws.merge_cells(start_row=15, start_column=13, end_row=footer_row + 3, end_column=16)
            ws.cell(15, 13).fill = fill_hatch
            ws.cell(15, 13).border = border_thin
            ws.merge_cells(start_row=footer_row + 4, start_column=13, end_row=footer_row + 4, end_column=16)
            ws.cell(footer_row + 4, 13).value = 'Gewichts-Etikette:'
            ws.cell(footer_row + 4, 13).font = font_head
            ws.cell(footer_row + 4, 13).alignment = Alignment(horizontal='center')
            ws.merge_cells(start_row=footer_row + 5, start_column=13, end_row=lower_top + 6, end_column=16)
            ws.cell(footer_row + 5, 13).fill = fill_hatch
            ws.cell(footer_row + 5, 13).border = border_thin

            # Logo im Pritschenzettel einfügen, falls vorhanden.
            if logo_bytes:
                try:
                    from openpyxl.drawing.image import Image as XLImage
                    if not hasattr(wb, '_verladeplan_excel_logo_buffers'):
                        wb._verladeplan_excel_logo_buffers = []
                    img_buf = io.BytesIO(logo_bytes)
                    wb._verladeplan_excel_logo_buffers.append(img_buf)
                    img = XLImage(img_buf)
                    img.width = 120
                    img.height = 85
                    ws.add_image(img, f'M{max(15, footer_row - 6)}')
                except Exception:
                    pass

            # Rahmen / allgemeine Formatierung.
            for row in ws.iter_rows(min_row=1, max_row=lower_top + 6, min_col=1, max_col=16):
                for cell in row:
                    if cell.font == Font():
                        cell.font = font_body
            ws.page_setup.orientation = 'landscape'
            ws.page_setup.paperSize = ws.PAPERSIZE_A3
            ws.page_margins.left = 0.25
            ws.page_margins.right = 0.25
            ws.page_margins.top = 0.25
            ws.page_margins.bottom = 0.25
            ws.print_area = f'A1:P{lower_top + 6}'
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1


        def add_visual_view_sheet(header: pd.Series):
            """Erzeugt pro Pritsche ein zusätzliches Excel-Blatt mit grafischen Ansichten.

            Dieses Blatt ersetzt die bisherige PDF-Ansicht in der Hauptausgabe:
            Seitenansicht, Rückansicht, Draufsicht und Vorderansicht werden als
            Bilder in Excel eingebettet. Der kleine BSD-/Pritschenzettel bleibt
            zusätzlich als eigenes Blatt erhalten.
            """
            pname = str(val(header, 'Pritsche', 'Pritsche'))
            ws = wb.create_sheet(unique_sheet_name(f'Ansicht_{pname}'))
            ws.sheet_view.showGridLines = False

            # Layout A3 Querformat.
            for col_idx in range(1, 19):
                ws.column_dimensions[get_column_letter(col_idx)].width = 10
            for row_idx in range(1, 67):
                ws.row_dimensions[row_idx].height = 18

            ws.merge_cells('A1:R1')
            ws['A1'] = f'Pritschenplan / Ansichten - {pname}'
            ws['A1'].font = Font(name='Arial', size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A1'].fill = fill_lightgray
            style_range(ws, 'A1:R1', border=border_thin)

            # Linker Kopfbereich.
            set_box(ws, 'A3:F10')
            write_label(ws, 'A3', 'Objekt:')
            ws.merge_cells('C3:F3')
            ws['C3'] = str(val(header, 'Objekt_Name', ''))
            write_label(ws, 'A4', 'Transport:')
            ws.merge_cells('C4:F4')
            ws['C4'] = str(val(header, 'Transport_Name', '') or val(header, 'Fuhrenoption', ''))
            write_label(ws, 'A5', 'Pritsche:')
            ws.merge_cells('C5:F5')
            ws['C5'] = pname
            ws['C5'].fill = fill_cyan
            write_label(ws, 'A6', 'Decke:')
            ws.merge_cells('C6:F6')
            ws['C6'] = str(val(header, 'Decke', ''))
            write_label(ws, 'A7', 'Bauabschnitt:')
            ws.merge_cells('C7:F7')
            ws['C7'] = str(val(header, 'Bauabschnitt', ''))
            write_label(ws, 'A8', 'Sachbearbeiter:')
            ws.merge_cells('C8:F8')
            ws['C8'] = str(val(header, 'Sachbearbeiter', ''))
            write_label(ws, 'A9', 'Datum:')
            ws.merge_cells('C9:F9')
            ws['C9'] = str(val(header, 'Datum', datetime.now().strftime('%d.%m.%Y')))

            # Rechter Infoblock ähnlich PDF-Vorlage.
            set_box(ws, 'H3:R10')
            ws.merge_cells('H3:R3')
            ws['H3'] = 'Info Pritsche'
            ws['H3'].font = font_title
            ws['H3'].alignment = Alignment(horizontal='center')
            fields = [
                ('H4', 'Gewicht:', 'K4', fmt_t(val(header, 'Ladegewicht_kg', 0)), 'L4', 'to'),
                ('H5', 'Breite:', 'K5', fmt_mm(val(header, 'Breite_gesamt_mm', 0)), 'L5', 'mm'),
                ('H6', 'Höhe:', 'K6', fmt_mm(val(header, 'Frachthöhe_mm', 0)), 'L6', 'mm'),
                ('H7', 'Länge:', 'K7', fmt_mm(val(header, 'Länge_gesamt_mm', 0)), 'L7', 'mm'),
                ('H8', 'Max. Breite:', 'K8', fmt_mm(val(header, 'Pritschenbreite_mm', 0)), 'L8', 'mm'),
                ('H9', 'Max. Ladehöhe:', 'K9', fmt_mm(val(header, 'Pritschenhöhe_mm', 0)), 'L9', 'mm'),
                ('H10', 'Warnungen:', 'K10', str(val(header, 'Warnungen', 0)), 'L10', ''),
            ]
            for lab_cell, lab, val_cell, value, unit_cell, unit in fields:
                write_label(ws, lab_cell, lab)
                ws[val_cell] = value
                ws[val_cell].font = font_head
                ws[val_cell].alignment = Alignment(horizontal='right')
                ws[unit_cell] = unit

            # Plattformdaten und Verladedaten.
            platform_match = platforms_df[platforms_df['Pritsche'].astype(str) == pname] if platforms_df is not None and not platforms_df.empty else pd.DataFrame()
            if platform_match.empty:
                ws['A13'] = 'Keine Pritschendaten für grafische Ansicht vorhanden.'
                return
            platform_row = platform_match.iloc[0]
            loaded = placements_df[placements_df['Pritsche'].astype(str) == pname].copy() if placements_df is not None and not placements_df.empty else pd.DataFrame()
            if not loaded.empty:
                loaded = loaded[loaded['X_mm'].notna() & loaded['Y_mm'].notna() & loaded['Z_mm'].notna()].copy()

            def make_view_png(view: str, title: str) -> Optional[io.BytesIO]:
                try:
                    import matplotlib
                    matplotlib.use('Agg')
                    import matplotlib.pyplot as plt
                    from matplotlib.patches import Rectangle
                except Exception:
                    return None

                eff_length = safe_number(platform_row.get('Länge_mm')) + safe_number(platform_row.get('Überhang_vorne_mm')) + safe_number(platform_row.get('Überhang_hinten_mm'))
                width = safe_number(platform_row.get('Breite_mm'))
                max_height = safe_number(platform_row.get('Max_Höhe_mm'))
                base_wood = safe_number(platform_row.get('Kantholz_erste_Lage_mm'))
                used_length = safe_number(val(header, 'Länge_gesamt_mm', 0))
                used_width = safe_number(val(header, 'Breite_gesamt_mm', 0))
                used_height = safe_number(val(header, 'Frachthöhe_mm', 0))

                fig, ax = plt.subplots(figsize=(7.1, 3.0), dpi=160)
                ax.set_title(title, fontsize=10, fontweight='bold', pad=6)
                ax.grid(True, linewidth=0.25, alpha=0.35)
                ax.tick_params(labelsize=7)
                ax.set_facecolor('#fbfbfb')

                if view == 'top':
                    ax.set_xlabel('Länge X (mm)', fontsize=7)
                    ax.set_ylabel('Breite Y (mm)', fontsize=7)
                    ax.add_patch(Rectangle((0, 0), eff_length, width, fill=False, edgecolor='black', linestyle='--', linewidth=1.5))
                    for _, row in loaded.iterrows():
                        x0 = safe_number(row.get('X_mm'))
                        y0 = safe_number(row.get('Y_mm'))
                        lx = safe_number(row.get('Länge_mm'))
                        by = safe_number(row.get('Breite_mm'))
                        ax.add_patch(Rectangle((x0, y0), lx, by, facecolor='#c9d6df', edgecolor='black', linewidth=0.5))
                        label = _view_label(row).replace('<br>', '\n')
                        ax.text(x0 + lx / 2, y0 + by / 2, label, ha='center', va='center', fontsize=5.5)
                    ax.set_xlim(0, max(eff_length, 1))
                    ax.set_ylim(0, max(width, 1))

                elif view == 'side':
                    ax.set_xlabel('Länge X (mm)', fontsize=7)
                    ax.set_ylabel('Höhe Z (mm)', fontsize=7)
                    ax.add_patch(Rectangle((0, 0), eff_length, max_height, fill=False, edgecolor='black', linestyle='--', linewidth=1.5))
                    ax.plot([0, eff_length], [0, 0], color='black', linewidth=2)
                    if base_wood > 0:
                        ax.plot([0, eff_length], [base_wood, base_wood], color='black', linestyle=':', linewidth=0.8)
                        ax.text(eff_length * 0.01, base_wood + 8, f'Kantholz {base_wood:.0f}', fontsize=6, va='bottom')
                    for _, row in loaded.iterrows():
                        x0 = safe_number(row.get('X_mm'))
                        z0 = safe_number(row.get('Z_mm'))
                        lx = safe_number(row.get('Länge_mm'))
                        hz = safe_number(row.get('Höhe_mm'))
                        ax.add_patch(Rectangle((x0, z0), lx, hz, facecolor='#c9d6df', edgecolor='black', linewidth=0.5))
                        label = _view_label(row).replace('<br>', '\n')
                        ax.text(x0 + lx / 2, z0 + hz / 2, label, ha='center', va='center', fontsize=5.5)
                    ax.set_xlim(0, max(eff_length, 1))
                    ax.set_ylim(0, max(max_height, used_height, 1))

                elif view in ('back', 'front'):
                    ax.set_xlabel('Breite Y (mm)', fontsize=7)
                    ax.set_ylabel('Höhe Z (mm)', fontsize=7)
                    ax.add_patch(Rectangle((0, 0), width, max_height, fill=False, edgecolor='black', linestyle='--', linewidth=1.5))
                    ax.plot([0, width], [0, 0], color='black', linewidth=2)
                    if base_wood > 0:
                        ax.plot([0, width], [base_wood, base_wood], color='black', linestyle=':', linewidth=0.8)
                    for _, row in loaded.iterrows():
                        y = safe_number(row.get('Y_mm'))
                        z0 = safe_number(row.get('Z_mm'))
                        by = safe_number(row.get('Breite_mm'))
                        hz = safe_number(row.get('Höhe_mm'))
                        x0 = width - y - by if view == 'front' else y
                        ax.add_patch(Rectangle((x0, z0), by, hz, facecolor='#c9d6df', edgecolor='black', linewidth=0.5))
                        label = _view_label(row).replace('<br>', '\n')
                        ax.text(x0 + by / 2, z0 + hz / 2, label, ha='center', va='center', fontsize=5.5)
                    ax.set_xlim(0, max(width, 1))
                    ax.set_ylim(0, max(max_height, used_height, 1))
                    if view == 'front':
                        ax.text(width * 0.5, max(max_height, used_height, 1) * 0.98, 'Vorne', ha='center', va='top', fontsize=7)
                    else:
                        ax.text(width * 0.5, max(max_height, used_height, 1) * 0.98, 'Hinten', ha='center', va='top', fontsize=7)

                dim_text = f'Belegt: L {used_length:.0f} mm / B {used_width:.0f} mm / H {used_height:.0f} mm'
                ax.text(0.01, -0.18, dim_text, transform=ax.transAxes, fontsize=7, ha='left', va='top')
                fig.tight_layout(pad=1.0)
                buf = io.BytesIO()
                fig.savefig(buf, format='png', bbox_inches='tight')
                plt.close(fig)
                buf.seek(0)
                return buf

            try:
                from openpyxl.drawing.image import Image as XLImage
            except Exception:
                ws['A13'] = 'Grafische Excel-Ansichten konnten nicht eingebettet werden, weil Pillow fehlt.'
                ws['A14'] = 'Installieren: pip install pillow matplotlib'
                return

            if not hasattr(wb, '_verladeplan_image_buffers'):
                wb._verladeplan_image_buffers = []

            view_defs = [
                ('B13', 'B14', 'side', 'Seitenansicht'),
                ('K13', 'K14', 'back', 'Rückansicht'),
                ('B36', 'B37', 'top', 'Draufsicht'),
                ('K36', 'K37', 'front', 'Vorderansicht'),
            ]
            for title_anchor, image_anchor, view, title in view_defs:
                ws[title_anchor] = title
                ws[title_anchor].font = font_head
                ws[title_anchor].alignment = Alignment(horizontal='center')
                img_buf = make_view_png(view, f'{title} - {pname}')
                if img_buf is None:
                    ws[image_anchor] = 'Grafik nicht erzeugt. Installieren: pip install matplotlib pillow'
                    continue
                wb._verladeplan_image_buffers.append(img_buf)
                img = XLImage(img_buf)
                img.width = 560
                img.height = 240
                ws.add_image(img, image_anchor)

            # Kurze Legende / Hinweise.
            ws.merge_cells('A61:R62')
            ws['A61'] = 'Hinweis: Dieses Blatt enthält die grafischen Verladeansichten für die Pritsche. Die kleinen Ladeplan-BSD/Pritschenzettel bleiben als separate BSD-Blätter erhalten.'
            ws['A61'].font = font_small
            ws['A61'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            ws.page_setup.orientation = 'landscape'
            ws.page_setup.paperSize = ws.PAPERSIZE_A3
            ws.page_margins.left = 0.25
            ws.page_margins.right = 0.25
            ws.page_margins.top = 0.25
            ws.page_margins.bottom = 0.25
            ws.print_area = 'A1:R62'
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1

        if bsd_header_df is not None and not bsd_header_df.empty and bsd_matrix_df is not None and not bsd_matrix_df.empty:
            for _, header in bsd_header_df.iterrows():
                pname = str(val(header, 'Pritsche', ''))
                matrix = bsd_matrix_df[bsd_matrix_df['Pritsche'].astype(str) == pname].copy()
                add_styled_bsd_sheet(header, matrix)

        if 'Hinweis' in wb.sheetnames and len(wb.sheetnames) > 1:
            wb.remove(wb['Hinweis'])

        # Verbleibende Blätter etwas lesbarer machen.
        for ws in wb.worksheets:
            if ws.title.startswith('BSD_') or ws.title.startswith('Ansicht_'):
                continue
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col[:50]:
                    try:
                        max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)
            if ws.max_row >= 1:
                for cell in ws[1]:
                    cell.font = font_head
                    cell.fill = fill_lightgray
                    cell.border = border_thin

    return output.getvalue()


def _view_label(row: pd.Series) -> str:
    label = _format_label_value(row.get('Ansicht_Label'))
    if not label:
        label = _format_label_value(row.get('Bauteile'))
    if not label:
        label = _format_label_value(row.get('Einheit_ID'))
    if str(row.get('Typ', '')).strip() == 'Bund':
        liste = _split_bsd_text_list(row.get('Ansicht_Liste', ''), label)
        if liste:
            # Bunde in der App ebenfalls untereinander darstellen.
            return '<br>'.join(str(v) for v in liste)
        return str(label).replace(', ', '<br>')
    return str(label)


def draw_loading_view(placements_df: pd.DataFrame, platforms_df: pd.DataFrame, platform_name: str, view: str) -> go.Figure:
    """Zeichnet Draufsicht, Seitenansicht oder Rückansicht der ausgewählten Pritsche.

    Wichtig: Plotly skaliert Achsen nicht zuverlässig nur anhand von Shapes.
    Deshalb werden die Achsbereiche hier fest gesetzt, damit die Rechtecke sichtbar sind.
    """
    platform_match = platforms_df[platforms_df['Pritsche'].astype(str) == str(platform_name)]
    fig = go.Figure()

    title_prefix = {
        'top': 'Draufsicht',
        'side': 'Seitenansicht',
        'back': 'Rückansicht',
        'front': 'Vorderansicht',
    }.get(view, 'Ansicht')

    if platform_match.empty:
        fig.update_layout(
            title=f'{title_prefix} - keine Pritsche gefunden',
            height=450,
            margin=dict(l=20, r=20, t=50, b=20),
        )
        fig.add_annotation(text='Keine Pritschendaten vorhanden', x=0.5, y=0.5, xref='paper', yref='paper', showarrow=False)
        return fig

    platform_row = platform_match.iloc[0]
    base_length = float(platform_row['Länge_mm'])
    over_front = float(platform_row.get('Überhang_vorne_mm', 0) or 0)
    over_back = float(platform_row.get('Überhang_hinten_mm', 0) or 0)
    eff_length = base_length + over_front + over_back
    platform_x0 = over_back
    platform_x1 = over_back + base_length
    width = float(platform_row['Breite_mm'])
    max_height = float(platform_row['Max_Höhe_mm'])
    base_wood = float(platform_row.get('Kantholz_erste_Lage_mm', 0) or 0)

    loaded = placements_df[placements_df['Pritsche'].astype(str) == str(platform_name)].copy()
    loaded = loaded[loaded['X_mm'].notna() & loaded['Y_mm'].notna() & loaded['Z_mm'].notna()].copy()

    # Unsichtbarer Punkt erzwingt, dass Plotly die Grafikfläche überhaupt aufspannt.
    fig.add_trace(go.Scatter(x=[0], y=[0], mode='markers', marker=dict(size=1, opacity=0), hoverinfo='skip', showlegend=False))

    if view == 'top':
        fig.update_layout(
            title=f'{title_prefix} - {platform_name}',
            xaxis_title='Länge X (mm)',
            yaxis_title='Breite Y (mm)',
        )
        # Zulässiger Ladebereich inkl. Überstand = gestrichelt.
        fig.add_shape(type='rect', x0=0, y0=0, x1=eff_length, y1=width, line=dict(width=2, dash='dash'))
        # Echte Pritsche ohne Überstand = durchgezogen.
        fig.add_shape(type='rect', x0=platform_x0, y0=0, x1=platform_x1, y1=width, line=dict(width=2))
        if over_back > 0:
            fig.add_annotation(x=over_back / 2, y=width * 0.05, text='Überstand hinten', showarrow=False, font=dict(size=9))
        if over_front > 0:
            fig.add_annotation(x=platform_x1 + over_front / 2, y=width * 0.05, text='Überstand vorne', showarrow=False, font=dict(size=9))
        for _, row in loaded.iterrows():
            x0, y0 = float(row['X_mm']), float(row['Y_mm'])
            x1, y1 = x0 + float(row['Länge_mm']), y0 + float(row['Breite_mm'])
            fig.add_shape(type='rect', x0=x0, y0=y0, x1=x1, y1=y1, line=dict(width=1), fillcolor='rgba(100,100,100,0.28)')
            fig.add_annotation(x=(x0 + x1) / 2, y=(y0 + y1) / 2, text=_view_label(row), showarrow=False, font=dict(size=10))
        fig.update_xaxes(range=[0, max(eff_length, 1)], constrain='domain')
        fig.update_yaxes(range=[0, max(width, 1)], scaleanchor='x', scaleratio=1)

    elif view == 'side':
        fig.update_layout(
            title=f'{title_prefix} - {platform_name}',
            xaxis_title='Länge X (mm)',
            yaxis_title='Höhe Z (mm)',
        )
        # Zulässiger Ladebereich inkl. Überstand = gestrichelt.
        fig.add_shape(type='rect', x0=0, y0=0, x1=eff_length, y1=max_height, line=dict(width=2, dash='dash'))
        # Echte Pritsche ohne Überstand = durchgezogen.
        fig.add_shape(type='rect', x0=platform_x0, y0=0, x1=platform_x1, y1=max_height, line=dict(width=2))
        if over_back > 0:
            fig.add_annotation(x=over_back / 2, y=max_height * 0.06, text='Überstand hinten', showarrow=False, font=dict(size=9))
        if over_front > 0:
            fig.add_annotation(x=platform_x1 + over_front / 2, y=max_height * 0.06, text='Überstand vorne', showarrow=False, font=dict(size=9))
        if base_wood > 0:
            fig.add_shape(type='line', x0=0, y0=base_wood, x1=eff_length, y1=base_wood, line=dict(width=1, dash='dot'))
        for _, row in loaded.iterrows():
            x0, z0 = float(row['X_mm']), float(row['Z_mm'])
            x1, z1 = x0 + float(row['Länge_mm']), z0 + float(row['Höhe_mm'])
            fig.add_shape(type='rect', x0=x0, y0=z0, x1=x1, y1=z1, line=dict(width=1), fillcolor='rgba(100,100,100,0.28)')
            fig.add_annotation(x=(x0 + x1) / 2, y=(z0 + z1) / 2, text=_view_label(row), showarrow=False, font=dict(size=10))
        fig.update_xaxes(range=[0, max(eff_length, 1)], constrain='domain')
        fig.update_yaxes(range=[0, max(max_height, 1)])

    elif view == 'back':
        fig.update_layout(
            title=f'{title_prefix} - {platform_name}',
            xaxis_title='Breite Y (mm)',
            yaxis_title='Höhe Z (mm)',
        )
        fig.add_shape(type='rect', x0=0, y0=0, x1=width, y1=max_height, line=dict(width=2, dash='dash'))
        if base_wood > 0:
            fig.add_shape(type='line', x0=0, y0=base_wood, x1=width, y1=base_wood, line=dict(width=1, dash='dot'))
        for _, row in loaded.iterrows():
            y0, z0 = float(row['Y_mm']), float(row['Z_mm'])
            y1, z1 = y0 + float(row['Breite_mm']), z0 + float(row['Höhe_mm'])
            fig.add_shape(type='rect', x0=y0, y0=z0, x1=y1, y1=z1, line=dict(width=1), fillcolor='rgba(100,100,100,0.28)')
            fig.add_annotation(x=(y0 + y1) / 2, y=(z0 + z1) / 2, text=_view_label(row), showarrow=False, font=dict(size=10))
        fig.update_xaxes(range=[0, max(width, 1)], constrain='domain')
        fig.update_yaxes(range=[0, max(max_height, 1)])

    else:
        # Vorderansicht: gleiche Projektion wie Rückansicht, aber links/rechts gespiegelt.
        fig.update_layout(
            title=f'{title_prefix} - {platform_name}',
            xaxis_title='Breite Y gespiegelt (mm)',
            yaxis_title='Höhe Z (mm)',
        )
        fig.add_shape(type='rect', x0=0, y0=0, x1=width, y1=max_height, line=dict(width=2, dash='dash'))
        if base_wood > 0:
            fig.add_shape(type='line', x0=0, y0=base_wood, x1=width, y1=base_wood, line=dict(width=1, dash='dot'))
        for _, row in loaded.iterrows():
            y0_raw, z0 = float(row['Y_mm']), float(row['Z_mm'])
            b = float(row['Breite_mm'])
            y0 = width - y0_raw - b
            y1, z1 = y0 + b, z0 + float(row['Höhe_mm'])
            fig.add_shape(type='rect', x0=y0, y0=z0, x1=y1, y1=z1, line=dict(width=1), fillcolor='rgba(100,100,100,0.28)')
            fig.add_annotation(x=(y0 + y1) / 2, y=(z0 + z1) / 2, text=_view_label(row), showarrow=False, font=dict(size=10))
        fig.update_xaxes(range=[0, max(width, 1)], constrain='domain')
        fig.update_yaxes(range=[0, max(max_height, 1)])

    if loaded.empty:
        fig.add_annotation(
            text='Auf dieser Pritsche sind keine Einheiten platziert',
            x=0.5,
            y=0.5,
            xref='paper',
            yref='paper',
            showarrow=False,
            font=dict(size=14),
        )

    fig.update_layout(height=450, showlegend=False, margin=dict(l=20, r=20, t=50, b=20))
    return fig



def clean_placements_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Konvertiert manuell bearbeitete Platzierungswerte wieder in saubere Zahlen."""
    result = df.copy()
    numeric_cols = ['Fuhre_Nr', 'Anzahl_Bauteile', 'X_mm', 'Y_mm', 'Z_mm', 'Länge_mm', 'Breite_mm', 'Höhe_mm', 'Drehung', 'Gewicht_kg', 'Einlage_unten_mm', 'Einlage_oben_mm', 'Z_manuell_mm', 'Manuelle_Reihenfolge']
    for col in numeric_cols:
        if col in result.columns:
            result[col] = pd.to_numeric(result[col], errors='coerce')
    if 'Pritsche' in result.columns:
        result['Pritsche'] = result['Pritsche'].fillna('').astype(str)
    if 'Einheit_ID' in result.columns:
        result['Einheit_ID'] = result['Einheit_ID'].fillna('').astype(str)
    return result




def _rect_overlap_area(ax: float, ay: float, aw: float, ah: float, bx: float, by: float, bw: float, bh: float) -> float:
    """Überlappungsfläche zweier Rechtecke in der Pritschen-Draufsicht."""
    x0 = max(ax, bx)
    y0 = max(ay, by)
    x1 = min(ax + aw, bx + bw)
    y1 = min(ay + ah, by + bh)
    if x1 <= x0 or y1 <= y0:
        return 0.0
    return float((x1 - x0) * (y1 - y0))


def _platform_row_for_name(platforms_df: pd.DataFrame, pname: str) -> Optional[pd.Series]:
    if platforms_df is None or platforms_df.empty:
        return None
    match = platforms_df[platforms_df['Pritsche'].astype(str) == str(pname)]
    if match.empty:
        return None
    return match.iloc[0]


def calculate_underbau_rows_for_platform(
    placements_df: pd.DataFrame,
    platform: pd.Series,
    min_support_ratio: float = 0.65,
    min_underbau_height: float = 20.0,
    tolerance: float = 1.0,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Ermittelt notwendigen Unterbau / Aufdopplung für eine Pritsche.

    Die Logik ist bewusst als Verlade-Hinweis aufgebaut:
    - Wenn über einer darunterliegenden Lage eine grössere freie Höhe entsteht,
      wird ein Unterbau-Eintrag erzeugt.
    - Wenn die Auflagefläche zu klein ist, wird zusätzlich eine Warnung erzeugt.
    - Es wird keine neue Verladung gerechnet und die Bauteile werden nicht verschoben.
    """
    columns = list(placements_df.columns) if placements_df is not None and not placements_df.empty else []
    warning_cols = ['Typ', 'Pritsche', 'Einheit_ID', 'Warnung', 'Details']
    if placements_df is None or placements_df.empty or platform is None:
        return pd.DataFrame(columns=columns), pd.DataFrame(columns=warning_cols)

    pname = str(platform.get('Pritsche', ''))
    rows = placements_df[placements_df['Pritsche'].astype(str) == pname].copy()
    rows = rows[rows['X_mm'].notna() & rows['Y_mm'].notna() & rows['Z_mm'].notna()].copy()
    if rows.empty:
        return pd.DataFrame(columns=columns), pd.DataFrame(columns=warning_cols)

    helper_types = {'Unterbau', 'Kantholz', 'Bundeinlage', 'Einlage', 'Lagenholz'}
    rows = rows[~rows.get('Typ', pd.Series(dtype=str)).astype(str).isin(helper_types)].copy()
    if rows.empty:
        return pd.DataFrame(columns=columns), pd.DataFrame(columns=warning_cols)

    for col in ['X_mm', 'Y_mm', 'Z_mm', 'Länge_mm', 'Breite_mm', 'Höhe_mm']:
        rows[col] = pd.to_numeric(rows[col], errors='coerce').fillna(0.0)

    base_height = safe_number(platform.get('Kantholz_erste_Lage_mm'), 0.0)
    bundle_spacer = safe_number(platform.get('Einlage_zwischen_Lagen_mm'), 0.0)
    general_spacer = safe_number(platform.get('Einlage_allgemein_mm'), 0.0)
    min_support_ratio = max(0.0, min(1.0, float(min_support_ratio)))
    min_underbau_height = max(0.0, float(min_underbau_height))

    underbau_rows: List[Dict[str, Any]] = []
    warnings: List[Dict[str, Any]] = []
    existing_keys = set()

    # Nur Bauteile/Bunde oberhalb der ersten Ebene prüfen.
    check_rows = rows[rows['Z_mm'] > base_height + tolerance].copy()
    for _, row in check_rows.sort_values(['Z_mm', 'X_mm', 'Y_mm'], kind='stable').iterrows():
        x = safe_number(row.get('X_mm'))
        y = safe_number(row.get('Y_mm'))
        z = safe_number(row.get('Z_mm'))
        lx = safe_number(row.get('Länge_mm'))
        by = safe_number(row.get('Breite_mm'))
        footprint = max(1.0, lx * by)
        row_typ = str(row.get('Typ', '') or '').strip()
        expected_spacer = bundle_spacer if row_typ == 'Bund' else general_spacer

        below = rows[(rows['Z_mm'] + rows['Höhe_mm']) <= z - tolerance].copy()
        overlap_area = 0.0
        closest_top = base_height
        overlap_rows = []
        for _, b in below.iterrows():
            ov = _rect_overlap_area(x, y, lx, by, safe_number(b.get('X_mm')), safe_number(b.get('Y_mm')), safe_number(b.get('Länge_mm')), safe_number(b.get('Breite_mm')))
            if ov <= 0:
                continue
            overlap_area += ov
            top_b = safe_number(b.get('Z_mm')) + safe_number(b.get('Höhe_mm'))
            closest_top = max(closest_top, top_b)
            overlap_rows.append((b, ov, top_b))
        support_ratio = min(1.0, overlap_area / footprint)

        # Unterbau nur dann erzwingen, wenn auf der höchsten realen Auflageebene
        # nicht genug tragende Fläche vorhanden ist. Das vermeidet falschen Unterbau
        # bei Teilen/Bunden, die über die Y-Mitte spannen oder links/rechts real aufliegen.
        top_overlap_area = sum(ov for _b, ov, top_b in overlap_rows if abs(top_b - closest_top) <= tolerance + 1e-6)
        top_support_ratio = min(1.0, top_overlap_area / footprint) if footprint > 0 else 0.0

        free_gap = max(0.0, z - closest_top - max(0.0, expected_spacer))
        einheit_id = str(row.get('Einheit_ID', ''))
        if support_ratio + 1e-6 < min_support_ratio:
            warnings.append({
                'Typ': 'Auflage',
                'Pritsche': pname,
                'Einheit_ID': einheit_id,
                'Warnung': 'Auflagefläche prüfen',
                'Details': f'Auflage ca. {support_ratio*100:.0f}% < {min_support_ratio*100:.0f}% bei Z={z:.0f} mm',
            })

        if free_gap >= min_underbau_height and top_support_ratio + 1e-6 < min_support_ratio:
            ub_h = round(free_gap, 1)
            ub_z = round(z - ub_h, 1)
            key = (round(x, 1), round(y, 1), round(ub_z, 1), round(lx, 1), round(by, 1), round(ub_h, 1))
            if key not in existing_keys:
                existing_keys.add(key)
                base = {col: '' for col in columns}
                for col in columns:
                    try:
                        base[col] = row.get(col, '')
                    except Exception:
                        pass
                label = _fmt_bsd_mm_label('Unterbau', ub_h)
                base.update({
                    'Typ': 'Unterbau',
                    'Einheit_ID': f'UB_{einheit_id}' if einheit_id else 'UB',
                    'Bauteile': label,
                    'Bauteile_Liste': label,
                    'Ansicht_Label': label,
                    'Ansicht_Liste': label,
                    'Anzahl_Bauteile': 0,
                    'X_mm': x,
                    'Y_mm': y,
                    'Z_mm': ub_z,
                    'Länge_mm': lx,
                    'Breite_mm': by,
                    'Höhe_mm': ub_h,
                    'Gewicht_kg': 0.0,
                    'Ebene': 'Unterbau automatisch',
                })
                underbau_rows.append(base)
                warnings.append({
                    'Typ': 'Unterbau',
                    'Pritsche': pname,
                    'Einheit_ID': einheit_id,
                    'Warnung': 'Unterbau erforderlich',
                    'Details': f'{ub_h:.0f} mm unter {einheit_id} / Z={z:.0f} mm',
                })

    return pd.DataFrame(underbau_rows, columns=columns), pd.DataFrame(warnings, columns=warning_cols)


def add_underbau_rows_to_placements(
    placements_df: pd.DataFrame,
    platforms_df: pd.DataFrame,
    min_support_ratio: float = 0.65,
    min_underbau_height: float = 20.0,
    enabled: bool = True,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Fügt Unterbau-Hilfszeilen für Darstellung/Ladeplan hinzu und liefert Warnungen."""
    if not enabled or placements_df is None or placements_df.empty or platforms_df is None or platforms_df.empty:
        return placements_df.copy() if placements_df is not None else pd.DataFrame(), pd.DataFrame(columns=['Typ', 'Pritsche', 'Einheit_ID', 'Warnung', 'Details'])
    helpers: List[pd.DataFrame] = []
    warnings: List[pd.DataFrame] = []
    for _, platform in platforms_df.iterrows():
        h, w = calculate_underbau_rows_for_platform(placements_df, platform, min_support_ratio=min_support_ratio, min_underbau_height=min_underbau_height)
        if h is not None and not h.empty:
            helpers.append(h)
        if w is not None and not w.empty:
            warnings.append(w)
    result = placements_df.copy()
    if helpers:
        result = pd.concat([result] + helpers, ignore_index=True, sort=False)
    warn_df = pd.concat(warnings, ignore_index=True, sort=False) if warnings else pd.DataFrame(columns=['Typ', 'Pritsche', 'Einheit_ID', 'Warnung', 'Details'])
    return result, warn_df

def compute_loading_warnings(placements_df: pd.DataFrame, platforms_df: pd.DataFrame) -> pd.DataFrame:
    """Prüft Länge, Breite, Höhe, Gewicht, negative Positionen und nicht verladen."""
    warnings: List[Dict[str, Any]] = []
    if placements_df is None or placements_df.empty:
        return pd.DataFrame([{'Typ': 'Info', 'Pritsche': '', 'Einheit_ID': '', 'Warnung': 'Keine Platzierung vorhanden', 'Details': ''}])

    platform_lookup: Dict[str, Dict[str, float]] = {}
    if platforms_df is not None and not platforms_df.empty:
        for _, row in platforms_df.iterrows():
            name = str(row.get('Pritsche', ''))
            platform_lookup[name] = {
                'eff_length': safe_number(row.get('Länge_mm')) + safe_number(row.get('Überhang_vorne_mm')) + safe_number(row.get('Überhang_hinten_mm')),
                'width': safe_number(row.get('Breite_mm')),
                'max_height': safe_number(row.get('Max_Höhe_mm')),
                'max_weight': safe_number(row.get('Max_Gewicht_kg')),
                'own_weight': safe_number(row.get('Eigengewicht_Pritsche_kg')),
                'base_height': safe_number(row.get('Kantholz_erste_Lage_mm')),
            }

    for _, row in placements_df.iterrows():
        pritsche = str(row.get('Pritsche', ''))
        einheit = str(row.get('Einheit_ID', ''))

        if pritsche == 'NICHT VERLADEN':
            warnings.append({'Typ': 'Nicht verladen', 'Pritsche': pritsche, 'Einheit_ID': einheit, 'Warnung': 'Einheit wurde nicht platziert', 'Details': str(row.get('Ebene', ''))})
            continue

        if pritsche not in platform_lookup:
            warnings.append({'Typ': 'Pritsche', 'Pritsche': pritsche, 'Einheit_ID': einheit, 'Warnung': 'Pritsche nicht in Stammdaten gefunden', 'Details': 'Name der Pritsche prüfen'})
            continue

        limits = platform_lookup[pritsche]
        x = safe_number(row.get('X_mm'), 0.0)
        y = safe_number(row.get('Y_mm'), 0.0)
        z = safe_number(row.get('Z_mm'), 0.0)
        length = safe_number(row.get('Länge_mm'), 0.0)
        width = safe_number(row.get('Breite_mm'), 0.0)
        height = safe_number(row.get('Höhe_mm'), 0.0)

        if x < 0 or y < 0 or z < 0:
            warnings.append({'Typ': 'Position', 'Pritsche': pritsche, 'Einheit_ID': einheit, 'Warnung': 'Negative Position', 'Details': f'X={x:.0f}, Y={y:.0f}, Z={z:.0f}'})
        if x + length > limits['eff_length']:
            warnings.append({'Typ': 'Länge', 'Pritsche': pritsche, 'Einheit_ID': einheit, 'Warnung': 'Länge / Überhang überschritten', 'Details': f'{x + length:.0f} mm > {limits["eff_length"]:.0f} mm'})
        if y + width > limits['width']:
            warnings.append({'Typ': 'Breite', 'Pritsche': pritsche, 'Einheit_ID': einheit, 'Warnung': 'Breite überschritten', 'Details': f'{y + width:.0f} mm > {limits["width"]:.0f} mm'})
        check_height = z + height if z >= limits.get('base_height', 0.0) - 0.1 else z + height + limits.get('base_height', 0.0)
        if check_height > limits['max_height']:
            warnings.append({'Typ': 'Höhe', 'Pritsche': pritsche, 'Einheit_ID': einheit, 'Warnung': 'Höhe überschritten', 'Details': f'{check_height:.0f} mm > {limits["max_height"]:.0f} mm'})

    if platforms_df is not None and not platforms_df.empty:
        for platform_name, group in placements_df[placements_df['Pritsche'] != 'NICHT VERLADEN'].groupby('Pritsche'):
            name = str(platform_name)
            if name not in platform_lookup:
                continue
            load_weight = pd.to_numeric(group.get('Gewicht_kg'), errors='coerce').fillna(0).sum()
            own_weight = platform_lookup[name].get('own_weight', 0.0)
            total_weight = load_weight + own_weight
            max_weight = platform_lookup[name]['max_weight']
            if total_weight > max_weight:
                warnings.append({'Typ': 'Gewicht', 'Pritsche': name, 'Einheit_ID': '', 'Warnung': 'Max. Gesamtgewicht überschritten', 'Details': f'{total_weight:.0f} kg inkl. Eigengewicht > {max_weight:.0f} kg'})

    return pd.DataFrame(warnings)


def recompute_summary_from_placements(placements_df: pd.DataFrame, platforms_df: pd.DataFrame) -> pd.DataFrame:
    """Erstellt die Pritschen-Zusammenfassung neu, damit manuelle Änderungen berücksichtigt werden."""
    rows: List[Dict[str, Any]] = []
    if platforms_df is None or platforms_df.empty:
        return pd.DataFrame()

    for _, prow in platforms_df.iterrows():
        pname = str(prow.get('Pritsche', ''))
        group = placements_df[placements_df.get('Pritsche', pd.Series(dtype=str)).astype(str) == pname] if not placements_df.empty else pd.DataFrame()
        group = group[group.get('X_mm', pd.Series(dtype=float)).notna()] if not group.empty else group
        used_length = float((group['X_mm'] + group['Länge_mm']).max()) if not group.empty else 0.0
        used_width = float((group['Y_mm'] + group['Breite_mm']).max()) if not group.empty else 0.0
        base_height = safe_number(prow.get('Kantholz_erste_Lage_mm'), 0.0)
        if not group.empty:
            raw_height = float((group['Z_mm'] + group['Höhe_mm']).max())
            min_z = float(group['Z_mm'].min())
            # Normal startet die Platzierung bereits auf dem Kantholz. Falls manuell ab Z=0 gesetzt wird,
            # wird das Kantholz trotzdem zur Frachthöhe addiert.
            used_height = raw_height if min_z >= base_height - 0.1 else raw_height + base_height
        else:
            used_height = base_height
        used_weight = float(pd.to_numeric(group.get('Gewicht_kg'), errors='coerce').fillna(0).sum()) if not group.empty else 0.0
        platform_own_weight = safe_number(prow.get('Eigengewicht_Pritsche_kg'), 0.0)
        rows.append({
            'Fuhre_Nr': prow.get('Fuhre_Nr'),
            'Fuhrenoption': prow.get('Fuhrenoption'),
            'Pritschenname': prow.get('Pritschenname'),
            'Pritsche': pname,
            'Länge genutzt_mm': round(used_length, 1),
            'Breite genutzt_mm': round(used_width, 1),
            'Höhe genutzt_mm': round(used_height, 1),
            'Gewicht genutzt_kg': round(used_weight, 2),
            'Eigengewicht Pritsche_kg': round(platform_own_weight, 2),
            'Gesamtgewicht inkl. Pritsche_kg': round(used_weight + platform_own_weight, 2),
            'Max Länge effektiv_mm': round(safe_number(prow.get('Länge_mm')) + safe_number(prow.get('Überhang_vorne_mm')) + safe_number(prow.get('Überhang_hinten_mm')), 1),
            'Max Breite_mm': round(safe_number(prow.get('Breite_mm')), 1),
            'Max Höhe_mm': round(safe_number(prow.get('Max_Höhe_mm')), 1),
            'Max Gewicht_kg': round(safe_number(prow.get('Max_Gewicht_kg')), 1),
        })
    return pd.DataFrame(rows)



def _format_bsd_cell(row: pd.Series) -> str:
    """Beschriftung für die Ladeplan-BSD-Matrix."""
    bauteile = str(row.get('Bauteile', '') or '').strip()
    einheit = str(row.get('Einheit_ID', '') or '').strip()
    typ = str(row.get('Typ', '') or '').strip()
    anzahl = int(safe_number(row.get('Anzahl_Bauteile'), 1))

    if bauteile and bauteile.lower() != 'nan':
        label = bauteile
    else:
        label = einheit

    if typ == 'Bund' and einheit:
        return f'{einheit} ({anzahl} Stk.)\n{label}'
    return label or einheit


def _orientation_flags_from_meta(project_meta: Optional[Dict[str, Any]] = None) -> Tuple[bool, bool]:
    """Fixe Orientierung für Ladeplan/PDF.

    V55:
    Die Orientierung ist nicht mehr über die Oberfläche verstellbar, damit die
    Darstellung nicht optisch "schön", aber fachlich falsch gedreht werden kann.

    Fix:
    - Vorne = X=max
    - Hinten = X=0
    - Links = Y=0
    - Rechts = Y=max
    """
    return True, False


def _position_slot_for_bsd(
    row: pd.Series,
    eff_length: float,
    platform_width: float,
    front_at_x_max: bool = False,
    left_at_y_max: bool = False,
) -> str:
    """Ordnet eine Einheit einer BSD-Position zu.

    V55:
    Die BSD-Tabelle verwendet dieselbe einfache Positionslogik wie die
    Visualisierung: Entscheidend ist der Mittelpunkt des sichtbaren Rechtecks.

    - X-Mitte bestimmt vorne/hinten.
    - Y-Mitte bestimmt links/rechts.
    - Exakt X-mittig bleibt gemäss BSD-Regel "vorne".
    - Ein Bund, der die X-Mitte nur berührt/überspannt, wird NICHT automatisch
      nach vorne gezwungen, wenn sein Mittelpunkt hinten liegt.
    """
    x0 = safe_number(row.get('X_mm'))
    y0 = safe_number(row.get('Y_mm'))
    length = safe_number(row.get('Länge_mm'))
    width = safe_number(row.get('Breite_mm'))
    x_mid = x0 + length / 2.0
    y_mid = y0 + width / 2.0

    x_center = eff_length / 2.0 if eff_length else 0.0
    y_center = platform_width / 2.0 if platform_width else 0.0

    # Nur wirklich mittig als "vorne" behandeln. Keine grosse Toleranz,
    # weil sonst hinten liegende Bauteile fälschlich in "vorne" landen.
    center_tolerance = 50.0
    x_is_middle = abs(x_mid - x_center) <= center_tolerance

    if x_is_middle:
        front_back = 'Vorne'
    elif front_at_x_max:
        front_back = 'Vorne' if x_mid >= x_center else 'Hinten'
    else:
        front_back = 'Vorne' if x_mid <= x_center else 'Hinten'

    if left_at_y_max:
        left_right = 'links' if y_mid >= y_center else 'rechts'
    else:
        left_right = 'links' if y_mid <= y_center else 'rechts'

    return f'{front_back} {left_right}'

def _position_slots_for_bsd(
    row: pd.Series,
    eff_length: float,
    platform_width: float,
    front_at_x_max: bool = False,
    left_at_y_max: bool = False,
) -> List[str]:
    """Wie _position_slot_for_bsd, aber breite mittige Elemente können beide Seiten belegen.

    - X mittig bleibt fachlich 'Vorne'.
    - Y wird über die Bauteil-/Bundmitte entschieden.
    - Elemente/Bunde, die nahezu die ganze Pritschenbreite belegen, werden in
      links UND rechts angezeigt. Für die Auflage- und BSD-Logik gelten sie als
      über beide Seiten belegt.
    """
    slot = _position_slot_for_bsd(row, eff_length, platform_width, front_at_x_max, left_at_y_max)
    width = safe_number(row.get('Breite_mm'))
    y0 = safe_number(row.get('Y_mm'))
    y1 = y0 + width
    y_center = platform_width / 2.0 if platform_width else 0.0
    spans_full_width = platform_width > 0 and width >= platform_width * 0.75 and y0 < y_center < y1
    if not spans_full_width:
        return [slot]
    # Breiter/mittiger Bund belegt links und rechts.
    # Vorne/hinten bleibt trotzdem aus der Mittelpunkt-Logik der Visualisierung.
    x0 = safe_number(row.get('X_mm'))
    length = safe_number(row.get('Länge_mm'))
    x_mid = x0 + length / 2.0
    x_center = eff_length / 2.0 if eff_length else 0.0
    x_is_middle = abs(x_mid - x_center) <= 50.0
    fb = 'Vorne' if x_is_middle or slot.startswith('Vorne') else 'Hinten'
    return [f'{fb} links', f'{fb} rechts']


def create_bsd_header_for_platform(
    platform: pd.Series,
    summary_df: pd.DataFrame,
    warnings_df: Optional[pd.DataFrame] = None,
    project_meta: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Kopfdaten je Pritsche ähnlich Ladeplan BSD."""
    pname = str(platform.get('Pritsche', ''))
    srow = summary_df[summary_df['Pritsche'].astype(str) == pname] if summary_df is not None and not summary_df.empty else pd.DataFrame()
    srow = srow.iloc[0] if not srow.empty else pd.Series(dtype=object)
    warn_count = 0
    if warnings_df is not None and not warnings_df.empty and 'Pritsche' in warnings_df.columns:
        warn_count = int((warnings_df['Pritsche'].astype(str) == pname).sum())

    project_meta = project_meta or {}

    return {
        'Pritsche': pname,
        'Fuhre_Nr': platform.get('Fuhre_Nr', ''),
        'Fuhrenoption': platform.get('Fuhrenoption', ''),
        'Pritschenname': platform.get('Pritschenname', ''),
        'Pritschenhöhe_mm': safe_number(platform.get('Max_Höhe_mm')),
        'Pritschenbreite_mm': safe_number(platform.get('Breite_mm')),
        'Pritschenlänge_effektiv_mm': safe_number(platform.get('Länge_mm')) + safe_number(platform.get('Überhang_vorne_mm')) + safe_number(platform.get('Überhang_hinten_mm')),
        'Frachthöhe_mm': safe_number(srow.get('Höhe genutzt_mm')),
        'Höhe_gesamt_mm': safe_number(srow.get('Höhe genutzt_mm')),
        'Länge_gesamt_mm': safe_number(srow.get('Länge genutzt_mm')),
        'Breite_gesamt_mm': safe_number(srow.get('Breite genutzt_mm')),
        'Ladegewicht_kg': safe_number(srow.get('Gewicht genutzt_kg')),
        'Eigengewicht_Pritsche_kg': safe_number(srow.get('Eigengewicht Pritsche_kg'), safe_number(platform.get('Eigengewicht_Pritsche_kg'))),
        'Gesamtgewicht_kg': safe_number(srow.get('Gesamtgewicht inkl. Pritsche_kg'), safe_number(srow.get('Gewicht genutzt_kg')) + safe_number(platform.get('Eigengewicht_Pritsche_kg'))),
        'Max_Gewicht_kg': safe_number(platform.get('Max_Gewicht_kg')),
        'Warnungen': warn_count,
        'Objekt_Name': project_meta.get('Objekt_Name', ''),
        'Transport_Name': project_meta.get('Transport_Name', ''),
        'Sachbearbeiter': project_meta.get('Sachbearbeiter', ''),
        'Datum': project_meta.get('Datum', datetime.now().strftime('%d.%m.%Y')),
        'Decke': project_meta.get('Decke', ''),
        'Bauabschnitt': project_meta.get('Bauabschnitt', ''),
        'Decke_Attribut': project_meta.get('Decke_Attribut', ''),
        'Bauabschnitt_Attribut': project_meta.get('Bauabschnitt_Attribut', ''),
    }



def _split_bsd_text_list(value: Any, fallback: str = '') -> List[str]:
    """Zerlegt Bauteillisten aus Verladeeinheiten robust."""
    text = str(value or '').strip()
    if not text or text.lower() == 'nan':
        text = str(fallback or '').strip()
    if not text:
        return []
    if '|' in text:
        items = [item.strip() for item in text.split('|')]
    else:
        items = [item.strip() for item in re.split(r'\s*,\s*', text)]
    return [item for item in items if item and item.lower() != 'nan']


def _split_bsd_number_list(value: Any, count: int, fallback_total: float, spacer_height: float) -> List[float]:
    """Zerlegt Einzelhöhen/Längen/Breiten. Fehlen Werte, wird sinnvoll aufgefüllt."""
    text = str(value or '').strip()
    values: List[float] = []
    if text and text.lower() != 'nan':
        raw_items = text.split('|') if '|' in text else re.split(r'[;,]', text)
        for item in raw_items:
            if str(item).strip():
                values.append(safe_number(item, 0.0))
    if count <= 0:
        return values
    if len(values) < count:
        # Wenn keine Einzelhöhen bekannt sind, wird aus der Gesamthöhe abzüglich Einlagen verteilt.
        if not values:
            remaining = max(0.0, fallback_total - max(0, count - 1) * spacer_height)
            default_each = remaining / count if count else fallback_total
        else:
            default_each = values[-1]
        values.extend([default_each] * (count - len(values)))
    return values[:count]


def _fmt_bsd_mm_label(prefix: str, value: float) -> str:
    v = safe_number(value)
    if abs(v - round(v)) < 0.001:
        text = f'{v:.0f}'
    else:
        text = f'{v:.2f}'.rstrip('0').rstrip('.')
    return f'{prefix} {text}'.strip()


def create_bsd_matrix_for_platform(
    placements_df: pd.DataFrame,
    platform: pd.Series,
    top_first: bool = True,
    front_at_x_max: bool = False,
    left_at_y_max: bool = False,
) -> pd.DataFrame:
    """Erstellt eine Ladeplan-BSD-Tabelle je Pritsche.

    V55:
    - Die 4 BSD-Spalten bleiben bestehen.
    - Jede BSD-Spalte wird als eigene kompakte Stapelliste aufgebaut.
    - Leere Zellen durch andere Spalten/Höhen werden nicht mehr erzeugt.
    - Eine Tabellenzeile bedeutet NICHT mehr "gleiche Höhe über alle Spalten".
    - Bundmarkierung B1/B2/... bleibt erhalten.
    - H/M/V wird nur verwendet, wenn auf gleicher Höhe/Seite drei X-Bereiche existieren.
    """
    columns = [
        'Pritsche', 'Fuhre_Nr', 'Lage', 'Z_mm',
        'Vorne links', 'Vorne rechts', 'Hinten links', 'Hinten rechts',
        'Bemerkung vorne links', 'Bemerkung vorne rechts', 'Bemerkung hinten links', 'Bemerkung hinten rechts',
        'Höhe_mm', 'Breite_mm', 'Gesamtlänge_mm', 'Gewicht_kg', 'Anzahl_Einheiten', 'Zeilentyp'
    ]
    if placements_df is None or placements_df.empty:
        return pd.DataFrame(columns=columns)

    pname = str(platform.get('Pritsche', ''))
    eff_length = safe_number(platform.get('Länge_mm')) + safe_number(platform.get('Überhang_vorne_mm')) + safe_number(platform.get('Überhang_hinten_mm'))
    platform_width = safe_number(platform.get('Breite_mm'))
    base_height = safe_number(platform.get('Kantholz_erste_Lage_mm'), 0.0)
    spacer_height = safe_number(platform.get('Einlage_zwischen_Lagen_mm'), 0.0)
    platform_general_spacer = safe_number(platform.get('Einlage_allgemein_mm'), 0.0)

    rows = placements_df[placements_df['Pritsche'].astype(str) == pname].copy()
    rows = rows[rows['X_mm'].notna() & rows['Y_mm'].notna() & rows['Z_mm'].notna()].copy()
    if rows.empty:
        return pd.DataFrame(columns=columns)

    for col in ['X_mm', 'Y_mm', 'Z_mm', 'Länge_mm', 'Breite_mm', 'Höhe_mm', 'Gewicht_kg']:
        rows[col] = pd.to_numeric(rows[col], errors='coerce').fillna(0.0)

    slots = ['Vorne links', 'Vorne rechts', 'Hinten links', 'Hinten rechts']
    slot_has_load = {slot: False for slot in slots}
    entries: List[Dict[str, Any]] = []

    def _row_x_mid(row: pd.Series) -> Optional[float]:
        length = safe_number(row.get('Länge_mm'))
        x0 = safe_number(row.get('X_mm'))
        if length <= 0:
            return None
        return x0 + length / 2.0

    def add_entry(
        z: float,
        slot: str,
        label: str,
        kind: str,
        height: float = 0.0,
        length: float = 0.0,
        width: float = 0.0,
        weight: float = 0.0,
        remark: str = '',
        x_mid: Optional[float] = None,
    ) -> None:
        if not slot or slot not in slots:
            return
        label_text = str(label or '').strip()
        if not label_text:
            return
        entries.append({
            'z': round(float(z), 1),
            'slot': slot,
            'label': label_text,
            'kind': str(kind or ''),
            'height': safe_number(height, 0.0),
            'length': safe_number(length, 0.0),
            'width': safe_number(width, 0.0),
            'weight': safe_number(weight, 0.0),
            'remark': str(remark or '').strip(),
            'x_zone': '',
            'x_mid': x_mid,
        })

    def _bsd_bundle_tag(row: pd.Series) -> str:
        raw = str(row.get('Einheit_ID', '') or '').strip()
        if not raw:
            return ''
        m = re.match(r'B0*(\d+)$', raw)
        if m:
            return f"B{m.group(1)}"
        return raw

    # Bauteile/Bunde in sichtbare Einträge je BSD-Spalte zerlegen.
    for _, row in rows.sort_values(['Z_mm', 'X_mm', 'Y_mm'], kind='stable').iterrows():
        slots_for_row = _position_slots_for_bsd(row, eff_length, platform_width, front_at_x_max, left_at_y_max)
        for _slot in slots_for_row:
            slot_has_load[_slot] = True

        typ = str(row.get('Typ', '') or '').strip()
        count = max(1, int(safe_number(row.get('Anzahl_Bauteile'), 1)))
        labels = _split_bsd_text_list(row.get('Ansicht_Liste', ''), row.get('Bauteile_Liste', row.get('Bauteile', row.get('Einheit_ID', ''))))
        if not labels:
            labels = [_format_bsd_cell(row)]

        x_mid = _row_x_mid(row)

        if typ == 'Unterbau':
            label = labels[0] if labels else _format_bsd_cell(row)
            for slot in slots_for_row:
                add_entry(row['Z_mm'], slot, label, 'Unterbau', row['Höhe_mm'], row['Länge_mm'], row['Breite_mm'], 0.0, 'Ausgleich / Aufdopplung', x_mid)
            continue

        if typ != 'Bund' or count <= 1:
            label = labels[0] if labels else _format_bsd_cell(row)
            for slot in slots_for_row:
                add_entry(row['Z_mm'], slot, label, 'Bauteil', row['Höhe_mm'], row['Länge_mm'], row['Breite_mm'], row['Gewicht_kg'], '', x_mid)
            continue

        # Bund: jedes enthaltene Bauteil bleibt sichtbar, aber mit Bundnummer.
        labels = (labels + [labels[-1]] * count)[:count]
        internal_spacer = safe_number(row.get('Einlage_allgemein_mm'), platform_general_spacer)
        if internal_spacer <= 0:
            internal_spacer = 0.0

        part_heights = _split_bsd_number_list(row.get('Einzelhöhen_mm', ''), count, row['Höhe_mm'], internal_spacer)
        part_lengths = _split_bsd_number_list(row.get('Einzellängen_mm', ''), count, row['Länge_mm'], 0.0)
        part_widths = _split_bsd_number_list(row.get('Einzelbreiten_mm', ''), count, row['Breite_mm'], 0.0)

        labels_bottom_to_top = list(reversed(labels))
        part_heights_bottom_to_top = list(reversed(part_heights))
        part_lengths_bottom_to_top = list(reversed(part_lengths))
        part_widths_bottom_to_top = list(reversed(part_widths))

        part_weight = safe_number(row.get('Gewicht_kg')) / count if count else safe_number(row.get('Gewicht_kg'))
        z_cursor = safe_number(row.get('Z_mm'))
        bund_tag = _bsd_bundle_tag(row)

        for i, label in enumerate(labels_bottom_to_top):
            ph = safe_number(part_heights_bottom_to_top[i] if i < len(part_heights_bottom_to_top) else row['Höhe_mm'])
            pl = safe_number(part_lengths_bottom_to_top[i] if i < len(part_lengths_bottom_to_top) else row['Länge_mm'])
            pw = safe_number(part_widths_bottom_to_top[i] if i < len(part_widths_bottom_to_top) else row['Breite_mm'])
            visible_label = f"{bund_tag}:{label}" if bund_tag else str(label)
            for slot in slots_for_row:
                add_entry(z_cursor, slot, visible_label, 'Bund-Bauteil', ph, pl, pw, part_weight, bund_tag, x_mid)
            z_cursor += ph
            if i < count - 1 and internal_spacer > 0:
                for slot in slots_for_row:
                    add_entry(z_cursor, slot, _fmt_bsd_mm_label('Einlage', internal_spacer), 'Einlage', internal_spacer, row['Länge_mm'], row['Breite_mm'], 0.0, '', x_mid)
                z_cursor += internal_spacer

    # Kantholz je belegter Spalte.
    if base_height > 0:
        for slot, has_load in slot_has_load.items():
            if has_load:
                add_entry(0.0, slot, _fmt_bsd_mm_label('Kantholz', base_height), 'Kantholz erste Lage', base_height, eff_length, platform_width, 0.0, '', None)

    # Einlagen/Lagenholz zwischen separaten Lagen je belegter Spalte.
    layer_z_values = sorted({round(float(v), 1) for v in rows['Z_mm'].tolist() if float(v) > base_height + 0.1})
    existing_spacer_keys = {(round(e['z'], 1), e['slot']) for e in entries if 'Einlage' in e['kind'] or 'Lagenholz' in e['kind']}
    existing_real_keys = {(round(e['z'], 1), e['slot']) for e in entries if 'Bauteil' in e['kind'] or e['kind'] == 'Unterbau'}

    for z_val in layer_z_values:
        layer_rows = rows[rows['Z_mm'].round(1) == z_val].copy()
        for _, lrow in layer_rows.iterrows():
            slots_for_lrow = _position_slots_for_bsd(lrow, eff_length, platform_width, front_at_x_max, left_at_y_max)
            is_bundle_layer = str(lrow.get('Typ', '')).strip() == 'Bund'
            if is_bundle_layer and spacer_height > 0:
                effective_layer_spacer = spacer_height
                spacer_kind = 'Lagenholz'
            elif platform_general_spacer > 0:
                effective_layer_spacer = platform_general_spacer
                spacer_kind = 'Einlage'
            else:
                continue
            z_spacer = round(max(0.0, z_val - effective_layer_spacer), 1)
            x_mid = _row_x_mid(lrow)
            for slot in slots_for_lrow:
                if (z_spacer, slot) not in existing_spacer_keys and (z_spacer, slot) not in existing_real_keys:
                    add_entry(z_spacer, slot, _fmt_bsd_mm_label('Einlage', effective_layer_spacer), spacer_kind, effective_layer_spacer, lrow['Länge_mm'], lrow['Breite_mm'], 0.0, '', x_mid)

    if not entries:
        return pd.DataFrame(columns=columns)

    # Einlage-/Lagenholz-Zeilen nie mit Bauteilzeilen auf exakt gleichem z mischen.
    real_z_levels = {round(e['z'], 1) for e in entries if ('Bauteil' in e['kind']) or e['kind'] == 'Unterbau'}
    for e in entries:
        if (('Einlage' in e['kind']) or ('Lagenholz' in e['kind'])) and round(e['z'], 1) in real_z_levels:
            e['z'] = round(max(0.1, float(e['z']) - 0.2), 1)

    # H/M/V nur bei drei X-Bereichen auf gleicher Höhe und gleicher Seite.
    grouped: Dict[Tuple[float, str], List[Dict[str, Any]]] = {}
    for e in entries:
        kind = str(e.get('kind', ''))
        if ('Bauteil' not in kind) and kind != 'Unterbau':
            continue
        slot = str(e.get('slot', ''))
        side = 'links' if 'links' in slot else ('rechts' if 'rechts' in slot else '')
        x_mid = e.get('x_mid')
        if not side or x_mid is None:
            continue
        grouped.setdefault((round(float(e.get('z', 0.0)), 1), side), []).append(e)

    for _, group in grouped.items():
        unique_positions = sorted({round(float(e.get('x_mid')), 1) for e in group if e.get('x_mid') is not None})
        if len(unique_positions) < 3:
            continue
        back_pos = unique_positions[0]
        front_pos = unique_positions[-1]
        for e in group:
            x_mid_val = round(float(e.get('x_mid')), 1)
            if abs(x_mid_val - back_pos) <= 1.0:
                e['x_zone'] = 'H'
            elif abs(x_mid_val - front_pos) <= 1.0:
                e['x_zone'] = 'V'
            else:
                e['x_zone'] = 'M'

    def _entry_sort_key(e: Dict[str, Any]) -> Tuple[float, float, str]:
        return (round(float(e.get('z', 0.0)), 1), safe_number(e.get('x_mid'), 0.0), str(e.get('label', '')))

    # Kompakte BSD-Spaltenliste: jede Spalte wird separat sortiert und dann
    # zeilenweise aufgefüllt. Leere Zellen entstehen nur, wenn eine Spalte kürzer ist.
    slot_lists: Dict[str, List[Dict[str, Any]]] = {}
    for slot in slots:
        slot_entries = [e for e in entries if e['slot'] == slot]
        slot_entries = sorted(slot_entries, key=_entry_sort_key, reverse=bool(top_first))
        slot_lists[slot] = slot_entries

    max_len = max((len(v) for v in slot_lists.values()), default=0)
    matrix_rows: List[Dict[str, Any]] = []
    for idx in range(max_len):
        out: Dict[str, Any] = {
            'Pritsche': pname,
            'Fuhre_Nr': platform.get('Fuhre_Nr', ''),
            'Lage': max_len - idx if top_first else idx + 1,
            'Z_mm': '',
            'Vorne links': '',
            'Vorne rechts': '',
            'Hinten links': '',
            'Hinten rechts': '',
            'Bemerkung vorne links': '',
            'Bemerkung vorne rechts': '',
            'Bemerkung hinten links': '',
            'Bemerkung hinten rechts': '',
            'Höhe_mm': '',
            'Breite_mm': '',
            'Gesamtlänge_mm': '',
            'Gewicht_kg': '',
            'Anzahl_Einheiten': '',
            'Zeilentyp': 'BSD-Spaltenliste',
        }

        row_entries: List[Dict[str, Any]] = []
        for slot in slots:
            if idx >= len(slot_lists[slot]):
                continue
            e = slot_lists[slot][idx]
            lbl = str(e.get('label', '') or '')
            zone = str(e.get('x_zone', '') or '').strip()
            kind = str(e.get('kind', '') or '')
            if zone and ('Bauteil' in kind or kind == 'Unterbau'):
                lbl = f"{zone}: {lbl}"
            out[slot] = lbl
            if e.get('remark'):
                out[f'Bemerkung {slot.lower()}'] = e.get('remark')
            row_entries.append(e)

        real_entries = [e for e in row_entries if ('Bauteil' in str(e.get('kind', ''))) or str(e.get('kind', '')) == 'Unterbau']
        visible_entries = row_entries if row_entries else []
        metric_entries = real_entries if real_entries else visible_entries

        if metric_entries:
            out['Z_mm'] = round(max(safe_number(e.get('z'), 0.0) for e in metric_entries), 1)
            out['Höhe_mm'] = round(max(safe_number(e.get('height'), 0.0) for e in metric_entries), 1)
            out['Breite_mm'] = round(max(safe_number(e.get('width'), 0.0) for e in metric_entries), 1)
            out['Gesamtlänge_mm'] = round(max(safe_number(e.get('length'), 0.0) for e in metric_entries), 1)
            out['Gewicht_kg'] = round(sum(safe_number(e.get('weight'), 0.0) for e in real_entries), 2)
            out['Anzahl_Einheiten'] = len(real_entries)
            out['Zeilentyp'] = ', '.join(sorted({str(e.get('kind', '')) for e in visible_entries}))

        matrix_rows.append(out)

    return pd.DataFrame(matrix_rows, columns=columns)

def create_all_bsd_matrices(
    placements_df: pd.DataFrame,
    platforms_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    warnings_df: Optional[pd.DataFrame] = None,
    project_meta: Optional[Dict[str, Any]] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Erstellt Kopfdaten und Ladeplan-BSD-Matrix für jede verwendete Pritsche."""
    if platforms_df is None or platforms_df.empty:
        return pd.DataFrame(), pd.DataFrame()

    header_rows: List[Dict[str, Any]] = []
    matrix_frames: List[pd.DataFrame] = []
    front_at_x_max, left_at_y_max = _orientation_flags_from_meta(project_meta)
    for _, platform in platforms_df.iterrows():
        pname = str(platform.get('Pritsche', ''))
        has_load = placements_df is not None and not placements_df.empty and (placements_df['Pritsche'].astype(str) == pname).any()
        if not has_load:
            # Leere Pritschen nicht als Ladeplan ausgeben.
            continue
        header_rows.append(create_bsd_header_for_platform(platform, summary_df, warnings_df, project_meta))
        matrix = create_bsd_matrix_for_platform(placements_df, platform, front_at_x_max=front_at_x_max, left_at_y_max=left_at_y_max)
        if not matrix.empty:
            matrix_frames.append(matrix)

    header_df = pd.DataFrame(header_rows)
    matrix_df = pd.concat(matrix_frames, ignore_index=True) if matrix_frames else pd.DataFrame()
    return header_df, matrix_df



def _pdf_projection_values(row: pd.Series, view: str, eff_length: float, width: float, front_at_x_max: bool = False, left_at_y_max: bool = False) -> Tuple[float, float, float, float, float]:
    """Gibt projiziertes Rechteck und Tiefenwert für eine Ansicht zurück."""
    x = safe_number(row.get('X_mm'))
    y = safe_number(row.get('Y_mm'))
    z = safe_number(row.get('Z_mm'))
    lx = safe_number(row.get('Länge_mm'))
    by = safe_number(row.get('Breite_mm'))
    hz = safe_number(row.get('Höhe_mm'))

    if view == 'top':
        return x, y, lx, by, z + hz

    # Linke/rechte Seitenansicht an die gewählte Links-Orientierung koppeln.
    # links = Y klein (Standard) oder links = Y gross.
    side_from_y_min = (view in ('side', 'side_left') and not left_at_y_max) or (view == 'side_right' and left_at_y_max)
    side_from_y_max = (view == 'side_right' and not left_at_y_max) or (view in ('side', 'side_left') and left_at_y_max)
    if side_from_y_min:
        return x, z, lx, hz, -y
    if side_from_y_max:
        return eff_length - x - lx, z, lx, hz, y + by

    # Vorder-/Rückansicht an die gewählte Vorne-Orientierung koppeln.
    # vorne = X klein (Standard) oder vorne = X gross.
    front_from_x_min = (view == 'front' and not front_at_x_max) or (view == 'back' and front_at_x_max)
    if front_from_x_min:
        return width - y - by, z, by, hz, -x

    # Ansicht von X gross.
    return y, z, by, hz, x + lx


def _pdf_visible_sort_value(row: pd.Series, view: str, eff_length: float, width: float, front_at_x_max: bool = False, left_at_y_max: bool = False) -> float:
    """Sortierwert: zuerst verdeckte, zuletzt sichtbare Bauteile zeichnen."""
    return _pdf_projection_values(row, view, eff_length, width, front_at_x_max, left_at_y_max)[4]


def _pdf_label_lines(row: pd.Series, view: str) -> List[str]:
    """Beschriftung für PDF-Ansichten.

    Bei Bunden werden die Nummern nicht nebeneinander geschrieben, sondern
    in der Verladereihenfolge untereinander. In der Draufsicht sieht man nur
    die oberste Nummer eines Bundes.
    """
    label = _format_label_value(row.get('Ansicht_Label'))
    if not label:
        label = _format_label_value(row.get('Bauteile'))
    if not label:
        label = _format_label_value(row.get('Einheit_ID'))

    if str(row.get('Typ', '')).strip() == 'Bund':
        labels = _split_bsd_text_list(row.get('Ansicht_Liste', ''), row.get('Bauteile_Liste', label))
        labels = [str(v).strip() for v in labels if str(v).strip()]
        if labels:
            if view == 'top':
                # Draufsicht: sichtbar ist die oberste Lage im Bund.
                # Die Liste im Bund ist logisch von oben nach unten.
                return [labels[0]]
            # Die PDF-Funktion zeichnet mehrere Zeilen von unten nach oben.
            # Damit die erste Nummer der Bundliste wirklich oben steht, wird nur
            # die Anzeige-Liste gedreht. Die Bundlogik selbst bleibt unverändert.
            return list(reversed(labels))
    return [str(label).replace('<br>', ' ').strip()]


def _pdf_truncate_to_width(c, text: str, max_width: float, font_name: str, font_size: float) -> str:
    """Kürzt Text so, dass er ungefähr in die Box passt."""
    text = str(text or '').strip()
    if not text:
        return ''
    if c.stringWidth(text, font_name, font_size) <= max_width:
        return text
    ell = '...'
    available = max_width - c.stringWidth(ell, font_name, font_size)
    if available <= 0:
        return ''
    result = ''
    for ch in text:
        if c.stringWidth(result + ch, font_name, font_size) > available:
            break
        result += ch
    return result + ell if result else ''


def _pdf_draw_label_lines(c, rx: float, ry: float, rw: float, rh: float, lines: List[str], view: str) -> None:
    """Zeichnet Beschriftungen ohne Überlappung innerhalb eines Rechtecks."""
    lines = [str(line).strip() for line in lines if str(line).strip()]
    if not lines or rw < 12 or rh < 7:
        return

    from reportlab.lib import colors
    font_name = 'Helvetica'
    n = len(lines)
    if n == 1:
        font_size = max(3.8, min(5.2, rw / max(8, len(lines[0]) * 2.0), rh * 0.42))
        c.setFont(font_name, font_size)
        c.setFillColor(colors.black)
        txt = _pdf_truncate_to_width(c, lines[0], max(4, rw - 4), font_name, font_size)
        if txt:
            c.drawCentredString(rx + rw / 2, ry + rh / 2 - font_size / 3, txt)
        return

    # Mehrere Nummern im Bund: vertikal von unten nach oben, Reihenfolge = Verladereihenfolge.
    font_size = min(5.0, max(2.7, (rh - 3) / max(1, n) * 0.72))
    line_step = max(font_size * 1.05, (rh - 4) / max(1, n))
    total_h = line_step * (n - 1)
    start_y = ry + rh / 2 - total_h / 2
    c.setFont(font_name, font_size)
    c.setFillColor(colors.black)
    for i, line in enumerate(lines):
        yy = start_y + i * line_step - font_size / 3
        if yy < ry + 1 or yy > ry + rh - font_size:
            continue
        txt = _pdf_truncate_to_width(c, line, max(4, rw - 4), font_name, font_size)
        if txt:
            c.drawCentredString(rx + rw / 2, yy, txt)




def _pdf_add_visible_spacer_rows(rows: pd.DataFrame, platform: pd.Series, view: str) -> pd.DataFrame:
    """Ergänzt Kantholz und Einlagen als sichtbare Hilfszeilen für PDF-Ansichten.

    Grund: Im Ladeplan BSD werden Kantholz und Bundeinlagen als eigene Lagen
    geführt. Damit Pritschenplan und Ladeplan zusammenpassen, müssen diese
    Hölzer auch in den Seiten-/Vorder-/Rückansichten gezeichnet werden.
    In der Draufsicht werden sie nicht gezeichnet, weil sie von oben nicht
    sichtbar sind, sobald Bauteile darüber liegen.
    """
    if rows is None or rows.empty or view == 'top':
        return rows

    base_height = safe_number(platform.get('Kantholz_erste_Lage_mm'), 0.0)
    bundle_spacer = safe_number(platform.get('Einlage_zwischen_Lagen_mm'), 0.0)
    general_spacer = safe_number(platform.get('Einlage_allgemein_mm'), 0.0)

    helpers = []
    required_cols = set(rows.columns)

    def helper_row(source: Optional[pd.Series], typ: str, label: str, x: float, y: float, z: float, lx: float, by: float, hz: float) -> Dict[str, Any]:
        base = {col: '' for col in required_cols}
        if source is not None:
            for col in required_cols:
                try:
                    base[col] = source.get(col, '')
                except Exception:
                    pass
        base.update({
            'Typ': typ,
            'Einheit_ID': label,
            'Bauteile': label,
            'Bauteile_Liste': label,
            'Ansicht_Label': label,
            'Ansicht_Liste': label,
            'X_mm': float(x),
            'Y_mm': float(y),
            'Z_mm': float(z),
            'Länge_mm': float(lx),
            'Breite_mm': float(by),
            'Höhe_mm': float(hz),
            'Gewicht_kg': 0.0,
        })
        return base

    # Kantholz erste Lage: als durchgehende Unterlage der tatsächlich belegten Fläche anzeigen.
    if base_height > 0:
        min_x = safe_number(rows['X_mm'].min(), 0.0)
        max_x = safe_number((rows['X_mm'] + rows['Länge_mm']).max(), min_x)
        min_y = safe_number(rows['Y_mm'].min(), 0.0)
        max_y = safe_number((rows['Y_mm'] + rows['Breite_mm']).max(), min_y)
        helpers.append(helper_row(None, 'Kantholz', _fmt_bsd_mm_label('Kantholz', base_height), min_x, min_y, 0.0, max(0.0, max_x - min_x), max(0.0, max_y - min_y), base_height))

    # Zwischenlagen: vor allem Bundeinlage/Lagenholz zwischen separaten Bundlagen.
    # Die Positionen entsprechen den Z-Sprüngen im Ladeplan BSD.
    existing = set()
    for _, r in rows.iterrows():
        z = safe_number(r.get('Z_mm'), 0.0)
        if z <= base_height + 0.1:
            continue
        typ = str(r.get('Typ', '') or '').strip()
        spacer_h = 0.0
        spacer_label = ''
        spacer_type = ''
        if typ == 'Bund' and bundle_spacer > 0:
            spacer_h = bundle_spacer
            spacer_label = _fmt_bsd_mm_label('Einlage', spacer_h)
            spacer_type = 'Bundeinlage'
        elif general_spacer > 0:
            spacer_h = general_spacer
            spacer_label = _fmt_bsd_mm_label('Einlage', spacer_h)
            spacer_type = 'Einlage'
        if spacer_h <= 0:
            continue
        z_spacer = round(max(0.0, z - spacer_h), 1)
        key = (
            round(safe_number(r.get('X_mm')), 1),
            round(safe_number(r.get('Y_mm')), 1),
            round(z_spacer, 1),
            round(safe_number(r.get('Länge_mm')), 1),
            round(safe_number(r.get('Breite_mm')), 1),
            spacer_type,
        )
        if key in existing:
            continue
        existing.add(key)
        helpers.append(helper_row(r, spacer_type, spacer_label, safe_number(r.get('X_mm')), safe_number(r.get('Y_mm')), z_spacer, safe_number(r.get('Länge_mm')), safe_number(r.get('Breite_mm')), spacer_h))

    if not helpers:
        return rows
    return pd.concat([rows, pd.DataFrame(helpers)], ignore_index=True, sort=False)



def _pdf_draw_underbau_blocks(c, rx: float, ry: float, rw: float, rh: float, view: str, label: str) -> None:
    """Zeichnet Unterbau als Hilfselement: in Draufsicht gestrichelt, in Ansichten als Klötze."""
    from reportlab.lib import colors
    if rw <= 0 or rh <= 0:
        return
    c.saveState()
    c.setStrokeColor(colors.HexColor('#555555'))
    c.setFillColor(colors.HexColor('#d9d9d9'))
    c.setLineWidth(0.45)

    if view == 'top':
        # Nur gestrichelte Kontur, damit Bauteilnummern nicht verdeckt werden.
        c.setDash(3, 2)
        c.rect(rx, ry, rw, rh, stroke=1, fill=0)
        c.setDash()
        if rw > 22 and rh > 8:
            c.setFont('Helvetica', 3.8)
            c.setFillColor(colors.HexColor('#555555'))
            c.drawString(rx + 2, ry + 2, str(label).replace('Unterbau', 'UB')[:10])
        c.restoreState()
        return

    # Seiten-/Front-/Rückansichten: nicht als durchgehender Balken, sondern als 2 Auflagerklötze.
    # Die Klötze werden annähernd quadratisch, aber nie breiter als 25% der Projektion.
    block_w = max(4.0, min(max(rh, 5.0), rw * 0.22, 38.0))
    if rw <= block_w * 2.4:
        # Wenn das Bauteil sehr schmal projiziert ist, nur einen mittigen Klotz zeichnen.
        bx = rx + (rw - block_w) / 2
        c.rect(bx, ry, block_w, rh, stroke=1, fill=1)
    else:
        c.rect(rx, ry, block_w, rh, stroke=1, fill=1)
        c.rect(rx + rw - block_w, ry, block_w, rh, stroke=1, fill=1)
    if rw > 35 and rh > 7:
        c.setFont('Helvetica', 4.2)
        c.setFillColor(colors.HexColor('#333333'))
        c.drawCentredString(rx + rw / 2, ry + rh / 2 - 1.4, str(label).replace('Unterbau', 'UB')[:14])
    c.restoreState()

def _pdf_draw_view_orientation_helpers(c, ox: float, oy: float, draw_w: float, draw_h: float, view: str, front_at_x_max: bool, left_at_y_max: bool) -> None:
    """Kleine Bezugsmarken, damit Seiten-/Vorder-/Rückansicht besser zusammen gelesen werden können."""
    from reportlab.lib import colors
    c.saveState()
    c.setStrokeColor(colors.grey)
    c.setFillColor(colors.black)
    c.setLineWidth(0.35)
    c.setDash(2, 2)
    # Mittellinie als Bezug
    if view in ('top', 'side', 'side_left', 'side_right'):
        c.line(ox + draw_w / 2.0, oy, ox + draw_w / 2.0, oy + draw_h)
    elif view in ('front', 'back'):
        c.line(ox + draw_w / 2.0, oy, ox + draw_w / 2.0, oy + draw_h)
    c.setDash()
    c.setFont('Helvetica', 5.8)

    if view in ('top', 'side', 'side_left', 'side_right'):
        left_lbl = 'Hinten' if front_at_x_max else 'Vorne'
        right_lbl = 'Vorne' if front_at_x_max else 'Hinten'
        c.drawString(ox, oy + draw_h + 2.5, left_lbl)
        c.drawRightString(ox + draw_w, oy + draw_h + 2.5, right_lbl)
        c.drawCentredString(ox + draw_w / 2.0, oy + draw_h + 2.5, 'Mitte X')
        if view == 'top':
            # In der Draufsicht sind die schmalen Stirnseiten vorne/hinten.
            # Die langen Seiten sind links/rechts.
            top_side = 'Links' if left_at_y_max else 'Rechts'
            bottom_side = 'Rechts' if left_at_y_max else 'Links'
            c.drawCentredString(ox + draw_w / 2.0, oy + draw_h + 10.5, top_side)
            c.drawCentredString(ox + draw_w / 2.0, oy - 14.5, bottom_side)
    if view in ('front', 'back'):
        left_lbl = 'Rechts' if left_at_y_max else 'Links'
        right_lbl = 'Links' if left_at_y_max else 'Rechts'
        c.drawString(ox, oy + draw_h + 2.5, left_lbl)
        c.drawRightString(ox + draw_w, oy + draw_h + 2.5, right_lbl)
        c.drawCentredString(ox + draw_w / 2.0, oy + draw_h + 2.5, 'Mitte Y')
    c.restoreState()


def _pdf_filter_rows_for_side_view(rows: pd.DataFrame, platform: pd.Series, view: str, left_at_y_max: bool = False) -> pd.DataFrame:
    """Filtert Seitenansichten auf die wirklich sichtbare Breiten-Seite.

    Dadurch laufen linke und rechte Stapel in der Seitenansicht nicht mehr
    optisch ineinander. Breite/mittige Elemente bleiben in beiden Seitenansichten sichtbar.
    """
    if rows is None or rows.empty or view not in ('side', 'side_left', 'side_right'):
        return rows
    width = safe_number(platform.get('Breite_mm'), 0.0)
    if width <= 0:
        return rows
    center = width / 2.0

    # Welche physische Seite schaut diese Ansicht an?
    view_from_y_min = (view in ('side', 'side_left') and not left_at_y_max) or (view == 'side_right' and left_at_y_max)

    keep = []
    for _, row in rows.iterrows():
        y0 = safe_number(row.get('Y_mm'))
        by = safe_number(row.get('Breite_mm'))
        y1 = y0 + by
        y_mid = y0 + by / 2.0
        spans_center = y0 < center < y1
        is_wide = by >= width * 0.75
        if is_wide or spans_center:
            keep.append(True)
        elif view_from_y_min:
            keep.append(y_mid <= center)
        else:
            keep.append(y_mid >= center)

    return rows.loc[keep].copy()


def _pdf_draw_view(c, placements: pd.DataFrame, platform: pd.Series, x: float, y: float, w: float, h: float, view: str, title: str, front_at_x_max: bool = False, left_at_y_max: bool = False) -> None:
    """Zeichnet eine PDF-Ansicht mit Pritschen- und Ladungsabmessungen.

    Beschriftung wurde bewusst reduziert:
    - Bunde werden nummernweise untereinander dargestellt.
    - Verdeckte Bauteile werden zuerst gezeichnet, sichtbare Seiten zuletzt.
    - In der Draufsicht wird je Bund nur die oberste Nummer beschriftet.
    """
    from reportlab.lib import colors

    eff_length = safe_number(platform.get('Länge_mm')) + safe_number(platform.get('Überhang_vorne_mm')) + safe_number(platform.get('Überhang_hinten_mm'))
    width = safe_number(platform.get('Breite_mm'))
    max_height = safe_number(platform.get('Max_Höhe_mm'))
    pname = str(platform.get('Pritsche', ''))

    rows = placements[placements['Pritsche'].astype(str) == pname].copy()
    rows = rows[rows['X_mm'].notna() & rows['Y_mm'].notna() & rows['Z_mm'].notna()].copy()
    if view == 'top' and not rows.empty:
        # In der Draufsicht sollen Bauteilnummern im Vordergrund bleiben.
        # Unterbau wird später nur gestrichelt gezeichnet; Einlagen/Kantholz bleiben verdeckt.
        typ_series = rows.get('Typ', pd.Series(dtype=str)).astype(str)
        rows = rows[~typ_series.isin(['Kantholz', 'Bundeinlage', 'Einlage', 'Lagenholz'])].copy()
    elif view in ('side', 'side_left', 'side_right') and not rows.empty:
        # Seitenansicht = nur die jeweilige sichtbare Breiten-Seite.
        # Damit überschneiden sich z. B. 1763 und 1764 nicht mehr optisch,
        # wenn sie links/rechts auf unterschiedlichen Stapeln liegen.
        rows = _pdf_filter_rows_for_side_view(rows, platform, view, left_at_y_max=left_at_y_max)
    rows = _pdf_add_visible_spacer_rows(rows, platform, view)

    used_len = safe_number((rows['X_mm'] + rows['Länge_mm']).max() - rows['X_mm'].min(), 0.0) if not rows.empty else 0.0
    used_wid = safe_number((rows['Y_mm'] + rows['Breite_mm']).max() - rows['Y_mm'].min(), 0.0) if not rows.empty else 0.0
    used_hei = safe_number((rows['Z_mm'] + rows['Höhe_mm']).max(), 0.0) if not rows.empty else 0.0

    c.setStrokeColor(colors.black)
    c.setFillColor(colors.white)
    c.rect(x - 1, y + h + 10, min(max(w, 90), 180), 11, stroke=0, fill=1)
    c.setFillColor(colors.black)
    c.setFont('Helvetica-Bold', 9.4)
    c.drawString(x, y + h + 13, title)

    if view == 'top':
        data_w, data_h = max(eff_length, 1), max(width, 1)
        x_label, y_label = 'Länge X', 'Breite Y'
        dim_line = f'Pritsche L {eff_length:.0f} / B {width:.0f} mm   Ladung L {used_len:.0f} / B {used_wid:.0f} mm'
    elif view in ('side', 'side_left', 'side_right'):
        data_w, data_h = max(eff_length, 1), max(max_height, used_hei, 1)
        x_label, y_label = 'Länge X', 'Höhe Z'
        dim_line = f'Pritsche L {eff_length:.0f} / H {max_height:.0f} mm   Ladung L {used_len:.0f} / H {used_hei:.0f} mm'
    else:
        data_w, data_h = max(width, 1), max(max_height, used_hei, 1)
        x_label, y_label = 'Breite Y', 'Höhe Z'
        dim_line = f'Pritsche B {width:.0f} / H {max_height:.0f} mm   Ladung B {used_wid:.0f} / H {used_hei:.0f} mm'

    scale = min(w / data_w, h / data_h)
    draw_w = data_w * scale
    draw_h = data_h * scale
    ox = x
    oy = y

    # Pritschenrahmen / Maximalbereich
    c.setStrokeColor(colors.black)
    c.setLineWidth(0.85)
    c.rect(ox, oy, draw_w, draw_h, stroke=1, fill=0)
    _pdf_draw_view_orientation_helpers(c, ox, oy, draw_w, draw_h, view, front_at_x_max, left_at_y_max)

    # Achsen- und Maßtexte mit Abstand, damit nichts in die Bauteile läuft.
    c.setFont('Helvetica', 6.0)
    c.setFillColor(colors.black)
    c.drawString(ox, oy - 9, x_label)
    c.saveState()
    c.translate(ox - 12, oy + 2)
    c.rotate(90)
    c.drawString(0, 0, y_label)
    c.restoreState()
    c.setFont('Helvetica', 6.0)
    c.drawString(ox, oy - 20, dim_line)

    # Maßlinie unten / links
    c.setLineWidth(0.45)
    c.line(ox, oy - 3.5, ox + draw_w, oy - 3.5)
    c.line(ox, oy - 5.5, ox, oy - 1.5)
    c.line(ox + draw_w, oy - 5.5, ox + draw_w, oy - 1.5)
    c.saveState()
    c.translate(ox - 5, oy)
    c.line(0, 0, 0, draw_h)
    c.line(-1.8, 0, 1.8, 0)
    c.line(-1.8, draw_h, 1.8, draw_h)
    c.restoreState()

    # Grenzwerte klein anschreiben
    c.setFont('Helvetica', 5.4)
    c.drawRightString(ox + draw_w, oy - 10, f'{data_w:.0f}')
    c.drawString(ox - 2, oy + draw_h + 2, f'{data_h:.0f}')

    # Verdeckte Bauteile zuerst, sichtbare zuletzt.
    if not rows.empty:
        rows['_pdf_sort'] = rows.apply(lambda r: _pdf_visible_sort_value(r, view, eff_length, width, front_at_x_max, left_at_y_max), axis=1)
        rows = rows.sort_values(['_pdf_sort', 'Z_mm', 'X_mm', 'Y_mm'], kind='stable')

    for _, row in rows.iterrows():
        px, py, pw, ph, _depth = _pdf_projection_values(row, view, eff_length, width, front_at_x_max, left_at_y_max)
        rx = ox + px * scale
        ry = oy + py * scale
        rw = pw * scale
        rh = ph * scale
        if rw <= 0 or rh <= 0:
            continue
        row_typ = str(row.get('Typ', '') or '').strip()
        if row_typ == 'Kantholz':
            c.setFillColor(colors.HexColor('#b8b8b8'))
            c.setStrokeColor(colors.black)
        elif row_typ in ['Bundeinlage', 'Einlage']:
            c.setFillColor(colors.HexColor('#e6e6e6'))
            c.setStrokeColor(colors.grey)
        elif row_typ == 'Unterbau':
            _pdf_draw_underbau_blocks(c, rx, ry, rw, rh, view, str(row.get('Ansicht_Label', row.get('Bauteile', 'Unterbau'))))
            continue
        else:
            c.setFillColor(colors.lightgrey)
            c.setStrokeColor(colors.darkgrey)
        c.setLineWidth(0.38)
        c.rect(rx, ry, rw, rh, stroke=1, fill=1)
        label_lines = _pdf_label_lines(row, view)
        if view == 'top':
            # Draufsicht nur beschriften, wenn genug Platz vorhanden ist.
            # Sonst laufen die Nummern ineinander und der Plan wird unbrauchbar.
            if rw < 42 or rh < 18:
                label_lines = []
            else:
                label_lines = label_lines[:1]
        _pdf_draw_label_lines(c, rx, ry, rw, rh, label_lines, view)
def _pdf_draw_bsd_matrix_page(c, page_w: float, page_h: float, margin: float, platform: pd.Series, matrix_df: pd.DataFrame, header: Dict[str, Any], project_name: str, logo_bytes: Optional[bytes] = None) -> None:
    """Zeichnet eine zweite PDF-Seite pro Pritsche mit Ladeplan-BSD-Matrix."""
    from reportlab.lib import colors

    pname = str(platform.get('Pritsche', 'Pritsche'))
    _pdf_draw_logo(c, logo_bytes, page_w - margin - 95, page_h - 78, 85, 45)
    c.setFont('Helvetica-Bold', 16)
    c.drawString(margin, page_h - 36, f'Ladeplan BSD - {pname}')
    c.setFont('Helvetica', 9)
    c.drawString(margin, page_h - 54, f'Objekt / Datei: {header.get("Objekt_Name", project_name) or project_name}')
    c.drawString(margin, page_h - 70, f'Erstellt: {datetime.now().strftime("%d.%m.%Y %H:%M")}')

    # Kopfbereich links/rechts ähnlich Tabellen-Ladeplan.
    left_x = margin
    right_x = page_w / 2 + 20
    top_y = page_h - 100
    c.setStrokeColor(colors.black)
    c.rect(left_x, top_y - 125, page_w / 2 - margin - 35, 125, stroke=1, fill=0)
    c.rect(right_x, top_y - 125, page_w / 2 - margin - 20, 125, stroke=1, fill=0)

    c.setFont('Helvetica-Bold', 10)
    c.drawString(left_x + 8, top_y - 18, 'Pritsche:')
    c.drawString(right_x + 8, top_y - 18, 'Pritschen- / Frachtdaten')
    c.setFont('Helvetica', 8)
    left_lines = [
        f'Fuhre: {header.get("Fuhre_Nr", "")}',
        f'Fuhrenoption: {header.get("Fuhrenoption", "")}',
        f'Pritschenname: {header.get("Pritschenname", "")}',
        f'Datum: {datetime.now().strftime("%d.%m.%Y")}',
    ]
    right_lines = [
        f'Pritschenhöhe: {safe_number(header.get("Pritschenhöhe_mm")):.0f} mm',
        f'Pritschenbreite: {safe_number(header.get("Pritschenbreite_mm")):.0f} mm',
        f'Frachthöhe: {safe_number(header.get("Frachthöhe_mm")):.0f} mm',
        f'Höhe gesamt: {safe_number(header.get("Höhe_gesamt_mm")):.0f} mm',
        f'Länge gesamt: {safe_number(header.get("Länge_gesamt_mm")):.0f} mm',
        f'Breite gesamt: {safe_number(header.get("Breite_gesamt_mm")):.0f} mm',
        f'Ladegewicht: {safe_number(header.get("Ladegewicht_kg")):.0f} kg',
        f'Eigengewicht Pritsche: {safe_number(header.get("Eigengewicht_Pritsche_kg")):.0f} kg',
        f'Gesamtgewicht: {safe_number(header.get("Gesamtgewicht_kg")):.0f} kg',
    ]
    for i, line in enumerate(left_lines):
        c.drawString(left_x + 8, top_y - 38 - i * 14, line)
    c.setFont('Helvetica', 7.2)
    for i, line in enumerate(right_lines):
        c.drawString(right_x + 8, top_y - 34 - i * 10.5, line)

    c.setFont('Helvetica', 6.4)
    c.drawString(left_x + 8, top_y - 110, 'Bezug: X = vorne/hinten, Y = links/rechts; X-mittig = vorne im BSD')

    table_y = top_y - 155
    table_x = margin
    row_h = 20
    # A3 quer: kompakte Spaltenbreiten.
    col_defs = [
        ('Lage', 38),
        ('Vorne links', 128),
        ('Vorne rechts', 128),
        ('Hinten links', 128),
        ('Hinten rechts', 128),
        ('Höhe_mm', 56),
        ('Breite_mm', 62),
        ('Gesamtlänge_mm', 78),
        ('Gewicht_kg', 64),
    ]
    total_w = sum(w for _, w in col_defs)

    c.setFillColor(colors.lightgrey)
    c.rect(table_x, table_y, total_w, row_h, stroke=1, fill=1)
    c.setFillColor(colors.black)
    c.setFont('Helvetica-Bold', 7)
    cx = table_x
    for col, w in col_defs:
        c.rect(cx, table_y, w, row_h, stroke=1, fill=0)
        c.drawString(cx + 3, table_y + 7, col.replace('_mm', ' mm').replace('_kg', ' kg'))
        cx += w

    c.setFont('Helvetica', 6)
    y = table_y - row_h
    max_rows = int((table_y - 35) / row_h)
    rows = matrix_df.head(max_rows).copy() if matrix_df is not None and not matrix_df.empty else pd.DataFrame()
    if rows.empty:
        c.drawString(table_x, y + 6, 'Keine Ladeplan-BSD-Daten vorhanden')
        return

    for _, row in rows.iterrows():
        cx = table_x
        for col, w in col_defs:
            c.rect(cx, y, w, row_h, stroke=1, fill=0)
            value = row.get(col, '')
            if col in ['Höhe_mm', 'Breite_mm', 'Gesamtlänge_mm']:
                text = f'{safe_number(value):.0f}' if safe_number(value) else ''
            elif col == 'Gewicht_kg':
                text = f'{safe_number(value):.0f}' if safe_number(value) else ''
            else:
                text = str(value)
            if '\n' in text:
                lines = [v.strip() for v in text.split('\n') if v.strip()][:3]
                c.setFont('Helvetica', 5.2)
                for li, line in enumerate(lines):
                    line = line[:32] if w >= 100 else line[:11]
                    c.drawString(cx + 3, y + row_h - 7 - li * 5.5, line)
                c.setFont('Helvetica', 6)
            else:
                text = text[:35] if w >= 100 else text[:12]
                c.drawString(cx + 3, y + 7, text)
            cx += w
        y -= row_h
        if y < 30:
            break




def _pdf_draw_logo(c, logo_bytes: Optional[bytes], x: float, y: float, max_w: float = 90, max_h: float = 55) -> None:
    """Zeichnet optional ein Logo in die PDF-Seite."""
    if not logo_bytes:
        return
    try:
        from reportlab.lib.utils import ImageReader
        img = ImageReader(io.BytesIO(logo_bytes))
        iw, ih = img.getSize()
        if not iw or not ih:
            return
        scale = min(max_w / float(iw), max_h / float(ih))
        draw_w = iw * scale
        draw_h = ih * scale
        c.drawImage(img, x, y, width=draw_w, height=draw_h, preserveAspectRatio=True, mask='auto')
    except Exception:
        return

def _pdf_draw_bsd_reference_sketch(c, x: float, y: float, w: float = 170, h: float = 72, front_at_x_max: bool = True, left_at_y_max: bool = False) -> None:
    """Kleine Bezugs-Skizze: Welche Seite der Draufsicht gehört zu BSD VL/VR/HL/HR."""
    from reportlab.lib import colors
    c.saveState()
    c.setStrokeColor(colors.black)
    c.setFillColor(colors.white)
    c.setLineWidth(0.6)
    c.rect(x, y, w, h, stroke=1, fill=0)

    c.setFont('Helvetica-Bold', 6.7)
    c.setFillColor(colors.black)
    c.drawString(x + 5, y + h - 10, 'Bezug Draufsicht / BSD')

    pad_x = 34
    pad_y = 17
    bx = x + pad_x
    by = y + pad_y
    bw = w - pad_x * 2
    bh = h - pad_y - 16
    mid_x = bx + bw / 2
    mid_y = by + bh / 2

    c.setLineWidth(0.7)
    c.rect(bx, by, bw, bh, stroke=1, fill=0)
    c.setDash(2, 2)
    c.line(mid_x, by, mid_x, by + bh)
    c.line(bx, mid_y, bx + bw, mid_y)
    c.setDash()

    # Aktuelle feste Darstellung: X läuft horizontal, Y vertikal.
    left_end = 'Hinten' if front_at_x_max else 'Vorne'
    right_end = 'Vorne' if front_at_x_max else 'Hinten'
    top_side = 'Links' if left_at_y_max else 'Rechts'
    bottom_side = 'Rechts' if left_at_y_max else 'Links'

    c.setFont('Helvetica', 5.8)
    c.drawRightString(bx - 3, mid_y - 2, left_end)
    c.drawString(bx + bw + 3, mid_y - 2, right_end)
    c.drawCentredString(mid_x, by + bh + 3, top_side)
    c.drawCentredString(mid_x, by - 8, bottom_side)

    c.setFont('Helvetica-Bold', 6.0)
    c.drawCentredString((bx + mid_x) / 2, (mid_y + by + bh) / 2 - 2, 'HL' if left_end == 'Hinten' and top_side == 'Links' else 'HR')
    c.drawCentredString((mid_x + bx + bw) / 2, (mid_y + by + bh) / 2 - 2, 'VL' if right_end == 'Vorne' and top_side == 'Links' else 'VR')
    c.drawCentredString((bx + mid_x) / 2, (by + mid_y) / 2 - 2, 'HL' if left_end == 'Hinten' and bottom_side == 'Links' else 'HR')
    c.drawCentredString((mid_x + bx + bw) / 2, (by + mid_y) / 2 - 2, 'VL' if right_end == 'Vorne' and bottom_side == 'Links' else 'VR')

    c.setFont('Helvetica', 5.4)
    c.drawString(x + 5, y + 4, 'kurze Stirnseiten = vorne / hinten')
    c.restoreState()


def create_loading_pdf(
    placements_df: pd.DataFrame,
    platforms_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    warnings_df: pd.DataFrame,
    project_name: str = 'BVX Verladeplanung',
    project_meta: Optional[Dict[str, Any]] = None,
    logo_bytes: Optional[bytes] = None,
) -> bytes:
    """Erstellt einen einfachen A3-Pritschenplan als PDF pro Pritsche."""
    try:
        from reportlab.lib.pagesizes import A3, landscape
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
    except ImportError as exc:
        raise RuntimeError('Für den PDF-Export muss reportlab installiert sein: pip install reportlab') from exc

    project_meta = project_meta or {}
    front_at_x_max, left_at_y_max = _orientation_flags_from_meta(project_meta)
    output = io.BytesIO()
    c = canvas.Canvas(output, pagesize=landscape(A3))
    page_w, page_h = landscape(A3)
    margin = 24

    if platforms_df is None or platforms_df.empty:
        c.drawString(margin, page_h - margin, 'Keine Pritschen vorhanden')
        c.save()
        return output.getvalue()

    for _, platform in platforms_df.iterrows():
        pname = str(platform.get('Pritsche', 'Pritsche'))
        srow = summary_df[summary_df['Pritsche'].astype(str) == pname]
        srow = srow.iloc[0] if not srow.empty else pd.Series(dtype=object)

        _pdf_draw_logo(c, logo_bytes, page_w - 120, page_h - 70, 90, 50)
        c.setFont('Helvetica-Bold', 16)
        c.drawString(margin, page_h - 36, f'Pritschenplan - {pname}')
        c.setFont('Helvetica', 9)
        c.drawString(margin, page_h - 54, f'Objekt / Datei: {project_meta.get("Objekt_Name", project_name) or project_name}')
        c.drawString(margin, page_h - 70, f'Erstellt: {datetime.now().strftime("%d.%m.%Y %H:%M")}')

        info_x = page_w - 245
        info_y = page_h - 40
        c.setFont('Helvetica-Bold', 10)
        c.drawString(info_x, info_y, 'Info Pritsche')
        c.setFont('Helvetica', 8)
        info_lines = [
            f'Transport: {project_meta.get("Transport_Name", platform.get("Fuhrenoption", ""))}',
            f'Fuhrenoption: {platform.get("Fuhrenoption", "")}',
            f'Länge genutzt: {safe_number(srow.get("Länge genutzt_mm")):.0f} / {safe_number(srow.get("Max Länge effektiv_mm")):.0f} mm',
            f'Breite genutzt: {safe_number(srow.get("Breite genutzt_mm")):.0f} / {safe_number(srow.get("Max Breite_mm")):.0f} mm',
            f'Höhe genutzt: {safe_number(srow.get("Höhe genutzt_mm")):.0f} / {safe_number(srow.get("Max Höhe_mm")):.0f} mm',
            f'Ladegewicht: {safe_number(srow.get("Gewicht genutzt_kg")):.0f} kg',
            f'Eigengewicht: {safe_number(srow.get("Eigengewicht Pritsche_kg")):.0f} kg',
            f'Gesamtgewicht: {safe_number(srow.get("Gesamtgewicht inkl. Pritsche_kg")):.0f} / {safe_number(srow.get("Max Gewicht_kg")):.0f} kg',
        ]
        for i, line in enumerate(info_lines):
            c.drawString(info_x, info_y - 16 - i * 13, line)

        hint_x = margin
        hint_y = page_h - 105
        c.setFont('Helvetica-Bold', 9)
        c.drawString(hint_x, hint_y, 'Infos zur Verladung')
        c.setFont('Helvetica', 8)
        hints = [
            '- Unterleghölzer gemäss Einstellung einlegen',
            '- Bunde / Bauteile gemäss Plan laden',
            '- Sichtseite / Schutz gemäss Projektvorgabe beachten',
            '- Ladung sichern und Verladereihenfolge prüfen',
        ]
        for i, line in enumerate(hints):
            c.drawString(hint_x, hint_y - 14 - i * 12, line)

        # Kleine fixe Bezugsskizze, damit Draufsicht, Seitenansichten und BSD gleich gelesen werden.
        _pdf_draw_bsd_reference_sketch(c, margin + 500, page_h - 165, 180, 78, front_at_x_max=front_at_x_max, left_at_y_max=left_at_y_max)

        # Zeichnungsbereiche: mehr Bezug zwischen Seiten- und Vorder/Rückansicht
        _pdf_draw_view(c, placements_df, platform, margin, 395, 710, 220, 'side_left', 'Linke Seitenansicht', front_at_x_max=front_at_x_max, left_at_y_max=left_at_y_max)
        _pdf_draw_view(c, placements_df, platform, margin, 155, 710, 220, 'side_right', 'Rechte Seitenansicht', front_at_x_max=front_at_x_max, left_at_y_max=left_at_y_max)
        _pdf_draw_view(c, placements_df, platform, margin + 745, 395, 330, 200, 'back', 'Rückansicht', front_at_x_max=front_at_x_max, left_at_y_max=left_at_y_max)
        _pdf_draw_view(c, placements_df, platform, margin + 745, 155, 330, 200, 'front', 'Vorderansicht', front_at_x_max=front_at_x_max, left_at_y_max=left_at_y_max)
        _pdf_draw_view(c, placements_df, platform, margin, 20, 1080, 120, 'top', 'Draufsicht', front_at_x_max=front_at_x_max, left_at_y_max=left_at_y_max)

        # Qualitätssicherung kompakt oben rechts, getrennt vom Infofeld.
        c.setStrokeColor(colors.black)
        c.rect(page_w - 245, page_h - 200, 210, 52, stroke=1, fill=0)
        c.setFont('Helvetica', 8)
        c.drawString(page_w - 235, page_h - 163, 'Qualitätssicherung')
        c.drawString(page_w - 235, page_h - 180, 'Datum: ______________')
        c.drawString(page_w - 125, page_h - 180, 'Visum: __________')

        c.showPage()

        # Zweite Seite: Ladeplan BSD Matrix je Pritsche, ähnlich Excel-Beispiel PB 6.
        matrix = create_bsd_matrix_for_platform(placements_df, platform, front_at_x_max=front_at_x_max, left_at_y_max=left_at_y_max)
        if not matrix.empty:
            header = create_bsd_header_for_platform(platform, summary_df, warnings_df, project_meta)
            _pdf_draw_bsd_matrix_page(c, page_w, page_h, margin, platform, matrix, header, project_meta.get('Objekt_Name', project_name) or project_name, logo_bytes=logo_bytes)
            c.showPage()

    c.save()
    return output.getvalue()


# =============================================================================
# Streamlit Bereiche
# =============================================================================

def render_analysis_module(uploaded_file) -> None:
    if uploaded_file is None:
        st.info('Bitte laden Sie eine BVX-Datei in der Seitenleiste hoch.')
        col1, col2 = st.columns(2)
        with col1:
            st.markdown('''
            ### Unterstützte Formate
            - XML Format BVX
            - Text Format BVX
            ''')
        with col2:
            st.markdown('''
            ### Erkannte Operationen
            - Bohrungen
            - Fräsungen
            - Sägeschnitte
            - weitere Operationen
            ''')
        return

    content = read_uploaded_text(uploaded_file)
    parser = BVXParser()
    result = parser.parse(content, uploaded_file.name)

    st.header('Übersicht')
    col1, col2, col3, col4 = st.columns(4)
    col1.metric(label='Bauteile', value=result.part_count)
    col2.metric(label='Operationstypen', value=result.operation_count, delta=f'{result.total_operation_count} gesamt')
    col3.metric(label='Bauteilvolumen', value=format_volume(result.total_volume))
    col4.metric(label='Bearbeitetes Volumen', value=format_volume(result.machined_volume))

    if result.part_dimensions:
        st.subheader('Bauteilabmessungen erstes Bauteil')
        dims = result.part_dimensions
        col1, col2, col3 = st.columns(3)
        col1.info(f"**Länge:** {dims['length']:.1f} mm")
        col2.info(f"**Breite:** {dims['width']:.1f} mm")
        col3.info(f"**Höhe:** {dims['height']:.1f} mm")

    tab1, tab2, tab3, tab4, tab5 = st.tabs(['Bauteile', 'Operationstabelle', 'Diagramme', 'Positionsübersicht', 'Excel Export'])

    with tab1:
        st.subheader('Bauteilliste')
        parts_df = parts_to_dataframe(result.parts, density_kg_m3=500)
        st.dataframe(parts_df, use_container_width=True, hide_index=True)

    with tab2:
        st.subheader('Bearbeitungen Details')
        col1, col2 = st.columns([1, 2])
        with col1:
            operation_types = sorted(set(op.op_type for op in result.operations))
            selected_type = st.selectbox('Nach Typ filtern', ['Alle'] + operation_types, key='filter_type')
        with col2:
            search_term = st.text_input('Suchen...', key='search', placeholder='Typ, Durchmesser oder Tiefe')

        filtered_ops = result.operations
        if selected_type != 'Alle':
            filtered_ops = [op for op in filtered_ops if op.op_type == selected_type]
        if search_term:
            search_lower = search_term.lower()
            filtered_ops = [op for op in filtered_ops if
                            search_lower in op.op_type.lower() or
                            (op.diameter and search_lower in str(op.diameter)) or
                            (op.depth and search_lower in str(op.depth))]

        df_data = []
        for op in filtered_ops:
            df_data.append({
                'Typ': op.op_type,
                'Anzahl': op.count,
                'Durchmesser (mm)': f'{op.diameter:.1f}' if op.diameter else '-',
                'Länge/Tiefe (mm)': f'{op.depth:.1f}' if op.depth else '-',
                'Volumen': format_volume(op.volume),
            })
        df = pd.DataFrame(df_data)
        if not df.empty:
            st.dataframe(df, use_container_width=True, hide_index=True)
            total_ops = sum(op.count for op in filtered_ops)
            st.caption(f'Zeigt {len(filtered_ops)} Operationstypen ({total_ops} Operationen gesamt)')
        else:
            st.warning('Keine Operationen gefunden.')

    with tab3:
        st.subheader('Visualisierungen')
        col1, col2 = st.columns(2)
        with col1:
            type_counts = Counter(op.op_type for op in result.operations for _ in range(op.count))
            if type_counts:
                fig_pie = px.pie(values=list(type_counts.values()), names=list(type_counts.keys()), title='Operationen nach Typ', hole=0.4)
                st.plotly_chart(fig_pie, use_container_width=True)
        with col2:
            volume_by_type: Dict[str, float] = {}
            for op in result.operations:
                volume_by_type[op.op_type] = volume_by_type.get(op.op_type, 0) + op.volume
            if volume_by_type:
                fig_bar = px.bar(
                    x=list(volume_by_type.keys()),
                    y=[v * 1_000_000 for v in volume_by_type.values()],
                    title='Volumen nach Operationstyp (mm³)',
                    labels={'x': 'Operationstyp', 'y': 'Volumen (mm³)'},
                )
                st.plotly_chart(fig_bar, use_container_width=True)

        diameters = [op.diameter for op in result.operations if op.diameter and op.diameter > 0]
        if diameters:
            fig_hist = px.histogram(x=diameters, nbins=20, title='Verteilung der Durchmesser', labels={'x': 'Durchmesser (mm)', 'y': 'Anzahl'})
            st.plotly_chart(fig_hist, use_container_width=True)

    with tab4:
        st.subheader('2D Positionsübersicht')
        ops_with_pos = [op for op in result.operations if op.x is not None and op.y is not None]
        if ops_with_pos:
            fig_scatter = go.Figure()
            for op_type in set(op.op_type for op in ops_with_pos):
                type_ops = [op for op in ops_with_pos if op.op_type == op_type]
                fig_scatter.add_trace(go.Scatter(
                    x=[op.x for op in type_ops],
                    y=[op.y for op in type_ops],
                    mode='markers',
                    name=op_type,
                    marker=dict(size=10),
                    text=[f'{op.op_type}<br>D: {op.diameter or "-"} mm' for op in type_ops],
                    hovertemplate='<b>%{text}</b><br>X: %{x}<br>Y: %{y}<extra></extra>'
                ))
            if result.part_dimensions:
                dims = result.part_dimensions
                fig_scatter.add_shape(type='rect', x0=0, y0=0, x1=dims['length'], y1=dims['width'], line=dict(dash='dash'))
            fig_scatter.update_layout(title='Operationspositionen auf Bauteil', xaxis_title='X Position (mm)', yaxis_title='Y Position (mm)', showlegend=True, height=500)
            st.plotly_chart(fig_scatter, use_container_width=True)
        else:
            st.info('Keine Positionsdaten in der BVX-Datei vorhanden.')

    with tab5:
        st.subheader('Daten exportieren')
        df_export = pd.DataFrame([{
            'Typ': op.op_type,
            'Anzahl': op.count,
            'Durchmesser_mm': op.diameter,
            'Tiefe_mm': op.depth,
            'Laenge_mm': op.length,
            'Volumen_m3': op.volume,
            'X': op.x,
            'Y': op.y,
            'Z': op.z,
        } for op in result.operations])
        parts_df = parts_to_dataframe(result.parts, density_kg_m3=500)

        col1, col2 = st.columns(2)
        with col1:
            csv = df_export.to_csv(index=False)
            st.download_button(
                label='CSV herunterladen',
                data=csv,
                file_name=f"{result.file_name.replace('.bvx', '').replace('.BVX', '')}_analyse.csv",
                mime='text/csv',
            )
        with col2:
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                parts_df.to_excel(writer, sheet_name='Bauteile', index=False)
                df_export.to_excel(writer, sheet_name='Operationen', index=False)
                summary_data = {
                    'Eigenschaft': ['Dateiname', 'Anzahl Bauteile', 'Anzahl Operationstypen', 'Gesamtoperationen', 'Bauteilvolumen (m³)', 'Bearbeitetes Volumen (m³)'],
                    'Wert': [result.file_name, result.part_count, result.operation_count, result.total_operation_count, result.total_volume, result.machined_volume]
                }
                pd.DataFrame(summary_data).to_excel(writer, sheet_name='Zusammenfassung', index=False)
            st.download_button(
                label='Excel herunterladen',
                data=output.getvalue(),
                file_name=f"{result.file_name.replace('.bvx', '').replace('.BVX', '')}_analyse.xlsx",
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )




def _control_extract_numbers(value: Any) -> List[int]:
    """Extrahiert numerische Bauteilnummern für Kontrolltabellen."""
    text = str(value or '')
    nums: List[int] = []
    for m in re.findall(r'\d+', text):
        try:
            nums.append(int(m))
        except Exception:
            pass
    return nums


def _control_nums_from_row(row: pd.Series) -> List[int]:
    labels = _split_bsd_text_list(row.get('Bauteile_Liste', ''), row.get('Bauteile', row.get('Ansicht_Liste', '')))
    nums: List[int] = []
    for label in labels:
        nums.extend(_control_extract_numbers(label))
    if not nums:
        nums = _control_extract_numbers(row.get('Ansicht_Label', row.get('Bauteile', '')))
    return nums


def _control_join_nums(nums: List[int], max_items: int = 12) -> str:
    if not nums:
        return ''
    if len(nums) <= max_items:
        return ' / '.join(str(n) for n in nums)
    return ' / '.join(str(n) for n in nums[:max_items]) + ' / ...'


def _control_platform_order(platforms_df: pd.DataFrame) -> Dict[str, int]:
    order: Dict[str, int] = {}
    if platforms_df is None or platforms_df.empty or 'Pritsche' not in platforms_df.columns:
        return order
    for idx, (_, row) in enumerate(platforms_df.reset_index(drop=True).iterrows(), start=1):
        order[str(row.get('Pritsche', ''))] = idx
    return order


def build_control_bundle_table(units_df: pd.DataFrame) -> pd.DataFrame:
    """Kontrolltabelle: Welche Bauteile stecken in welchem Bund/Verladeeinheit?"""
    if units_df is None or units_df.empty:
        return pd.DataFrame()
    rows: List[Dict[str, Any]] = []
    for idx, row in units_df.reset_index(drop=True).iterrows():
        nums = _control_nums_from_row(row)
        is_ascending = nums == sorted(nums) if nums else True
        rows.append({
            'Reihenfolge': idx + 1,
            'Einheit_ID': row.get('Einheit_ID', ''),
            'Typ': row.get('Typ', ''),
            'Anzahl_Bauteile': int(safe_number(row.get('Anzahl_Bauteile'), 0)),
            'Bauteile_im_Bund': _control_join_nums(nums),
            'Nummer_min': min(nums) if nums else '',
            'Nummer_max': max(nums) if nums else '',
            'Bund_intern_aufsteigend': 'OK' if is_ascending else 'prüfen',
            'Länge_mm': safe_number(row.get('Länge_mm')),
            'Breite_mm': safe_number(row.get('Breite_mm')),
            'Höhe_mm': safe_number(row.get('Höhe_mm')),
            'Gewicht_kg': safe_number(row.get('Gewicht_kg')),
            'Warnung': row.get('Warnung', ''),
        })
    return pd.DataFrame(rows)


def build_control_assignment_table(placements_df: pd.DataFrame, platforms_df: pd.DataFrame) -> pd.DataFrame:
    """Kontrolltabelle: Welche fortlaufenden Nummernbereiche landen auf welcher Pritsche?"""
    if placements_df is None or placements_df.empty:
        return pd.DataFrame()
    porder = _control_platform_order(platforms_df)
    loaded = placements_df[placements_df.get('Pritsche', '').astype(str) != 'NICHT VERLADEN'].copy()
    if loaded.empty:
        return pd.DataFrame()
    rows: List[Dict[str, Any]] = []
    last_max: Optional[int] = None
    for pname in sorted(loaded['Pritsche'].astype(str).unique(), key=lambda p: porder.get(p, 999999)):
        group = loaded[loaded['Pritsche'].astype(str) == pname]
        nums: List[int] = []
        for _, r in group.iterrows():
            nums.extend(_control_nums_from_row(r))
        nums_unique = sorted(set(nums))
        min_num = min(nums_unique) if nums_unique else None
        max_num = max(nums_unique) if nums_unique else None
        ruecksprung = bool(last_max is not None and min_num is not None and min_num <= last_max)
        missing_inside = ''
        if min_num is not None and max_num is not None:
            expected = set(range(min_num, max_num + 1))
            missing = sorted(expected.difference(set(nums_unique)))
            if missing:
                missing_inside = _control_join_nums(missing, 20)
        rows.append({
            'Pritschen_Reihenfolge': porder.get(pname, 999999),
            'Pritsche': pname,
            'Fuhre_Nr': int(safe_number(group['Fuhre_Nr'].iloc[0], 0)) if 'Fuhre_Nr' in group.columns else '',
            'Einheiten': int(group['Einheit_ID'].nunique()) if 'Einheit_ID' in group.columns else len(group),
            'Bauteilnummer_von': min_num if min_num is not None else '',
            'Bauteilnummer_bis': max_num if max_num is not None else '',
            'Rücksprung_zur_vorigen_Pritsche': 'prüfen' if ruecksprung else 'OK',
            'Fehlende_Nummern_innerhalb_Bereich': missing_inside,
            'Ladegewicht_kg': round(float(pd.to_numeric(group.get('Gewicht_kg'), errors='coerce').fillna(0).sum()), 1),
        })
        if max_num is not None:
            last_max = max(last_max or max_num, max_num)
    return pd.DataFrame(rows)


def build_control_layer_table(placements_df: pd.DataFrame, platforms_df: pd.DataFrame) -> pd.DataFrame:
    """Kontrolltabelle: Lage, Position und Reihenfolge je Pritsche."""
    if placements_df is None or placements_df.empty:
        return pd.DataFrame()
    porder = _control_platform_order(platforms_df)
    platform_lookup = {}
    if platforms_df is not None and not platforms_df.empty:
        for _, p in platforms_df.iterrows():
            platform_lookup[str(p.get('Pritsche', ''))] = p
    rows: List[Dict[str, Any]] = []
    loaded = placements_df[placements_df.get('Pritsche', '').astype(str) != 'NICHT VERLADEN'].copy()
    for pname in sorted(loaded['Pritsche'].astype(str).unique(), key=lambda p: porder.get(p, 999999)):
        group = loaded[loaded['Pritsche'].astype(str) == pname].copy()
        if group.empty:
            continue
        group['Z_mm'] = pd.to_numeric(group.get('Z_mm'), errors='coerce').fillna(0.0)
        unique_z_bottom = sorted(group['Z_mm'].unique())
        bottom_rank = {z: i + 1 for i, z in enumerate(unique_z_bottom)}
        top_rank = {z: len(unique_z_bottom) - i for i, z in enumerate(unique_z_bottom)}
        platform = platform_lookup.get(pname)
        p_width = safe_number(platform.get('Breite_mm')) if platform is not None else safe_number(group.get('Breite_mm', pd.Series([0])).max())
        p_center_y = p_width / 2 if p_width else 0.0
        sort_cols = ['Z_mm', 'Y_mm', 'X_mm']
        group = group.sort_values([c for c in sort_cols if c in group.columns], ascending=[False, True, True][:len([c for c in sort_cols if c in group.columns])], kind='stable')
        for _, r in group.iterrows():
            nums = _control_nums_from_row(r)
            y = safe_number(r.get('Y_mm'))
            w = safe_number(r.get('Breite_mm'))
            y_center = y + w / 2
            if p_width and w >= p_width * 0.82:
                y_pos = 'volle Breite / mittig'
            elif p_width and y_center < p_center_y:
                y_pos = 'links'
            elif p_width:
                y_pos = 'rechts'
            else:
                y_pos = ''
            z = safe_number(r.get('Z_mm'))
            rows.append({
                'Pritschen_Reihenfolge': porder.get(pname, 999999),
                'Pritsche': pname,
                'Lage_von_oben': top_rank.get(z, ''),
                'Lage_von_unten': bottom_rank.get(z, ''),
                'Z_mm': z,
                'Einheit_ID': r.get('Einheit_ID', ''),
                'Typ': r.get('Typ', ''),
                'Bauteile': _control_join_nums(nums),
                'Nummer_min': min(nums) if nums else '',
                'Nummer_max': max(nums) if nums else '',
                'Y_Position': y_pos,
                'X_mm': safe_number(r.get('X_mm')),
                'Y_mm': y,
                'Länge_mm': safe_number(r.get('Länge_mm')),
                'Breite_mm': w,
                'Höhe_mm': safe_number(r.get('Höhe_mm')),
                'Gewicht_kg': safe_number(r.get('Gewicht_kg')),
            })
    return pd.DataFrame(rows)


def build_control_issue_table(bundle_control_df: pd.DataFrame, assignment_control_df: pd.DataFrame, layer_control_df: pd.DataFrame) -> pd.DataFrame:
    """Erzeugt eine kurze, fachliche Prüfliste vor dem PDF-Export."""
    issues: List[Dict[str, Any]] = []
    if bundle_control_df is not None and not bundle_control_df.empty and 'Bund_intern_aufsteigend' in bundle_control_df.columns:
        for _, r in bundle_control_df[bundle_control_df['Bund_intern_aufsteigend'] == 'prüfen'].iterrows():
            issues.append({'Bereich': 'Bundbildung', 'Pritsche': '', 'Schwere': 'mittel', 'Hinweis': f"Einheit {r.get('Einheit_ID')} ist intern nicht aufsteigend: {r.get('Bauteile_im_Bund')}"})
    if assignment_control_df is not None and not assignment_control_df.empty:
        for _, r in assignment_control_df[assignment_control_df['Rücksprung_zur_vorigen_Pritsche'] == 'prüfen'].iterrows():
            issues.append({'Bereich': 'Pritschen-Zuteilung', 'Pritsche': r.get('Pritsche', ''), 'Schwere': 'hoch', 'Hinweis': 'Bauteilnummern springen gegenüber der vorherigen Pritsche zurück.'})
        for _, r in assignment_control_df[assignment_control_df['Fehlende_Nummern_innerhalb_Bereich'].astype(str) != ''].iterrows():
            issues.append({'Bereich': 'Pritschen-Zuteilung', 'Pritsche': r.get('Pritsche', ''), 'Schwere': 'mittel', 'Hinweis': f"Nummernbereich ist nicht geschlossen. Fehlend: {r.get('Fehlende_Nummern_innerhalb_Bereich')}"})
    if layer_control_df is not None and not layer_control_df.empty:
        for pname, grp in layer_control_df.groupby('Pritsche'):
            g = grp.copy()
            g['Lage_von_oben_num'] = pd.to_numeric(g.get('Lage_von_oben'), errors='coerce')
            top = g[g['Lage_von_oben_num'] == 1]
            lower = g[g['Lage_von_oben_num'] > 1]
            top_nums = pd.to_numeric(top.get('Nummer_min'), errors='coerce').dropna()
            lower_nums = pd.to_numeric(lower.get('Nummer_min'), errors='coerce').dropna()
            if not top_nums.empty and not lower_nums.empty and float(top_nums.min()) > float(lower_nums.min()):
                issues.append({'Bereich': 'Lagenrichtung', 'Pritsche': pname, 'Schwere': 'hoch', 'Hinweis': 'Niedrigste Nummer liegt nicht oben. Stapelrichtung prüfen.'})
    return pd.DataFrame(issues)



def _manual_plan_signature(placements_df: pd.DataFrame) -> str:
    """Signatur, damit manuelle Korrekturen nur zurückgesetzt werden, wenn der Automatikplan neu ist."""
    if placements_df is None or placements_df.empty:
        return 'empty'
    cols = [c for c in ['Einheit_ID', 'Bauteile_Liste', 'Bauteile', 'Pritsche', 'Länge_mm', 'Breite_mm', 'Höhe_mm'] if c in placements_df.columns]
    if not cols:
        return str(len(placements_df))
    try:
        tmp = placements_df[cols].fillna('').astype(str)
        return '|'.join(tmp.apply(lambda r: ';'.join(r.values.tolist()), axis=1).tolist())
    except Exception:
        return str(len(placements_df))



def _manual_platform_signature(platforms_df: pd.DataFrame) -> str:
    """Signatur für verwendete Pritschen, damit manuelle Pritschenwerte stabil bleiben."""
    if platforms_df is None or platforms_df.empty:
        return 'empty'
    cols = [c for c in [
        'Pritsche', 'Fuhre_Nr', 'Pritschenname', 'Länge_mm', 'Breite_mm', 'Max_Höhe_mm',
        'Überhang_vorne_mm', 'Überhang_hinten_mm', 'Max_Gewicht_kg',
        'Eigengewicht_Pritsche_kg', 'Kantholz_erste_Lage_mm',
        'Einlage_zwischen_Lagen_mm', 'Einlage_allgemein_mm'
    ] if c in platforms_df.columns]
    try:
        tmp = platforms_df[cols].fillna('').astype(str)
        return '|'.join(tmp.apply(lambda r: ';'.join(r.values.tolist()), axis=1).tolist())
    except Exception:
        return str(len(platforms_df))


def _clean_manual_platforms_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Bereinigt manuell bearbeitete Pritschenparameter."""
    if df is None:
        return pd.DataFrame()
    result = df.copy()
    numeric_cols = [
        'Fuhre_Nr', 'Pritschen_Reihenfolge', 'Länge_mm', 'Breite_mm', 'Max_Höhe_mm',
        'Überhang_vorne_mm', 'Überhang_hinten_mm', 'Max_Gewicht_kg',
        'Eigengewicht_Pritsche_kg', 'Kantholz_erste_Lage_mm',
        'Einlage_zwischen_Lagen_mm', 'Einlage_allgemein_mm'
    ]
    for col in numeric_cols:
        if col in result.columns:
            result[col] = pd.to_numeric(result[col], errors='coerce').fillna(0.0)
    if 'Freigabe' in result.columns:
        result['Freigabe'] = result['Freigabe'].fillna(True)
    return result


def _manual_default_platform_editor_columns(platforms_df: pd.DataFrame) -> List[str]:
    wanted = [
        'Pritsche', 'Fuhre_Nr', 'Pritschenname', 'Länge_mm', 'Breite_mm', 'Max_Höhe_mm',
        'Überhang_vorne_mm', 'Überhang_hinten_mm', 'Max_Gewicht_kg',
        'Eigengewicht_Pritsche_kg', 'Kantholz_erste_Lage_mm',
        'Einlage_zwischen_Lagen_mm', 'Einlage_allgemein_mm'
    ]
    return [c for c in wanted if platforms_df is not None and c in platforms_df.columns]


def _manual_place_by_position(
    df: pd.DataFrame,
    row_idx: int,
    platform_row: Optional[pd.Series],
    position: str,
    layer_action: str,
    spacer_under: float = 0.0,
    custom_z: Optional[float] = None,
) -> pd.DataFrame:
    """Fachliche manuelle Vergabe: Position + Lage -> X/Y/Z."""
    if df is None or df.empty or row_idx not in df.index:
        return df
    row = df.loc[row_idx]
    length = safe_number(row.get('Länge_mm'), 0.0)
    width = safe_number(row.get('Breite_mm'), 0.0)
    height = safe_number(row.get('Höhe_mm'), 0.0)
    x_now = safe_number(row.get('X_mm'), 0.0)
    y_now = safe_number(row.get('Y_mm'), 0.0)
    z_now = safe_number(row.get('Z_mm'), 0.0)

    eff_len = _manual_effective_length(platform_row)
    p_width = safe_number(platform_row.get('Breite_mm'), 0.0) if platform_row is not None else 0.0
    base_z = safe_number(platform_row.get('Kantholz_erste_Lage_mm'), 0.0) if platform_row is not None else 0.0
    layer_spacer = safe_number(platform_row.get('Einlage_zwischen_Lagen_mm'), 0.0) if platform_row is not None else 0.0
    center_y = p_width / 2.0 if p_width > 0 else 0.0

    pos = str(position or '').lower()
    x_new = x_now
    y_new = y_now

    # X: vorne/hinten/mittig
    if 'vorne' in pos:
        x_new = max(0.0, eff_len - length) if eff_len > 0 else x_now
    elif 'hinten' in pos:
        x_new = 0.0
    elif 'x-mitte' in pos or 'x mitte' in pos or 'x mittig' in pos:
        x_new = max(0.0, (eff_len - length) / 2.0) if eff_len > 0 else x_now

    # Y: links/rechts/mittig
    if 'links' in pos:
        y_new = max(0.0, center_y - width) if p_width > 0 else y_now
    elif 'rechts' in pos:
        y_new = min(max(0.0, p_width - width), center_y) if p_width > 0 else y_now
    elif 'y-mitte' in pos or 'y mitte' in pos or 'y mittig' in pos or pos.strip() == 'mitte':
        y_new = max(0.0, (p_width - width) / 2.0) if p_width > 0 else y_now

    # Z: automatisch auf bestehende Lage in dieser Zone.
    # Keine Schaltfläche "eine Lage höher": die App sucht die höchste belegte Auflage
    # an der gewählten Position und legt das Teil/Bund darauf ab.
    z_new = z_now
    action = str(layer_action or '').lower()
    spacer_val = safe_number(spacer_under, 0.0)

    def _overlap_area(ax, ay, al, ab, bx, by, bl, bb) -> float:
        ox = max(0.0, min(ax + al, bx + bl) - max(ax, bx))
        oy = max(0.0, min(ay + ab, by + bb) - max(ay, by))
        return ox * oy

    if 'manuell' in action and custom_z is not None:
        z_new = max(0.0, safe_number(custom_z, z_now))
    elif 'ganz unten' in action:
        z_new = base_z + spacer_val
    else:
        # automatisch: auf höchste vorhandene Lage setzen, die unter der neuen
        # Position eine reale Überdeckung hat. Falls keine vorhanden ist: unten.
        best_top = base_z
        try:
            helper_types = {'Unterbau', 'Kantholz', 'Bundeinlage', 'Einlage', 'Lagenholz'}
            same_platform = df[
                df.get('Pritsche', pd.Series(dtype=str)).astype(str).eq(str(platform_row.get('Pritsche', '')))
            ].copy()
            for b_idx, b in same_platform.iterrows():
                if b_idx == row_idx:
                    continue
                if str(b.get('Typ', '')).strip() in helper_types:
                    continue
                if pd.isna(b.get('X_mm')) or pd.isna(b.get('Y_mm')) or pd.isna(b.get('Z_mm')):
                    continue
                ov = _overlap_area(
                    x_new, y_new, length, width,
                    safe_number(b.get('X_mm')), safe_number(b.get('Y_mm')),
                    safe_number(b.get('Länge_mm')), safe_number(b.get('Breite_mm')),
                )
                if ov <= 0:
                    continue
                # Nur echte Auflage zählen. Kleine Kantenberührung ignorieren.
                if ov < max(1.0, length * width * 0.10):
                    continue
                top_b = safe_number(b.get('Z_mm')) + safe_number(b.get('Höhe_mm'))
                best_top = max(best_top, top_b)
        except Exception:
            best_top = base_z
        z_new = best_top + spacer_val

    updates = {
        'X_mm': round(x_new, 1),
        'Y_mm': round(y_new, 1),
        'Z_mm': round(z_new, 1),
    }
    if platform_row is not None:
        updates.update({
            'Pritsche': str(platform_row.get('Pritsche', row.get('Pritsche', ''))),
            'Fuhre_Nr': platform_row.get('Fuhre_Nr', row.get('Fuhre_Nr')),
            'Fuhrenoption': platform_row.get('Fuhrenoption', row.get('Fuhrenoption')),
            'Pritschenname': platform_row.get('Pritschenname', row.get('Pritschenname')),
        })
    if 'Einlage_unten_mm' in df.columns:
        updates['Einlage_unten_mm'] = safe_number(spacer_under, 0.0)
    if 'Bundeinlage_mm' in df.columns:
        updates['Bundeinlage_mm'] = max(safe_number(row.get('Bundeinlage_mm'), 0.0), safe_number(spacer_under, 0.0))
    return _manual_update_row(df, row_idx, **updates)



def _build_manual_platforms_from_excel(pritschen_df: pd.DataFrame, selected_option: str, standards: Dict[str, Any]) -> pd.DataFrame:
    """Erzeugt F01-Pritschen aus der Excel-Pritschen/Fuhren-Vorlage für den manuellen Modus."""
    if pritschen_df is None or pritschen_df.empty:
        return pd.DataFrame()

    rows = pritschen_df.copy()
    if 'Fuhrenoption' in rows.columns and selected_option:
        rows = rows[rows['Fuhrenoption'].astype(str).eq(str(selected_option))].copy()
    if 'Freigabe' in rows.columns:
        try:
            rows = rows[rows['Freigabe'].astype(bool)].copy()
        except Exception:
            pass
    if rows.empty:
        return pd.DataFrame()

    sort_cols = [c for c in ['Pritschen_Reihenfolge', 'Pritschenname'] if c in rows.columns]
    if sort_cols:
        rows = rows.sort_values(sort_cols, kind='stable').reset_index(drop=True)
    else:
        rows = rows.reset_index(drop=True)

    rows['Fuhre_Nr'] = 1
    if 'Fuhrenoption' not in rows.columns:
        rows['Fuhrenoption'] = selected_option
    rows['Fuhrenoption'] = selected_option or rows['Fuhrenoption'].astype(str)
    if 'Pritschenname' not in rows.columns:
        rows['Pritschenname'] = [f'Pritsche {i+1}' for i in range(len(rows))]
    rows['Pritschenname'] = rows['Pritschenname'].astype(str)
    rows['Pritsche'] = rows.apply(lambda r: f"F01 {r['Pritschenname']}", axis=1)
    rows['Manueller_Status'] = 'Leer'

    # Defaults aus Standards ergänzen, falls Excel-Zellen leer sind.
    defaults = {
        'Kantholz_erste_Lage_mm': safe_number(standards.get('Standard_Kantholz_erste_Lage'), 80.0),
        'Einlage_zwischen_Lagen_mm': safe_number(standards.get('Standard_Einlage_zwischen_Lagen'), 40.0),
        'Einlage_allgemein_mm': safe_number(standards.get('Standard_Einlage_allgemein'), 0.0),
        'Überhang_vorne_mm': 0.0,
        'Überhang_hinten_mm': 0.0,
        'Eigengewicht_Pritsche_kg': 0.0,
    }
    for col, default in defaults.items():
        if col not in rows.columns:
            rows[col] = default
        rows[col] = pd.to_numeric(rows[col], errors='coerce').fillna(default)

    return _clean_manual_platforms_dataframe(rows)


def _build_manual_empty_placements_from_units(units_df: pd.DataFrame) -> pd.DataFrame:
    """Manueller Modus: alle Verladeeinheiten starten unverladen und werden per Tabelle gesetzt."""
    if units_df is None or units_df.empty:
        return pd.DataFrame()
    rows: List[Dict[str, Any]] = []
    for _, unit in units_df.reset_index(drop=True).iterrows():
        rows.append({
            'Fuhre_Nr': None,
            'Fuhrenoption': '',
            'Pritschenname': '',
            'Pritsche': 'NICHT VERLADEN',
            'Einheit_ID': unit.get('Einheit_ID', ''),
            'Typ': unit.get('Typ', ''),
            'Anzahl_Bauteile': int(safe_number(unit.get('Anzahl_Bauteile'), 1)),
            'Bauteile': unit.get('Bauteile', ''),
            'Bauteile_Liste': unit.get('Bauteile_Liste', unit.get('Bauteile', '')),
            'Ansicht_Attribut': unit.get('Ansicht_Attribut', ''),
            'Ansicht_Label': unit.get('Ansicht_Label', unit.get('Einheit_ID', '')),
            'Ansicht_Liste': unit.get('Ansicht_Liste', unit.get('Ansicht_Label', '')),
            'Einzellängen_mm': unit.get('Einzellängen_mm', ''),
            'Einzelbreiten_mm': unit.get('Einzelbreiten_mm', ''),
            'Einzelhöhen_mm': unit.get('Einzelhöhen_mm', ''),
            'Einlage_allgemein_mm': safe_number(unit.get('Einlage_allgemein_mm'), 0.0),
            'Bundeinlage_mm': safe_number(unit.get('Bundeinlage_mm'), 0.0),
            'Einlage_unten_mm': 0.0,
            'Einlage_oben_mm': 0.0,
            'X_mm': None,
            'Y_mm': None,
            'Z_mm': None,
            'Länge_mm': safe_number(unit.get('Länge_mm'), 0.0),
            'Breite_mm': safe_number(unit.get('Breite_mm'), 0.0),
            'Höhe_mm': safe_number(unit.get('Höhe_mm'), 0.0),
            'Drehung': 0,
            'Ebene': 'manuell nicht gesetzt',
            'Gewicht_kg': round(safe_number(unit.get('Gewicht_kg'), 0.0), 2),
            'Volumen_m3': safe_number(unit.get('Volumen_m3'), 0.0),
            'Warnung': '',
            'Manuelle_Position': 'vorne links',
            'Manuelle_Lage': 'ganz unten',
            'Z_manuell_mm': 0.0,
        })
    return clean_placements_dataframe(pd.DataFrame(rows))


def _derive_manual_position(row: pd.Series, platform_row: Optional[pd.Series]) -> str:
    """Ermittelt eine grobe Positionsbeschriftung für die manuelle Tabelle."""
    if platform_row is None or row is None or pd.isna(row.get('X_mm')) or pd.isna(row.get('Y_mm')):
        return str(row.get('Manuelle_Position', 'vorne links') or 'vorne links')
    eff_len = _manual_effective_length(platform_row)
    width = safe_number(platform_row.get('Breite_mm'), 0.0)
    x_mid = safe_number(row.get('X_mm')) + safe_number(row.get('Länge_mm')) / 2.0
    y_mid = safe_number(row.get('Y_mm')) + safe_number(row.get('Breite_mm')) / 2.0
    fb = 'vorne' if x_mid >= eff_len / 2.0 else 'hinten'
    lr = 'links' if y_mid <= width / 2.0 else 'rechts'
    return f'{fb} {lr}'


def _apply_manual_assignment_table(
    manual_df: pd.DataFrame,
    assignment_df: pd.DataFrame,
    selected_platform_row: Optional[pd.Series],
    selected_platform: str,
) -> pd.DataFrame:
    """Setzt alle in der Tabelle gewählten Verladeeinheiten auf die ausgewählte Pritsche."""
    if manual_df is None or manual_df.empty or assignment_df is None or assignment_df.empty:
        return manual_df
    result = manual_df.copy()
    id_to_idx = {str(r.get('Einheit_ID')): idx for idx, r in result.iterrows()}

    base_z = safe_number(selected_platform_row.get('Kantholz_erste_Lage_mm'), 0.0) if selected_platform_row is not None else 0.0
    for _, erow in assignment_df.iterrows():
        eid = str(erow.get('Einheit_ID', ''))
        if eid not in id_to_idx:
            continue
        idx = id_to_idx[eid]
        should_load = bool(erow.get('Verladen', False))
        if not should_load:
            if str(result.at[idx, 'Pritsche']) == str(selected_platform):
                result = _manual_update_row(
                    result, idx,
                    Pritsche='NICHT VERLADEN',
                    Fuhre_Nr=None,
                    Fuhrenoption='',
                    Pritschenname='',
                    X_mm=None,
                    Y_mm=None,
                    Z_mm=None,
                    Ebene='manuell nicht gesetzt',
                )
            continue

        position = str(erow.get('Position') or result.at[idx, 'Manuelle_Position'] if 'Manuelle_Position' in result.columns else 'vorne links')
        layer_action = str(erow.get('Lage') or 'ganz unten')
        spacer = safe_number(erow.get('Einlage unten mm'), safe_number(result.at[idx, 'Einlage_unten_mm'] if 'Einlage_unten_mm' in result.columns else 0.0))
        z_manual = safe_number(erow.get('Z manuell mm'), base_z)
        result = _manual_place_by_position(
            result,
            idx,
            selected_platform_row,
            position=position,
            layer_action=layer_action,
            spacer_under=spacer,
            custom_z=z_manual,
        )
        if 'Einlage_oben_mm' not in result.columns:
            result['Einlage_oben_mm'] = 0.0
        result.at[idx, 'Einlage_oben_mm'] = safe_number(erow.get('Einlage oben mm'), 0.0)
        if 'Manuelle_Position' not in result.columns:
            result['Manuelle_Position'] = ''
        if 'Manuelle_Lage' not in result.columns:
            result['Manuelle_Lage'] = ''
        if 'Z_manuell_mm' not in result.columns:
            result['Z_manuell_mm'] = 0.0
        result.at[idx, 'Manuelle_Position'] = position
        result.at[idx, 'Manuelle_Lage'] = layer_action
        result.at[idx, 'Z_manuell_mm'] = z_manual

    return clean_placements_dataframe(result)



_MANUAL_ZONE_LABELS = {
    'VL': 'vorne links',
    'VR': 'vorne rechts',
    'HL': 'hinten links',
    'HR': 'hinten rechts',
}
_MANUAL_ZONE_ORDER = ['VL', 'VR', 'HL', 'HR']


def _manual_zone_label(zone: str) -> str:
    return _MANUAL_ZONE_LABELS.get(str(zone or '').upper(), '')


def _manual_zone_from_label(label: str) -> str:
    txt = str(label or '').lower().strip()
    for k, v in _MANUAL_ZONE_LABELS.items():
        if txt == v:
            return k
    return str(label or '').upper().strip()


def _manual_zone_position(zone: str) -> str:
    return _manual_zone_label(zone) or 'vorne links'


def _manual_side_of_zone(zone: str) -> str:
    z = str(zone or '').upper()
    return 'links' if z in ['VL', 'HL'] else 'rechts'


def _manual_frontback_of_zone(zone: str) -> str:
    z = str(zone or '').upper()
    return 'vorne' if z in ['VL', 'VR'] else 'hinten'


def _manual_reflow_zone_platform(manual_df: pd.DataFrame, platform_row: Optional[pd.Series]) -> pd.DataFrame:
    """Berechnet X/Y/Z aus Zonen VL/VR/HL/HR. Benutzer setzt nur die Zone."""
    if manual_df is None or manual_df.empty or platform_row is None:
        return manual_df
    result = manual_df.copy()
    pname = str(platform_row.get('Pritsche', ''))
    if not pname:
        return result

    for col, default in {
        'Manuelle_Zone': '',
        'Manuelle_Reihenfolge': 0.0,
        'Einlage_unten_mm': 0.0,
        'Einlage_oben_mm': 0.0,
        'Manuelle_Position': '',
    }.items():
        if col not in result.columns:
            result[col] = default

    eff_len = _manual_effective_length(platform_row)
    p_width = safe_number(platform_row.get('Breite_mm'), 0.0)
    base_z = safe_number(platform_row.get('Kantholz_erste_Lage_mm'), 0.0)
    standard_spacer = safe_number(platform_row.get('Einlage_zwischen_Lagen_mm'), 0.0)
    center_y = p_width / 2.0 if p_width > 0 else 0.0

    p_mask = result.get('Pritsche', pd.Series(dtype=str)).astype(str).eq(pname)
    zone_mask = result.get('Manuelle_Zone', pd.Series(dtype=str)).astype(str).str.upper().isin(_MANUAL_ZONE_ORDER)
    rows_p = result[p_mask & zone_mask].copy()
    if rows_p.empty:
        return clean_placements_dataframe(result)

    result['Manuelle_Reihenfolge'] = pd.to_numeric(result['Manuelle_Reihenfolge'], errors='coerce').fillna(0.0)

    # Wenn links nur vorne ODER nur hinten belegt ist, wird in X mittig gesetzt.
    # Sobald vorne und hinten links belegt sind, werden sie nach vorne/hinten gezogen.
    # Dasselbe gilt rechts.
    side_active = {
        'links': {
            'front': not rows_p[rows_p['Manuelle_Zone'].astype(str).str.upper().eq('VL')].empty,
            'back': not rows_p[rows_p['Manuelle_Zone'].astype(str).str.upper().eq('HL')].empty,
        },
        'rechts': {
            'front': not rows_p[rows_p['Manuelle_Zone'].astype(str).str.upper().eq('VR')].empty,
            'back': not rows_p[rows_p['Manuelle_Zone'].astype(str).str.upper().eq('HR')].empty,
        },
    }

    for zone in _MANUAL_ZONE_ORDER:
        zrows = rows_p[rows_p['Manuelle_Zone'].astype(str).str.upper().eq(zone)].copy()
        if zrows.empty:
            continue
        zrows = zrows.sort_values(['Manuelle_Reihenfolge', 'Einheit_ID'], kind='stable')
        side = _manual_side_of_zone(zone)
        fb = _manual_frontback_of_zone(zone)
        both_front_back = bool(side_active[side]['front'] and side_active[side]['back'])

        z_cursor = base_z
        for idx, row in zrows.iterrows():
            length = safe_number(row.get('Länge_mm'), 0.0)
            width = safe_number(row.get('Breite_mm'), 0.0)
            height = safe_number(row.get('Höhe_mm'), 0.0)
            spacer = safe_number(row.get('Einlage_unten_mm'), standard_spacer)

            # Y-Position: links/rechts auf halber Pritsche. Sehr breite Einheiten werden zentriert.
            if p_width > 0 and width >= p_width * 0.75:
                y_new = max(0.0, (p_width - width) / 2.0)
            elif side == 'links':
                y_new = max(0.0, center_y - width)
            else:
                y_new = min(max(0.0, p_width - width), center_y)

            # X-Position: Immer aus der Mitte heraus.
            # Wenn nur eine Zone in dieser Seite aktiv ist: mittig auf Pritschenmitte.
            # Wenn vorne + hinten aktiv sind:
            # - vorne startet an der Mitte und wächst nach vorne
            # - hinten endet an der Mitte und wächst nach hinten
            # Die Pritschenenden sind kein Anschlag für die Grundausrichtung.
            center_x = eff_len / 2.0 if eff_len > 0 else 0.0
            if not both_front_back:
                x_new = max(0.0, (eff_len - length) / 2.0) if eff_len > 0 else 0.0
            elif fb == 'vorne':
                x_new = min(max(0.0, center_x), max(0.0, eff_len - length)) if eff_len > 0 else 0.0
            else:
                x_new = max(0.0, center_x - length) if eff_len > 0 else 0.0

            z_new = max(0.0, z_cursor + spacer)
            updates = {
                'X_mm': round(x_new, 1),
                'Y_mm': round(y_new, 1),
                'Z_mm': round(z_new, 1),
                'Manuelle_Position': _manual_zone_position(zone),
                'Ebene': f'manuell {_manual_zone_position(zone)}',
            }
            result = _manual_update_row(result, idx, **updates)
            z_cursor = z_new + height

    return clean_placements_dataframe(result)


def _manual_reflow_all_zones(manual_df: pd.DataFrame, platforms_df: pd.DataFrame) -> pd.DataFrame:
    if manual_df is None or manual_df.empty or platforms_df is None or platforms_df.empty:
        return manual_df
    result = manual_df.copy()
    for _, prow in platforms_df.iterrows():
        result = _manual_reflow_zone_platform(result, prow)
    return clean_placements_dataframe(result)


def _manual_assign_unit_to_zone(
    manual_df: pd.DataFrame,
    platforms_df: pd.DataFrame,
    unit_id: str,
    platform_row: Optional[pd.Series],
    zone: str,
    spacer_under: float,
) -> pd.DataFrame:
    if manual_df is None or manual_df.empty or platform_row is None:
        return manual_df
    result = manual_df.copy()
    for col, default in {
        'Manuelle_Zone': '',
        'Manuelle_Reihenfolge': 0.0,
        'Einlage_unten_mm': 0.0,
        'Einlage_oben_mm': 0.0,
        'Manuelle_Position': '',
    }.items():
        if col not in result.columns:
            result[col] = default

    ids = result.get('Einheit_ID', pd.Series(dtype=str)).astype(str)
    matches = result.index[ids.eq(str(unit_id))].tolist()
    if not matches:
        return clean_placements_dataframe(result)
    idx = matches[0]

    max_order = pd.to_numeric(result.get('Manuelle_Reihenfolge', pd.Series(dtype=float)), errors='coerce').fillna(0.0).max()
    old_order = safe_number(result.at[idx, 'Manuelle_Reihenfolge'], 0.0)
    if old_order <= 0:
        result.at[idx, 'Manuelle_Reihenfolge'] = float(max_order) + 1.0

    result.at[idx, 'Pritsche'] = str(platform_row.get('Pritsche', ''))
    result.at[idx, 'Fuhre_Nr'] = platform_row.get('Fuhre_Nr', result.at[idx, 'Fuhre_Nr'] if 'Fuhre_Nr' in result.columns else 1)
    result.at[idx, 'Fuhrenoption'] = platform_row.get('Fuhrenoption', result.at[idx, 'Fuhrenoption'] if 'Fuhrenoption' in result.columns else '')
    result.at[idx, 'Pritschenname'] = platform_row.get('Pritschenname', result.at[idx, 'Pritschenname'] if 'Pritschenname' in result.columns else '')
    result.at[idx, 'Manuelle_Zone'] = _manual_zone_from_label(zone)
    result.at[idx, 'Manuelle_Position'] = _manual_zone_position(zone)
    result.at[idx, 'Einlage_unten_mm'] = safe_number(spacer_under, 0.0)
    result.at[idx, 'Warnung'] = str(result.at[idx, 'Warnung']) if 'Warnung' in result.columns else ''
    # Turbo-Manuell: beim Klick nur Zuordnung speichern.
    # X/Y/Z werden erst mit "Ladeplan berechnen / prüfen" fachlich berechnet.
    for col in ['X_mm', 'Y_mm', 'Z_mm']:
        if col not in result.columns:
            result[col] = None
        result.at[idx, col] = None
    if 'Ebene' not in result.columns:
        result['Ebene'] = ''
    result.at[idx, 'Ebene'] = 'Zone gesetzt, noch nicht berechnet'

    return clean_placements_dataframe(result)


def _manual_unload_unit(manual_df: pd.DataFrame, platforms_df: pd.DataFrame, unit_id: str) -> pd.DataFrame:
    if manual_df is None or manual_df.empty:
        return manual_df
    result = manual_df.copy()
    ids = result.get('Einheit_ID', pd.Series(dtype=str)).astype(str)
    matches = result.index[ids.eq(str(unit_id))].tolist()
    if not matches:
        return clean_placements_dataframe(result)
    idx = matches[0]
    for col, val in {
        'Pritsche': 'NICHT VERLADEN',
        'Fuhre_Nr': None,
        'Fuhrenoption': '',
        'Pritschenname': '',
        'X_mm': None,
        'Y_mm': None,
        'Z_mm': None,
        'Ebene': 'manuell nicht gesetzt',
        'Manuelle_Zone': '',
        'Manuelle_Position': '',
    }.items():
        if col not in result.columns:
            result[col] = ''
        result.at[idx, col] = val
    return clean_placements_dataframe(result)


def _manual_zone_grid(manual_df: pd.DataFrame, platform_name: str) -> pd.DataFrame:
    """Erzeugt eine einfache BSD-ähnliche Zonentabelle."""
    cols = ['Vorne links', 'Vorne rechts', 'Hinten links', 'Hinten rechts']
    if manual_df is None or manual_df.empty:
        return pd.DataFrame(columns=['Lage'] + cols)

    p_mask = manual_df.get('Pritsche', pd.Series(dtype=str)).astype(str).eq(str(platform_name))
    rows = manual_df[p_mask].copy()
    if rows.empty or 'Manuelle_Zone' not in rows.columns:
        return pd.DataFrame(columns=['Lage'] + cols)

    def _label(row: pd.Series) -> str:
        b = str(row.get('Bauteile') or row.get('Einheit_ID') or '')
        return b if len(b) <= 28 else b[:25] + '...'

    stacks = {}
    for zone in _MANUAL_ZONE_ORDER:
        zr = rows[rows['Manuelle_Zone'].astype(str).str.upper().eq(zone)].copy()
        if zr.empty:
            stacks[zone] = []
        else:
            zr['Z_sort'] = pd.to_numeric(zr.get('Z_mm'), errors='coerce').fillna(0.0)
            zr = zr.sort_values(['Z_sort', 'Manuelle_Reihenfolge', 'Einheit_ID'], kind='stable')
            stacks[zone] = [_label(r) for _, r in zr.iterrows()]

    max_lage = max([len(v) for v in stacks.values()] + [0])
    if max_lage <= 0:
        return pd.DataFrame(columns=['Lage'] + cols)

    out_rows = []
    for lage in range(max_lage, 0, -1):
        out_rows.append({
            'Lage': lage,
            'Vorne links': stacks['VL'][lage - 1] if len(stacks['VL']) >= lage else '',
            'Vorne rechts': stacks['VR'][lage - 1] if len(stacks['VR']) >= lage else '',
            'Hinten links': stacks['HL'][lage - 1] if len(stacks['HL']) >= lage else '',
            'Hinten rechts': stacks['HR'][lage - 1] if len(stacks['HR']) >= lage else '',
        })
    return pd.DataFrame(out_rows)



def _manual_get_platform_status(platforms_df: pd.DataFrame, platform_name: str) -> str:
    if platforms_df is None or platforms_df.empty or 'Pritsche' not in platforms_df.columns:
        return 'Leer'
    if 'Manueller_Status' not in platforms_df.columns:
        return 'Leer'
    rows = platforms_df[platforms_df['Pritsche'].astype(str).eq(str(platform_name))]
    if rows.empty:
        return 'Leer'
    status = str(rows.iloc[0].get('Manueller_Status', '') or '').strip()
    return status or 'Leer'


def _manual_set_platform_status(platforms_df: pd.DataFrame, platform_name: str, status: str) -> pd.DataFrame:
    result = platforms_df.copy() if platforms_df is not None else pd.DataFrame()
    if result.empty or 'Pritsche' not in result.columns:
        return result
    if 'Manueller_Status' not in result.columns:
        result['Manueller_Status'] = 'Leer'
    mask = result['Pritsche'].astype(str).eq(str(platform_name))
    result.loc[mask, 'Manueller_Status'] = str(status)
    return result


def _manual_update_platform_statuses(platforms_df: pd.DataFrame, manual_df: pd.DataFrame) -> pd.DataFrame:
    """Leer/In Bearbeitung ergänzen, abgeschlossene Pritschen bleiben abgeschlossen."""
    result = platforms_df.copy() if platforms_df is not None else pd.DataFrame()
    if result.empty or 'Pritsche' not in result.columns:
        return result
    if 'Manueller_Status' not in result.columns:
        result['Manueller_Status'] = 'Leer'
    loaded_names = set()
    if manual_df is not None and not manual_df.empty and 'Pritsche' in manual_df.columns:
        loaded_names = set(manual_df[manual_df['Pritsche'].astype(str) != 'NICHT VERLADEN']['Pritsche'].astype(str).tolist())
    for idx, row in result.iterrows():
        status = str(row.get('Manueller_Status', '') or '').strip()
        if status in ['Abgeschlossen', 'Berechnet / geprüft']:
            continue
        result.at[idx, 'Manueller_Status'] = 'In Bearbeitung' if str(row.get('Pritsche')) in loaded_names else 'Leer'
    return result


def _manual_start_next_fuhre(platforms_df: pd.DataFrame) -> pd.DataFrame:
    """Erzeugt aus der vorhandenen Fuhrenoption die nächste Fuhre, z. B. F02 LKW/F02 Anhänger."""
    if platforms_df is None or platforms_df.empty:
        return platforms_df
    result = platforms_df.copy()
    if 'Fuhre_Nr' not in result.columns:
        result['Fuhre_Nr'] = 1
    result['Fuhre_Nr'] = pd.to_numeric(result['Fuhre_Nr'], errors='coerce').fillna(1).astype(int)
    max_f = int(result['Fuhre_Nr'].max()) if not result.empty else 1
    next_f = max_f + 1

    template = result[result['Fuhre_Nr'] == 1].copy()
    if template.empty:
        template = result.copy()
    new_rows = template.copy()
    new_rows['Fuhre_Nr'] = next_f
    if 'Pritschenname' not in new_rows.columns:
        new_rows['Pritschenname'] = new_rows['Pritsche'].astype(str).str.replace(r'^F\\d+\\s+', '', regex=True)
    new_rows['Pritschenname'] = new_rows['Pritschenname'].astype(str)
    new_rows['Pritsche'] = new_rows.apply(lambda r: f"F{next_f:02d} {r['Pritschenname']}", axis=1)
    new_rows['Manueller_Status'] = 'Leer'
    combined = pd.concat([result, new_rows], ignore_index=True, sort=False)
    return _clean_manual_platforms_dataframe(combined)


def _manual_validate_zone_plan(manual_df: pd.DataFrame, platforms_df: pd.DataFrame) -> pd.DataFrame:
    """Prüft vor dem fachlichen Berechnen, ob die manuelle Zonenzuordnung grundsätzlich verladen werden darf."""
    issues: List[Dict[str, Any]] = []
    if manual_df is None or manual_df.empty or platforms_df is None or platforms_df.empty:
        return pd.DataFrame(issues)

    for _, prow in platforms_df.iterrows():
        pname = str(prow.get('Pritsche', ''))
        if not pname:
            continue
        p_rows = manual_df[manual_df.get('Pritsche', pd.Series(dtype=str)).astype(str).eq(pname)].copy()
        if p_rows.empty:
            continue

        eff_len = _manual_effective_length(prow)
        p_width = safe_number(prow.get('Breite_mm'), 0.0)
        max_h = safe_number(prow.get('Max_Höhe_mm'), 0.0)
        base_z = safe_number(prow.get('Kantholz_erste_Lage_mm'), 0.0)
        max_weight = safe_number(prow.get('Max_Gewicht_kg'), 0.0)
        own_weight = safe_number(prow.get('Eigengewicht_Pritsche_kg'), 0.0)

        zone_rows = p_rows[p_rows.get('Manuelle_Zone', pd.Series(dtype=str)).astype(str).str.upper().isin(_MANUAL_ZONE_ORDER)].copy()
        if zone_rows.empty:
            continue

        total_weight = safe_number(zone_rows.get('Gewicht_kg', pd.Series(dtype=float)).sum(), 0.0)
        if max_weight > 0 and total_weight + own_weight > max_weight:
            issues.append({
                'Pritsche': pname,
                'Bereich': 'Gewicht',
                'Problem': f'Ladegewicht {total_weight:.0f} kg + Eigengewicht {own_weight:.0f} kg > max. {max_weight:.0f} kg',
            })

        for zone in _MANUAL_ZONE_ORDER:
            zr = zone_rows[zone_rows['Manuelle_Zone'].astype(str).str.upper().eq(zone)].copy()
            if zr.empty:
                continue
            max_len_zone = pd.to_numeric(zr.get('Länge_mm'), errors='coerce').fillna(0.0).max()
            max_width_zone = pd.to_numeric(zr.get('Breite_mm'), errors='coerce').fillna(0.0).max()
            stack_height = base_z
            for _, r in zr.iterrows():
                stack_height += safe_number(r.get('Einlage_unten_mm'), safe_number(prow.get('Einlage_zwischen_Lagen_mm'), 0.0))
                stack_height += safe_number(r.get('Höhe_mm'), 0.0)

            if eff_len > 0 and max_len_zone > eff_len:
                issues.append({
                    'Pritsche': pname,
                    'Bereich': _manual_zone_label(zone),
                    'Problem': f'Länge {max_len_zone:.0f} mm > Lade-Länge {eff_len:.0f} mm',
                })
            if p_width > 0 and max_width_zone > p_width:
                issues.append({
                    'Pritsche': pname,
                    'Bereich': _manual_zone_label(zone),
                    'Problem': f'Breite {max_width_zone:.0f} mm > Pritschenbreite {p_width:.0f} mm',
                })
            if max_h > 0 and stack_height > max_h:
                issues.append({
                    'Pritsche': pname,
                    'Bereich': _manual_zone_label(zone),
                    'Problem': f'Stapelhöhe {stack_height:.0f} mm > max. Ladehöhe {max_h:.0f} mm',
                })

        # Center-out X-Regel: vorne und hinten auf derselben Seite wachsen ab Mitte.
        # Darum darf die Summe aus maximaler Front-Länge + maximaler Heck-Länge die Lade-Länge nicht überschreiten.
        for side_name, front_zone, back_zone in [('links', 'VL', 'HL'), ('rechts', 'VR', 'HR')]:
            front = zone_rows[zone_rows['Manuelle_Zone'].astype(str).str.upper().eq(front_zone)].copy()
            back = zone_rows[zone_rows['Manuelle_Zone'].astype(str).str.upper().eq(back_zone)].copy()
            if front.empty or back.empty:
                continue
            front_len = pd.to_numeric(front.get('Länge_mm'), errors='coerce').fillna(0.0).max()
            back_len = pd.to_numeric(back.get('Länge_mm'), errors='coerce').fillna(0.0).max()
            if eff_len > 0 and front_len + back_len > eff_len:
                issues.append({
                    'Pritsche': pname,
                    'Bereich': f'{side_name} vorne+hinten',
                    'Problem': f'Längen zusammen {front_len + back_len:.0f} mm > Lade-Länge inkl. Überstand {eff_len:.0f} mm',
                })

    return pd.DataFrame(issues)



def _manual_small_platform_overview(platforms_df: pd.DataFrame, manual_df: pd.DataFrame) -> pd.DataFrame:
    """Kleine Übersicht statt grosser Matrix: Status je Pritsche."""
    if platforms_df is None or platforms_df.empty:
        return pd.DataFrame()
    rows: List[Dict[str, Any]] = []
    for _, prow in platforms_df.iterrows():
        pname = str(prow.get('Pritsche', ''))
        p_rows = pd.DataFrame()
        if manual_df is not None and not manual_df.empty and 'Pritsche' in manual_df.columns:
            p_rows = manual_df[manual_df['Pritsche'].astype(str).eq(pname)].copy()
        count_units = int(len(p_rows)) if p_rows is not None else 0

        used_height = 0.0
        used_length = 0.0
        used_weight = 0.0
        if p_rows is not None and not p_rows.empty:
            z_vals = pd.to_numeric(p_rows.get('Z_mm'), errors='coerce')
            h_vals = pd.to_numeric(p_rows.get('Höhe_mm'), errors='coerce')
            if not (z_vals + h_vals).dropna().empty:
                used_height = float((z_vals + h_vals).dropna().max())
            x_vals = pd.to_numeric(p_rows.get('X_mm'), errors='coerce')
            l_vals = pd.to_numeric(p_rows.get('Länge_mm'), errors='coerce')
            if not x_vals.dropna().empty and not (x_vals + l_vals).dropna().empty:
                used_length = float((x_vals + l_vals).dropna().max() - x_vals.dropna().min())
            used_weight = float(pd.to_numeric(p_rows.get('Gewicht_kg'), errors='coerce').fillna(0.0).sum())

        rows.append({
            'Pritsche': pname,
            'Status': _manual_get_platform_status(platforms_df, pname),
            'Einheiten': count_units,
            'Länge': f"{used_length:.0f} / {_manual_effective_length(prow):.0f} mm",
            'Höhe': f"{used_height:.0f} / {safe_number(prow.get('Max_Höhe_mm')):.0f} mm",
            'Gewicht': f"{used_weight:.0f} / {safe_number(prow.get('Max_Gewicht_kg')):.0f} kg" if safe_number(prow.get('Max_Gewicht_kg')) > 0 else f"{used_weight:.0f} kg",
        })
    return pd.DataFrame(rows)


def _manual_platform_row(platforms_used_df: pd.DataFrame, platform_name: str) -> Optional[pd.Series]:
    if platforms_used_df is None or platforms_used_df.empty or 'Pritsche' not in platforms_used_df.columns:
        return None
    rows = platforms_used_df[platforms_used_df['Pritsche'].astype(str) == str(platform_name)]
    if rows.empty:
        return None
    return rows.iloc[0]


def _manual_effective_length(platform_row: Optional[pd.Series]) -> float:
    if platform_row is None:
        return 0.0
    return (
        safe_number(platform_row.get('Länge_mm'))
        + safe_number(platform_row.get('Überhang_vorne_mm'))
        + safe_number(platform_row.get('Überhang_hinten_mm'))
    )


def _manual_update_row(df: pd.DataFrame, row_idx: int, **updates) -> pd.DataFrame:
    result = df.copy()
    for col, val in updates.items():
        if col not in result.columns:
            result[col] = ''
        result.at[row_idx, col] = val
    if 'Ebene' in result.columns:
        old = str(result.at[row_idx, 'Ebene']) if row_idx in result.index else ''
        if 'manuell' not in old.lower():
            result.at[row_idx, 'Ebene'] = f'{old} / manuell geändert' if old else 'manuell geändert'
    return clean_placements_dataframe(result)


def _manual_split_bundle_row(df: pd.DataFrame, row_idx: int) -> pd.DataFrame:
    """Teilt einen Bund in einzelne manuell verschiebbare Bauteil-Zeilen."""
    if df is None or df.empty or row_idx not in df.index:
        return df
    row = df.loc[row_idx].copy()
    labels = [x.strip() for x in str(row.get('Bauteile_Liste') or row.get('Bauteile') or '').replace(',', '|').split('|') if x.strip()]
    if len(labels) <= 1:
        return df

    def _split_nums(value: Any, default: float, n: int) -> List[float]:
        raw = [x.strip() for x in str(value or '').split('|') if x.strip()]
        vals: List[float] = []
        for x in raw:
            try:
                vals.append(float(x))
            except Exception:
                vals.append(float(default))
        if not vals:
            vals = [float(default)]
        while len(vals) < n:
            vals.append(vals[-1])
        return vals[:n]

    n = len(labels)
    lengths = _split_nums(row.get('Einzellängen_mm'), safe_number(row.get('Länge_mm')), n)
    widths = _split_nums(row.get('Einzelbreiten_mm'), safe_number(row.get('Breite_mm')), n)
    heights = _split_nums(row.get('Einzelhöhen_mm'), safe_number(row.get('Höhe_mm')) / max(1, n), n)
    total_weight = safe_number(row.get('Gewicht_kg'), 0.0)
    total_volume = safe_number(row.get('Volumen_m3'), 0.0) if 'Volumen_m3' in df.columns else 0.0
    general_spacer = safe_number(row.get('Einlage_allgemein_mm'), 0.0)
    z_cursor = safe_number(row.get('Z_mm'), 0.0)

    new_rows: List[pd.Series] = []
    base_id = str(row.get('Einheit_ID') or 'MAN')
    for i, label in enumerate(labels):
        nr = row.copy()
        nr['Einheit_ID'] = f'{base_id}.{i+1}'
        nr['Typ'] = 'Bauteil'
        nr['Anzahl_Bauteile'] = 1
        nr['Bauteile'] = label
        nr['Bauteile_Liste'] = label
        nr['Ansicht_Label'] = label
        nr['Ansicht_Liste'] = label
        nr['Einzellängen_mm'] = str(lengths[i])
        nr['Einzelbreiten_mm'] = str(widths[i])
        nr['Einzelhöhen_mm'] = str(heights[i])
        nr['Länge_mm'] = lengths[i]
        nr['Breite_mm'] = widths[i]
        nr['Höhe_mm'] = heights[i]
        nr['Z_mm'] = round(z_cursor, 1)
        nr['Gewicht_kg'] = round(total_weight / n, 2)
        if 'Volumen_m3' in df.columns:
            nr['Volumen_m3'] = round(total_volume / n, 4) if total_volume else 0.0
        if 'Warnung' in df.columns:
            nr['Warnung'] = 'manuell aus Bund geteilt'
        if 'Ebene' in df.columns:
            nr['Ebene'] = 'manuell aus Bund geteilt'
        new_rows.append(nr)
        z_cursor += heights[i] + general_spacer

    before = df.loc[df.index < row_idx].copy()
    after = df.loc[df.index > row_idx].copy()
    result = pd.concat([before, pd.DataFrame(new_rows), after], ignore_index=True, sort=False)
    return clean_placements_dataframe(result)

def render_loading_module(uploaded_file, transport_excel_file=None, logo_file=None, parts_excel_file=None) -> None:
    st.header('Verladeplanung')

    if uploaded_file is None and parts_excel_file is None:
        st.info('Bitte laden Sie links eine BVX-Datei oder eine Bauteile-Excel für die Verladeplanung hoch.')
        st.markdown('''
        Die Verladeplanung ist getrennt von der normalen BVX-Auswertung aufgebaut.
        Sie arbeitet mit Bauteilen, Verladeeinheiten, Bunden, Fuhrenoptionen, Pritschen und Positionen.

        Falls keine BVX vorhanden ist, kann eine Excel-Datei mit Blatt **Bauteile** geladen werden.
        ''')
        return

    source_name = uploaded_file.name if uploaded_file is not None else parts_excel_file.name
    parsed_result = None
    if uploaded_file is not None:
        content = read_uploaded_text(uploaded_file)
        parser = BVXParser()
        parsed_result = parser.parse(content, uploaded_file.name)

    options_df, pritschen_df, standards, config_messages = read_transport_config_excel(transport_excel_file)
    for msg in config_messages:
        st.warning(msg)

    if transport_excel_file is not None:
        st.success(f'Pritschen-/Fuhren-Stammdaten geladen: {transport_excel_file.name}')
    else:
        st.info('Noch keine Excel-Stammdaten geladen. Es werden nur Beispielwerte verwendet.')

    st.subheader('1. Grunddaten aus BVX / Bauteile-Excel und Pritschen-Excel')
    default_density = safe_number(standards.get('Holzdichte'), 500.0)
    default_bundle_weight = safe_number(standards.get('Max_Bundgewicht'), 1000.0)
    default_base_wood = safe_number(standards.get('Standard_Kantholz_erste_Lage'), 80.0)
    default_layer_spacer = safe_number(standards.get('Standard_Einlage_zwischen_Lagen'), 40.0)
    default_general_spacer = safe_number(standards.get('Standard_Einlage_allgemein'), 0.0)
    default_gap = safe_number(standards.get('Längenversatz_je_Lage'), 100.0)

    col1, col2 = st.columns([1, 2])
    density = col1.number_input('Holzdichte kg/m³', min_value=100.0, max_value=1000.0, value=float(default_density), step=10.0)
    col2.caption('Bundbildung, Bundgewicht und Einlagen sind gesammelt im Abschnitt 4.')

    if parsed_result is not None:
        parts_df = parts_to_dataframe(parsed_result.parts, density_kg_m3=density)
    else:
        parts_df, parts_messages = read_parts_excel_to_dataframe(parts_excel_file, density_kg_m3=density)
        for msg in parts_messages:
            st.info(msg)

    if parts_df.empty:
        st.error('Keine Bauteile vorhanden. Bitte BVX oder Bauteile-Excel prüfen.')
        return

    col1, col2, col3, col4 = st.columns(4)
    col1.metric('Bauteile', len(parts_df))
    col2.metric('Gesamtvolumen', f"{parts_df['Volumen_m3'].sum():.3f} m³")
    col3.metric('Gesamtgewicht', f"{parts_df['Gewicht_kg'].sum():.0f} kg")
    col4.metric('Excel-Fuhrenoptionen', int(options_df['Freigegeben'].sum()) if not options_df.empty else 0)

    st.subheader('2. Projektdaten für Ladeplan-Kopf')
    st.caption('Objekt, Transport, Sachbearbeiter und Datum werden manuell eingegeben. Decke und Bauabschnitt können aus BVX-Attributen übernommen oder manuell überschrieben werden.')
    meta_fields = ['Manuell'] + _available_bvx_meta_fields(parts_df)
    mcol1, mcol2, mcol3, mcol4 = st.columns(4)
    objekt_name = mcol1.text_input('Objekt Name', value='')
    transport_name = mcol2.text_input('Transport Name / Unternehmer', value='')
    sachbearbeiter = mcol3.text_input('Sachbearbeiter', value='')
    ladeplan_datum = mcol4.text_input('Datum', value=datetime.now().strftime('%d.%m.%Y'))

    mcol5, mcol6, mcol7, mcol8 = st.columns(4)
    decke_attr = mcol5.selectbox('Decke aus BVX-Attribut', meta_fields, index=_default_attr_index(meta_fields, ['User_Attribut_2', 'BVX_User_Attribut_2', 'Name']))
    decke_auto = _unique_display_values_from_field(parts_df, decke_attr)
    decke_text = mcol6.text_input('Decke', value=decke_auto)
    bauabschnitt_attr = mcol7.selectbox('Bauabschnitt aus BVX-Attribut', meta_fields, index=_default_attr_index(meta_fields, ['Pak/Unit', 'Unit', 'BVX_Unit', 'BVX_User_Attribut_2']))
    bauabschnitt_auto = _unique_display_values_from_field(parts_df, bauabschnitt_attr)
    bauabschnitt_text = mcol8.text_input('Bauabschnitt', value=bauabschnitt_auto)

    st.caption('Orientierung fix: Vorne = X=max, Hinten = X=0, Links = Y=0, Rechts = Y=max. Diese Zuordnung ist nicht mehr verstellbar.')

    vorne_orientation = 'Vorne = X=max (rechts im Plan)'
    links_orientation = 'Links = Y=0'

    project_meta = {
        'Objekt_Name': objekt_name,
        'Transport_Name': transport_name,
        'Sachbearbeiter': sachbearbeiter,
        'Datum': ladeplan_datum,
        'Decke': decke_text,
        'Bauabschnitt': bauabschnitt_text,
        'Decke_Attribut': decke_attr,
        'Bauabschnitt_Attribut': bauabschnitt_attr,
        'Vorne_Orientierung': vorne_orientation,
        'Links_Orientierung': links_orientation,
    }

    logo_bytes = None
    if logo_file is not None:
        try:
            logo_bytes = logo_file.getvalue()
        except Exception:
            logo_bytes = None
    if logo_bytes is None:
        logo_bytes = get_embedded_default_logo()

    st.subheader('3. Sortierung')
    sort_options = [
        'Bauteilnummer', 'Pak/Unit', 'Name', 'PartId', 'Profil', 'Oberfläche', 'Qualität',
        'Länge_mm', 'Breite_mm', 'Höhe_mm', 'Gewicht_kg', 'Volumen_m3'
    ]
    col1, col2, col3, col4 = st.columns(4)
    main_attr = col1.selectbox('Hauptattribut', sort_options, index=0)
    main_direction = col2.selectbox('Richtung Hauptattribut', ['aufsteigend', 'absteigend'], index=0)
    second_attr = col3.selectbox('Nebenattribut', ['Keine'] + sort_options, index=0)
    second_direction = col4.selectbox('Richtung Nebenattribut', ['aufsteigend', 'absteigend'], index=0)

    sorted_parts = sort_parts_dataframe(
        parts_df,
        main_attr=main_attr,
        main_asc=(main_direction == 'aufsteigend'),
        second_attr=second_attr,
        second_asc=(second_direction == 'aufsteigend'),
    )

    label_options = _unique_options(
        ['Bauteilnummer', 'Name', 'Pak/Unit', 'Bauteilnummer + Pak/Unit', 'Bauteilnummer + Name', 'Name + Pak/Unit']
        + sort_options
        + _available_bvx_meta_fields(parts_df)
    )
    default_label_index = label_options.index('Bauteilnummer') if 'Bauteilnummer' in label_options else 0
    display_label_attr = st.selectbox(
        'Beschriftung in Verladeplänen / BSD',
        label_options,
        index=default_label_index,
        key='display_label_attr_v55',
        help='Diese Auswahl bestimmt nur die Beschriftung in Ansicht, PDF und BSD. Die Sortierung bleibt separat über Hauptattribut/Nebenattribut.'
    )

    project_meta['Beschriftungsattribut'] = display_label_attr

    with st.expander('Bauteile anzeigen', expanded=False):
        st.dataframe(sorted_parts, use_container_width=True, hide_index=True)

    st.subheader('4. Bundbildung und Unterlegholz')
    st.caption('Alle Einstellungen zur Bundbildung und zu Unterleghölzern sind hier gesammelt.')

    with st.expander('4a. Bundbildung', expanded=True):
        bcol1, bcol2 = st.columns(2)
        use_bundles = bcol1.checkbox('Bunde automatisch bilden', value=True)
        max_bundle_weight = bcol2.number_input('Max. Bundgewicht kg', min_value=100.0, max_value=5000.0, value=float(default_bundle_weight), step=50.0)

        bcol3, bcol4, bcol5, bcol6 = st.columns(4)
        same_height = bcol3.checkbox('Nur gleiche Höhe im Bund', value=True)
        same_width = bcol4.checkbox('Nur gleiche Breite im Bund', value=False)
        same_quality = bcol5.checkbox('Nur gleiche Qualität im Bund', value=False)
        same_profile = bcol6.checkbox('Nur gleiches Profil im Bund', value=False)

    with st.expander('4b. Unterlegholz / Einlagen / Längenversatz', expanded=True):
        st.caption('Kantholz = direkt auf der Pritsche. Bundeinlage/Lagenholz = zwischen Bunden/Lagen. Einlage allgemein = zwischen einzelnen Bauteilen.')
        ucol1, ucol2, ucol3, ucol4 = st.columns(4)
        base_wood_height = ucol1.number_input('Standard Kantholz erste Lage mm', min_value=0.0, max_value=300.0, value=float(default_base_wood), step=5.0)
        bundle_spacer_height = ucol2.number_input('Standard Bundeinlage / Lagenholz mm', min_value=0.0, max_value=200.0, value=float(default_layer_spacer), step=5.0)
        general_spacer_height = ucol3.number_input('Einlage allgemein zwischen jedem Bauteil mm', min_value=0.0, max_value=200.0, value=float(default_general_spacer), step=5.0, help='0 = aus. Wenn grösser 0, wird die Einlage zwischen gestapelten Einzelbauteilen und innerhalb eines Bundes eingezeichnet und in der Höhe berücksichtigt.')
        gap_length = ucol4.number_input('Längenversatz je Lage mm', min_value=0.0, max_value=500.0, value=float(default_gap), step=10.0)

    # Standards werden für die Berechnung aktualisiert. Pritschenwerte aus Excel überschreiben diese Defaults pro Pritsche.
    standards['Holzdichte'] = density
    standards['Max_Bundgewicht'] = max_bundle_weight
    standards['Standard_Kantholz_erste_Lage'] = base_wood_height
    standards['Standard_Einlage_zwischen_Lagen'] = bundle_spacer_height
    standards['Standard_Einlage_allgemein'] = general_spacer_height
    standards['Längenversatz_je_Lage'] = gap_length

    # Vorschau: globale Bundbildung nur zur ersten Orientierung.
    # Für die echte Verladung ab V21 werden Bunde erst nach der Pritschenblockwahl gebildet.
    units_preview_df = build_loading_units(
        sorted_parts,
        use_bundles=use_bundles,
        max_bundle_weight=max_bundle_weight,
        bundle_spacer_height=bundle_spacer_height,
        general_spacer_height=general_spacer_height,
        same_height=same_height,
        same_width=same_width,
        same_quality=same_quality,
        same_profile=same_profile,
        label_attr=display_label_attr,
    )
    units_df = units_preview_df.copy()

    st.subheader('5. Fuhrenoptionen und Pritschen aus Excel')
    st.caption('V22: Sortierung und Stapel-/Anzeige-Richtung sind getrennt. Die erste Nummer der sortierten Bundliste steht im Bund oben; physisch wird weiter von unten nach oben geladen.')

    fcol1, fcol2 = st.columns([1, 2])
    with fcol1:
        options_edit = st.data_editor(
            options_df,
            use_container_width=True,
            hide_index=True,
            column_config={
                'Freigegeben': st.column_config.CheckboxColumn('Freigegeben'),
                'Wiederholen_bis_alles_verladen': st.column_config.CheckboxColumn('Wiederholen'),
                'Priorität': st.column_config.NumberColumn('Priorität'),
            },
            key='fuhrenoptionen_editor',
        )
    with fcol2:
        pritschen_edit = st.data_editor(
            pritschen_df,
            use_container_width=True,
            hide_index=True,
            column_config={
                'Freigabe': st.column_config.CheckboxColumn('Aktiv'),
                'Drehen_90_erlaubt': st.column_config.CheckboxColumn('Drehen 90°'),
                'Pritschen_Reihenfolge': st.column_config.NumberColumn('Reihenfolge'),
                'Länge_mm': st.column_config.NumberColumn('Länge mm'),
                'Breite_mm': st.column_config.NumberColumn('Breite mm'),
                'Max_Höhe_mm': st.column_config.NumberColumn('Max. Höhe mm'),
                'Max_Gewicht_kg': st.column_config.NumberColumn('Max. Gesamtgewicht kg'),
                'Eigengewicht_Pritsche_kg': st.column_config.NumberColumn('Eigengewicht Pritsche kg'),
                'Kantholz_erste_Lage_mm': st.column_config.NumberColumn('Kantholz erste Lage mm'),
                'Einlage_zwischen_Lagen_mm': st.column_config.NumberColumn('Bundeinlage / Lagenholz mm'),
                'Einlage_allgemein_mm': st.column_config.NumberColumn('Einlage allgemein mm'),
            },
            key='pritschen_editor_excel',
        )

    # Aktuelle Unterlegholz-Einstellungen global auf alle Pritschen anwenden.
    # Damit überschreiben die Eingabefelder veraltete Werte aus der Excel-Stammdatendatei.
    if not pritschen_edit.empty:
        pritschen_edit = pritschen_edit.copy()
        pritschen_edit['Kantholz_erste_Lage_mm'] = float(base_wood_height)
        pritschen_edit['Einlage_zwischen_Lagen_mm'] = float(bundle_spacer_height)
        pritschen_edit['Einlage_allgemein_mm'] = float(general_spacer_height)

    st.subheader('6. Verladeart')
    verladeart = st.radio('Verladeart wählen', ['Automatisch', 'Manuell'], horizontal=True, key='verladeart_v55')
    st.caption('Automatisch = App erstellt Vorschlag. Manuell = du setzt Bauteile/Bunde nur in Zonen; die App berechnet die genaue Lage.')

    manual_fuhrenoption = ''
    if verladeart == 'Manuell':
        option_choices_manual = options_edit['Fuhrenoption'].dropna().astype(str).drop_duplicates().tolist() if options_edit is not None and not options_edit.empty and 'Fuhrenoption' in options_edit.columns else []
        if not option_choices_manual:
            option_choices_manual = pritschen_edit['Fuhrenoption'].dropna().astype(str).drop_duplicates().tolist() if pritschen_edit is not None and not pritschen_edit.empty and 'Fuhrenoption' in pritschen_edit.columns else []
        manual_fuhrenoption = st.selectbox('Fuhrenoption für manuelle Verladung', option_choices_manual, key='manual_fuhrenoption_v55') if option_choices_manual else ''

    st.subheader('6a. Platzierung / Automatik')
    st.caption('Bei Automatik: ruhige Ausrichtung ist fest aktiv. Bei Manuell: du bestimmst nur die Zone, die Geometrie wird automatisch gesetzt.')
    col1, col2, col3, col4 = st.columns(4)
    allow_beside = col1.checkbox('Nebeneinander erlauben', value=True)
    allow_stack = col2.checkbox('Übereinander erlauben', value=True)
    allow_rotation = col3.checkbox('90° drehen erlauben, wenn Pritsche es erlaubt', value=False)
    center_geometric = True
    max_fuhren = col4.number_input('Max. Fuhren Sicherheitslimit', min_value=1, max_value=200, value=50, step=1)

    st.subheader('7. Auflage / Unterbau')
    st.caption('Ruhige Logik: Unterbau ist nur Kontrolle/Notlösung. Zuerst werden weniger Fuhren und saubere Lagen angestrebt; Unterbau sollte möglichst selten verwendet werden.')
    ucol1, ucol2, ucol3 = st.columns(3)
    underbau_enabled = ucol1.checkbox('Unterbau / Auflage prüfen und anzeigen', value=False)
    min_support_ratio = ucol2.number_input('Mindestauflagefläche %', min_value=0, max_value=100, value=65, step=5) / 100.0
    min_underbau_height = ucol3.number_input('Unterbau anzeigen ab mm', min_value=0.0, max_value=500.0, value=20.0, step=5.0)

    if verladeart == 'Automatisch':
        placements_df, summary_df, platforms_used_df, fuhren_log_df, plan_units_df = create_variant_a_loading_plan(
            sorted_parts,
            options_edit,
            pritschen_edit,
            standards=standards,
            allow_beside=allow_beside,
            allow_stack=allow_stack,
            allow_rotation=allow_rotation,
            center_geometric=center_geometric,
            max_fuhren=int(max_fuhren),
            use_bundles=use_bundles,
            max_bundle_weight=max_bundle_weight,
            bundle_spacer_height=bundle_spacer_height,
            general_spacer_height=general_spacer_height,
            same_height=same_height,
            same_width=same_width,
            same_quality=same_quality,
            same_profile=same_profile,
            label_attr=display_label_attr,
        )
        # Ab hier arbeitet die App mit den tatsächlich je Pritschenblock gebildeten Verladeeinheiten.
        if plan_units_df is not None and not plan_units_df.empty:
            units_df = plan_units_df.copy()
    else:
        plan_units_df = units_df.copy()
        platforms_used_df = _build_manual_platforms_from_excel(pritschen_edit, manual_fuhrenoption, standards)
        placements_df = _build_manual_empty_placements_from_units(units_df)
        summary_df = recompute_summary_from_placements(placements_df, platforms_used_df) if not platforms_used_df.empty else pd.DataFrame()
        fuhren_log_df = pd.DataFrame([{
            'Fuhre_Nr': 1,
            'Fuhrenoption': manual_fuhrenoption,
            'Option_Rang': 1,
            'Pritschen': ', '.join(platforms_used_df['Pritsche'].astype(str).tolist()) if not platforms_used_df.empty else '',
            'Bauteile von': '',
            'Bauteile bis': '',
            'Bauteile Anzahl': 0,
            'Status': 'Manuell',
        }])

    # Manueller Planstand:
    # Die Automatik erzeugt den Vorschlag. Danach arbeiten Tabelle, Ansichten und Export
    # mit dem manuellen Planstand aus st.session_state, bis bewusst zurückgesetzt wird.
    manual_signature = f"{verladeart}|{manual_fuhrenoption}|" + _manual_plan_signature(placements_df)
    platform_signature = f"{verladeart}|{manual_fuhrenoption}|" + _manual_platform_signature(platforms_used_df)
    if (
        st.session_state.get('manual_plan_signature') != manual_signature
        or 'manual_placements_df' not in st.session_state
    ):
        st.session_state['manual_plan_signature'] = manual_signature
        manual_init = clean_placements_dataframe(placements_df)
        if 'Einlage_unten_mm' not in manual_init.columns:
            manual_init['Einlage_unten_mm'] = 0.0
        if 'Einlage_oben_mm' not in manual_init.columns:
            manual_init['Einlage_oben_mm'] = 0.0
        if 'Manuelle_Zone' not in manual_init.columns:
            manual_init['Manuelle_Zone'] = ''
        if 'Manuelle_Reihenfolge' not in manual_init.columns:
            manual_init['Manuelle_Reihenfolge'] = 0.0
        st.session_state['manual_placements_df'] = manual_init

    if (
        st.session_state.get('manual_platform_signature') != platform_signature
        or 'manual_platforms_df' not in st.session_state
    ):
        st.session_state['manual_platform_signature'] = platform_signature
        st.session_state['manual_platforms_df'] = _clean_manual_platforms_dataframe(platforms_used_df)

    platforms_used_df = _clean_manual_platforms_dataframe(st.session_state['manual_platforms_df'])
    edited_placements_df = clean_placements_dataframe(st.session_state['manual_placements_df'])
    loaded_count = int((edited_placements_df['Pritsche'] != 'NICHT VERLADEN').sum()) if not edited_placements_df.empty and 'Pritsche' in edited_placements_df.columns else 0
    not_loaded_count = int((edited_placements_df['Pritsche'] == 'NICHT VERLADEN').sum()) if not edited_placements_df.empty and 'Pritsche' in edited_placements_df.columns else 0
    fuhren_count = int(fuhren_log_df['Fuhre_Nr'].nunique()) if not fuhren_log_df.empty else 0

    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric('Verladeeinheiten', len(edited_placements_df) if not edited_placements_df.empty else len(units_df))
    col2.metric('Fuhren erzeugt', fuhren_count)
    col3.metric('Verladen', loaded_count)
    col4.metric('Nicht verladen', not_loaded_count)
    col5.metric('Pritschen genutzt', len(platforms_used_df) if not platforms_used_df.empty else 0)

    manual_plan_ready = bool(st.session_state.get('manual_plan_ready_v55', False)) if verladeart == 'Manuell' else True
    if verladeart == 'Manuell' and not manual_plan_ready:
        # Turbo-Manuell: während dem Anklicken nur Zuordnung anzeigen, keine schwere Zwischenberechnung.
        edited_summary_df = recompute_summary_from_placements(edited_placements_df, platforms_used_df) if not platforms_used_df.empty else summary_df
        warnings_plan_df = pd.DataFrame()
        display_placements_df = edited_placements_df.copy()
    else:
        edited_summary_df = recompute_summary_from_placements(edited_placements_df, platforms_used_df) if not platforms_used_df.empty else summary_df
        warnings_plan_df = compute_loading_warnings(edited_placements_df, platforms_used_df)
        display_placements_df, underbau_warnings_df = add_underbau_rows_to_placements(
            edited_placements_df,
            platforms_used_df,
            min_support_ratio=float(min_support_ratio),
            min_underbau_height=float(min_underbau_height),
            enabled=bool(underbau_enabled),
        )
        if underbau_warnings_df is not None and not underbau_warnings_df.empty:
            warnings_plan_df = pd.concat([warnings_plan_df, underbau_warnings_df], ignore_index=True, sort=False) if not warnings_plan_df.empty else underbau_warnings_df

    tab0, tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs(['Kontrolle vor PDF', 'Verladeeinheiten', 'Fuhrenübersicht', 'Manuelle Verladung', 'Ladeplan BSD', 'Warnungen', 'Ansichten', 'Excel Export'])


    with tab0:
        st.subheader('Kontrollansicht vor PDF / BSD')
        st.caption('Diese Tabellen sind bewusst vor dem PDF platziert. Erst wenn Bauteile, Bunde, Pritschen-Zuteilung und Lagenplan logisch stimmen, sollte der PDF-Plan als verbindlich betrachtet werden.')

        bundle_control_df = build_control_bundle_table(units_df)
        assignment_control_df = build_control_assignment_table(edited_placements_df, platforms_used_df)
        layer_control_df = build_control_layer_table(edited_placements_df, platforms_used_df)
        control_issue_df = build_control_issue_table(bundle_control_df, assignment_control_df, layer_control_df)

        st.markdown('**1. Bauteile sortiert**')
        part_cols = [c for c in ['Bauteilnummer', 'Name', 'Pak/Unit', 'Länge_mm', 'Breite_mm', 'Höhe_mm', 'Gewicht_kg', 'Volumen_m3', 'Qualität', 'Profil'] if c in sorted_parts.columns]
        st.dataframe(sorted_parts[part_cols] if part_cols else sorted_parts, use_container_width=True, hide_index=True)

        st.markdown('**2. Bunde / Verladeeinheiten gebildet**')
        st.dataframe(bundle_control_df, use_container_width=True, hide_index=True)

        st.markdown('**3. Pritschen-Zuteilung / Nummernbereiche**')
        if assignment_control_df.empty:
            st.warning('Noch keine Pritschen-Zuteilung vorhanden.')
        else:
            st.dataframe(assignment_control_df, use_container_width=True, hide_index=True)

        st.markdown('**4. Lagenplan je Pritsche**')
        if layer_control_df.empty:
            st.warning('Noch kein Lagenplan vorhanden.')
        else:
            platform_names = layer_control_df['Pritsche'].astype(str).drop_duplicates().tolist()
            selected_control_platform = st.selectbox('Pritsche für Kontroll-Lagenplan', platform_names, key='control_layer_platform')
            st.dataframe(
                layer_control_df[layer_control_df['Pritsche'].astype(str) == selected_control_platform],
                use_container_width=True,
                hide_index=True,
            )

        st.markdown('**5. Prüfpunkte**')
        if control_issue_df.empty:
            st.success('Kontrolltabellen zeigen keine offensichtlichen Rücksprünge oder Lagenrichtungsfehler.')
        else:
            st.warning('Es gibt Punkte, die vor dem PDF geprüft werden sollten.')
            st.dataframe(control_issue_df, use_container_width=True, hide_index=True)

    with tab1:
        st.dataframe(units_df, use_container_width=True, hide_index=True)
        warnings_df = units_df[units_df['Warnung'] != ''] if not units_df.empty and 'Warnung' in units_df.columns else pd.DataFrame()
        if not warnings_df.empty:
            st.warning('Einige Verladeeinheiten überschreiten das definierte Bundgewicht.')
            st.dataframe(warnings_df, use_container_width=True, hide_index=True)

    with tab2:
        if not fuhren_log_df.empty:
            st.dataframe(fuhren_log_df, use_container_width=True, hide_index=True)
        else:
            st.warning('Es wurde keine Fuhre erzeugt.')
        st.markdown('**Pritschen-Zusammenfassung**')
        st.dataframe(edited_summary_df, use_container_width=True, hide_index=True)

    with tab3:
        if verladeart == 'Automatisch':
            st.markdown('**Automatik-Modus**')
            st.info('Automatik ist aktiv. Manuelle Zonenverladung ist nur aktiv, wenn oben „Manuell“ gewählt wird.')
            st.dataframe(edited_placements_df, use_container_width=True, hide_index=True)
        else:
            st.markdown('**Manuelle Verladung – Zonenplan**')
            st.caption('Du setzt nur die Zone. Die App berechnet X/Y/Z und die Lagen automatisch nach Standardlogik.')

            manual_df = clean_placements_dataframe(st.session_state.get('manual_placements_df', edited_placements_df))
            for col, default in {
                'Einlage_unten_mm': 0.0,
                'Einlage_oben_mm': 0.0,
                'Manuelle_Zone': '',
                'Manuelle_Reihenfolge': 0.0,
                'Manuelle_Position': '',
            }.items():
                if col not in manual_df.columns:
                    manual_df[col] = default

            platforms_used_df = _clean_manual_platforms_dataframe(st.session_state.get('manual_platforms_df', platforms_used_df))
            platforms_used_df = _manual_update_platform_statuses(platforms_used_df, manual_df)
            st.session_state['manual_platforms_df'] = platforms_used_df
            platform_names = platforms_used_df['Pritsche'].astype(str).tolist() if platforms_used_df is not None and not platforms_used_df.empty else []
            if not platform_names:
                st.warning('Keine Pritsche für manuelle Verladung vorhanden.')
            else:
                top1, top2, top3 = st.columns([1.2, 0.7, 0.6])
                selected_manual_platform = top1.selectbox('Pritsche', platform_names, key='manual_zone_platform_v55')
                if top2.button('Neue Pritsche starten', key='manual_start_next_platform_v55'):
                    platforms_used_df = _manual_start_next_fuhre(platforms_used_df)
                    st.session_state['manual_platforms_df'] = platforms_used_df
                    st.session_state['manual_plan_ready_v55'] = False
                    st.rerun()
                if top3.button('Reset', key='manual_zone_reset_v55'):
                    manual_init = clean_placements_dataframe(placements_df)
                    manual_init['Einlage_unten_mm'] = 0.0
                    manual_init['Einlage_oben_mm'] = 0.0
                    manual_init['Manuelle_Zone'] = ''
                    manual_init['Manuelle_Reihenfolge'] = 0.0
                    st.session_state['manual_placements_df'] = manual_init
                    st.session_state['manual_plan_ready_v55'] = False
                    st.session_state['manual_platforms_df'] = _clean_manual_platforms_dataframe(platforms_used_df)
                    st.rerun()

                selected_platform_row = _manual_platform_row(platforms_used_df, selected_manual_platform)
                default_spacer = safe_number(selected_platform_row.get('Einlage_zwischen_Lagen_mm'), 0.0) if selected_platform_row is not None else 40.0
                platform_status = _manual_get_platform_status(platforms_used_df, selected_manual_platform)
                s1, s2, s3 = st.columns([1.0, 0.7, 0.7])
                s1.info(f'Status Pritsche: {platform_status}')
                if s2.button('Pritsche abschliessen', key='manual_close_platform_v55', disabled=(platform_status == 'Abgeschlossen')):
                    platforms_used_df = _manual_set_platform_status(platforms_used_df, selected_manual_platform, 'Abgeschlossen')
                    st.session_state['manual_platforms_df'] = platforms_used_df
                    st.rerun()
                if s3.button('Pritsche wieder öffnen', key='manual_open_platform_v55', disabled=(platform_status != 'Abgeschlossen')):
                    platforms_used_df = _manual_set_platform_status(platforms_used_df, selected_manual_platform, 'In Bearbeitung')
                    st.session_state['manual_platforms_df'] = platforms_used_df
                    st.rerun()
                platform_locked = platform_status == 'Abgeschlossen'

                with st.expander('Kleine Pritschenübersicht', expanded=False):
                    overview_df = _manual_small_platform_overview(platforms_used_df, manual_df)
                    if overview_df.empty:
                        st.info('Noch keine Pritschenübersicht vorhanden.')
                    else:
                        st.dataframe(overview_df, use_container_width=True, hide_index=True)

                c_left, c_right = st.columns([1.0, 1.05])

                with c_left:
                    st.markdown('**Bauteil/Bund anklicken**')
                    search_text = st.text_input('Suche', value='', key='manual_zone_search_v55')
                    attr_candidates = [c for c in manual_df.columns if c not in ['X_mm', 'Y_mm', 'Z_mm', 'Manuelle_Reihenfolge', 'Manuelle_Zone']]
                    default_attrs = [c for c in ['Bauteile', 'Länge_mm', 'Breite_mm', 'Höhe_mm', 'Gewicht_kg', 'Ansicht_Attribut'] if c in attr_candidates]
                    selected_attr_cols = st.multiselect(
                        'Attribute in Auswahl anzeigen',
                        attr_candidates,
                        default=default_attrs[:5],
                        key='manual_select_attrs_v55',
                    )
                    p_series = manual_df.get('Pritsche', pd.Series(dtype=str)).astype(str)
                    # Arbeitsliste: unverladen + bereits auf dieser Pritsche.
                    work_mask = p_series.eq('NICHT VERLADEN') | p_series.eq(str(selected_manual_platform))
                    work_df = manual_df.loc[work_mask].copy()
                    if search_text.strip():
                        needle = search_text.strip().lower()
                        work_df = work_df[
                            work_df.get('Bauteile', pd.Series(dtype=str)).astype(str).str.lower().str.contains(needle, na=False)
                            | work_df.get('Einheit_ID', pd.Series(dtype=str)).astype(str).str.lower().str.contains(needle, na=False)
                        ].copy()

                    if work_df.empty:
                        st.warning('Keine Bauteile/Bunde gefunden.')
                    else:
                        # Schnellwahl über Selectbox, damit es auch bei vielen Bauteilen schnell bleibt.
                        def _unit_choice_label(r: pd.Series) -> str:
                            zone_txt = _manual_zone_label(r.get('Manuelle_Zone', ''))
                            p_txt = str(r.get('Pritsche', ''))
                            status = 'frei' if p_txt == 'NICHT VERLADEN' else zone_txt or p_txt
                            pieces = [str(r.get('Einheit_ID',''))]
                            for acol in selected_attr_cols:
                                val = r.get(acol, '')
                                if pd.isna(val):
                                    continue
                                if isinstance(val, float):
                                    pieces.append(f'{acol}: {val:.0f}')
                                else:
                                    sval = str(val)
                                    if len(sval) > 35:
                                        sval = sval[:32] + '...'
                                    pieces.append(f'{acol}: {sval}')
                            pieces.append(status)
                            return ' | '.join([p for p in pieces if p])

                        work_df = work_df.sort_values(['Manuelle_Reihenfolge', 'Einheit_ID'], kind='stable')
                        labels = [_unit_choice_label(r) for _, r in work_df.iterrows()]
                        label_to_id = {label: str(r.get('Einheit_ID')) for label, (_, r) in zip(labels, work_df.iterrows())}
                        selected_label = st.selectbox('Auswahl', labels, key='manual_zone_unit_v55')
                        selected_unit_id = label_to_id.get(selected_label, '')

                        spacer_under = st.number_input(
                            'Einlage unten mm',
                            min_value=0.0,
                            max_value=300.0,
                            value=float(default_spacer),
                            step=5.0,
                            key='manual_zone_spacer_v55',
                        )

                        if st.button('ausgewähltes Element entladen / rückgängig', key='manual_zone_unload_v55', disabled=platform_locked):
                            manual_df = _manual_unload_unit(manual_df, platforms_used_df, selected_unit_id)
                            st.session_state['manual_placements_df'] = clean_placements_dataframe(manual_df)
                            st.session_state['manual_plan_ready_v55'] = False
                            st.rerun()

                        st.markdown('**Zone anklicken**')
                        if platform_locked:
                            st.warning('Diese Pritsche ist abgeschlossen. Zum Ändern zuerst „Pritsche wieder öffnen“.')
                        b1, b2, b3, b4 = st.columns(4)
                        direct_buttons = [
                            (b1, 'VL', 'VL'),
                            (b2, 'VR', 'VR'),
                            (b3, 'HL', 'HL'),
                            (b4, 'HR', 'HR'),
                        ]
                        for col_obj, zone_code, label in direct_buttons:
                            if col_obj.button(label, key=f'manual_zone_direct_{zone_code}_v55', disabled=platform_locked):
                                manual_df = _manual_assign_unit_to_zone(
                                    manual_df,
                                    platforms_used_df,
                                    selected_unit_id,
                                    selected_platform_row,
                                    zone_code,
                                    float(spacer_under),
                                )
                                st.session_state['manual_placements_df'] = clean_placements_dataframe(manual_df)
                                st.session_state['manual_plan_ready_v55'] = False
                                st.rerun()

                    with st.expander('Pritschenwerte kurz ändern', expanded=False):
                        p_mask = platforms_used_df['Pritsche'].astype(str).eq(str(selected_manual_platform))
                        p_cols = [c for c in [
                            'Pritsche', 'Länge_mm', 'Breite_mm', 'Max_Höhe_mm',
                            'Überhang_hinten_mm', 'Überhang_vorne_mm',
                            'Kantholz_erste_Lage_mm', 'Einlage_zwischen_Lagen_mm',
                            'Max_Gewicht_kg', 'Eigengewicht_Pritsche_kg'
                        ] if c in platforms_used_df.columns]
                        original_platform_edit = platforms_used_df.loc[p_mask, p_cols].copy().reset_index(drop=True)
                        edited_platform = st.data_editor(
                            original_platform_edit,
                            use_container_width=True,
                            hide_index=True,
                            num_rows='fixed',
                            key='manual_zone_platform_values_v55',
                        )
                        platform_changed = (
                            not edited_platform.empty
                            and not edited_platform.reset_index(drop=True).equals(original_platform_edit)
                        )
                        if platform_changed:
                            for col in edited_platform.columns:
                                platforms_used_df.loc[p_mask, col] = edited_platform[col].values
                            platforms_used_df = _clean_manual_platforms_dataframe(platforms_used_df)
                            st.session_state['manual_platforms_df'] = platforms_used_df
                            st.session_state['manual_plan_ready_v55'] = False
                            st.rerun()

                with c_right:
                    st.markdown('**Ladeplan BSD / Zonenansicht**')
                    if selected_platform_row is not None:
                        p_mask_info = manual_df.get('Pritsche', pd.Series(dtype=str)).astype(str).eq(str(selected_manual_platform))
                        placed_info = manual_df[p_mask_info].copy()
                        used_height = 0.0
                        used_width = 0.0
                        if not placed_info.empty and 'Z_mm' in placed_info.columns:
                            z_vals = pd.to_numeric(placed_info.get('Z_mm'), errors='coerce')
                            h_vals = pd.to_numeric(placed_info.get('Höhe_mm'), errors='coerce')
                            if not (z_vals + h_vals).dropna().empty:
                                used_height = float((z_vals + h_vals).dropna().max())
                            y_vals = pd.to_numeric(placed_info.get('Y_mm'), errors='coerce')
                            b_vals = pd.to_numeric(placed_info.get('Breite_mm'), errors='coerce')
                            if not y_vals.dropna().empty and not (y_vals + b_vals).dropna().empty:
                                used_width = float((y_vals + b_vals).dropna().max() - y_vals.dropna().min())
                        used_length = 0.0
                        x_vals = pd.to_numeric(placed_info.get('X_mm'), errors='coerce') if not placed_info.empty and 'X_mm' in placed_info.columns else pd.Series(dtype=float)
                        l_vals = pd.to_numeric(placed_info.get('Länge_mm'), errors='coerce') if not placed_info.empty and 'Länge_mm' in placed_info.columns else pd.Series(dtype=float)
                        if not x_vals.dropna().empty and not (x_vals + l_vals).dropna().empty:
                            used_length = float((x_vals + l_vals).dropna().max() - x_vals.dropna().min())

                        max_lade_length = _manual_effective_length(selected_platform_row)
                        base_length = safe_number(selected_platform_row.get('Länge_mm'))
                        st.info(
                            f"Lade-Länge max.: {max_lade_length:.0f} mm "
                            f"(Pritsche {base_length:.0f} mm) · "
                            f"aktuelle Länge belegt: {used_length:.0f} mm · "
                            f"max. Breite: {safe_number(selected_platform_row.get('Breite_mm')):.0f} mm · "
                            f"aktuelle Breite: {used_width:.0f} mm · "
                            f"max. Ladehöhe: {safe_number(selected_platform_row.get('Max_Höhe_mm')):.0f} mm · "
                            f"aktuelle Ladehöhe: {used_height:.0f} mm"
                        )
                    zone_grid = _manual_zone_grid(manual_df, selected_manual_platform)
                    if zone_grid.empty:
                        st.info('Noch keine Bauteile auf dieser Pritsche gesetzt.')
                    else:
                        st.dataframe(zone_grid, use_container_width=True, hide_index=True)

                    loaded_on_platform = manual_df[
                        manual_df.get('Pritsche', pd.Series(dtype=str)).astype(str).eq(str(selected_manual_platform))
                    ].copy()
                    if not loaded_on_platform.empty:
                        def _undo_label(r: pd.Series) -> str:
                            zone_txt = _manual_zone_label(r.get('Manuelle_Zone', ''))
                            b = str(r.get('Bauteile') or r.get('Einheit_ID') or '')
                            return f"{r.get('Einheit_ID','')} | {b} | {zone_txt}"
                        undo_labels = [_undo_label(r) for _, r in loaded_on_platform.iterrows()]
                        undo_map = {label: str(r.get('Einheit_ID')) for label, (_, r) in zip(undo_labels, loaded_on_platform.iterrows())}
                        u1, u2 = st.columns([1.4, 0.6])
                        undo_choice = u1.selectbox('verladenes Element rückgängig', undo_labels, key='manual_zone_undo_choice_v55')
                        if u2.button('rückgängig', key='manual_zone_undo_button_v55', disabled=platform_locked):
                            manual_df = _manual_unload_unit(manual_df, platforms_used_df, undo_map.get(undo_choice, ''))
                            st.session_state['manual_placements_df'] = clean_placements_dataframe(manual_df)
                            st.session_state['manual_plan_ready_v55'] = False
                            st.rerun()

                    counts_df = manual_df.copy()
                    loaded_now = int((counts_df.get('Pritsche', pd.Series(dtype=str)).astype(str) != 'NICHT VERLADEN').sum()) if not counts_df.empty else 0
                    free_now = int((counts_df.get('Pritsche', pd.Series(dtype=str)).astype(str) == 'NICHT VERLADEN').sum()) if not counts_df.empty else 0
                    ready_txt = 'berechnet' if bool(st.session_state.get('manual_plan_ready_v55', False)) else 'noch nicht berechnet'
                    st.write(f'Verladen: {loaded_now} · offen: {free_now} · Status: {ready_txt}')

                    if st.button('Ladeplan berechnen / prüfen', key='manual_zone_calculate_plan_v55', type='primary'):
                        validation_df = _manual_validate_zone_plan(st.session_state.get('manual_placements_df', manual_df), platforms_used_df)
                        if validation_df is not None and not validation_df.empty:
                            st.session_state['manual_validation_issues_v55'] = validation_df
                            st.session_state['manual_plan_ready_v55'] = False
                            st.error('Ladeplan wurde nicht berechnet: Grenzen überschritten.')
                            st.dataframe(validation_df, use_container_width=True, hide_index=True)
                        else:
                            calculated_df = _manual_reflow_all_zones(st.session_state.get('manual_placements_df', manual_df), platforms_used_df)
                            st.session_state['manual_placements_df'] = clean_placements_dataframe(calculated_df)
                            platforms_used_df = _manual_set_platform_status(platforms_used_df, selected_manual_platform, 'Berechnet / geprüft')
                            st.session_state['manual_platforms_df'] = platforms_used_df
                            st.session_state['manual_validation_issues_v55'] = pd.DataFrame()
                            st.session_state['manual_plan_ready_v55'] = True
                            st.rerun()

                    validation_show = st.session_state.get('manual_validation_issues_v55', pd.DataFrame())
                    if validation_show is not None and not validation_show.empty:
                        st.warning('Letzte Kontrolle hat Fehler gefunden:')
                        st.dataframe(validation_show, use_container_width=True, hide_index=True)

                    with st.expander('Ansicht prüfen', expanded=False):
                        st.caption('Die Ansicht wird bewusst erst nach Klick erstellt. Dadurch reagiert die manuelle Verladung schneller.')
                        view_choice = st.radio(
                            'Ansicht',
                            ['Draufsicht', 'Seitenansicht', 'Vorderansicht', 'Rückansicht'],
                            horizontal=True,
                            key='manual_zone_view_v55',
                        )
                        create_view = st.button('Ansicht erstellen / aktualisieren', key='manual_zone_create_view_v55')
                        if create_view:
                            if not bool(st.session_state.get('manual_plan_ready_v55', False)):
                                st.warning('Bitte zuerst „Ladeplan berechnen / prüfen“ klicken.')
                                st.stop()
                            manual_view_df = clean_placements_dataframe(st.session_state.get('manual_placements_df', manual_df))
                            display_tmp, _ = add_underbau_rows_to_placements(
                                manual_view_df,
                                platforms_used_df,
                                min_support_ratio=float(min_support_ratio),
                                min_underbau_height=float(min_underbau_height),
                                enabled=False,
                            )
                            view_map = {
                                'Draufsicht': 'top',
                                'Seitenansicht': 'side',
                                'Vorderansicht': 'front',
                                'Rückansicht': 'back',
                            }
                            st.plotly_chart(
                                draw_loading_view(display_tmp, platforms_used_df, selected_manual_platform, view_map[view_choice]),
                                use_container_width=True,
                                key=f'manual_zone_plot_v55_{selected_manual_platform}_{view_map[view_choice]}',
                            )
                        else:
                            st.caption('Noch keine Ansicht erstellt.')

                # Turbo-Manuell:
                # Beim Anklicken nur Zonen speichern. Fachliche Berechnung erst über Button.
                edited_placements_df = clean_placements_dataframe(st.session_state.get('manual_placements_df', manual_df))
                if bool(st.session_state.get('manual_plan_ready_v55', False)):
                    edited_summary_df = recompute_summary_from_placements(edited_placements_df, platforms_used_df) if not platforms_used_df.empty else summary_df
                    warnings_plan_df = compute_loading_warnings(edited_placements_df, platforms_used_df)
                    display_placements_df = edited_placements_df.copy()
                    if bool(underbau_enabled):
                        display_placements_df, underbau_warnings_df = add_underbau_rows_to_placements(
                            edited_placements_df,
                            platforms_used_df,
                            min_support_ratio=float(min_support_ratio),
                            min_underbau_height=float(min_underbau_height),
                            enabled=True,
                        )
                        if underbau_warnings_df is not None and not underbau_warnings_df.empty:
                            warnings_plan_df = pd.concat([warnings_plan_df, underbau_warnings_df], ignore_index=True, sort=False) if not warnings_plan_df.empty else underbau_warnings_df
                else:
                    edited_summary_df = recompute_summary_from_placements(edited_placements_df, platforms_used_df) if not platforms_used_df.empty else summary_df
                    warnings_plan_df = pd.DataFrame()
                    display_placements_df = edited_placements_df.copy()

    with tab4:
        bsd_header_df, bsd_matrix_df = create_all_bsd_matrices(
            display_placements_df,
            platforms_used_df,
            edited_summary_df,
            warnings_plan_df,
            project_meta=project_meta,
        )
        st.markdown('**Ladeplan BSD je Pritsche**')
        st.caption('Die Tabelle wird automatisch für jede belegte Pritsche erstellt. V55: kompakte BSD-Spaltenliste; jede Position wird separat ohne künstliche Leerzellen aufgebaut.')

        if bsd_header_df.empty or bsd_matrix_df.empty:
            st.warning('Für die aktuelle Verladung wurde kein Ladeplan BSD erzeugt.')
        else:
            selected_bsd_platform = st.selectbox('Pritsche für Ladeplan BSD', bsd_header_df['Pritsche'].astype(str).tolist(), key='bsd_platform_select')
            bsd_info = bsd_header_df[bsd_header_df['Pritsche'].astype(str) == selected_bsd_platform]
            if not bsd_info.empty:
                hrow = bsd_info.iloc[0]
                c1, c2, c3, c4, c5 = st.columns(5)
                c1.metric('Gesamtgewicht', f"{safe_number(hrow.get('Gesamtgewicht_kg')):.0f} kg")
                c2.metric('Frachthöhe', f"{safe_number(hrow.get('Frachthöhe_mm')):.0f} mm")
                c3.metric('Länge gesamt', f"{safe_number(hrow.get('Länge_gesamt_mm')):.0f} mm")
                c4.metric('Breite gesamt', f"{safe_number(hrow.get('Breite_gesamt_mm')):.0f} mm")
                c5.metric('Warnungen', int(safe_number(hrow.get('Warnungen'))))

            selected_matrix = bsd_matrix_df[bsd_matrix_df['Pritsche'].astype(str) == selected_bsd_platform].copy()
            display_cols = [
                'Lage', 'Vorne links', 'Vorne rechts', 'Hinten links', 'Hinten rechts',
                'Höhe_mm', 'Breite_mm', 'Gesamtlänge_mm', 'Gewicht_kg', 'Anzahl_Einheiten'
            ]
            st.dataframe(selected_matrix[display_cols], use_container_width=True, hide_index=True)

            with st.expander('Alle Ladeplan-BSD-Kopfdaten und Matrizen anzeigen', expanded=False):
                st.markdown('**Kopfdaten**')
                st.dataframe(bsd_header_df, use_container_width=True, hide_index=True)
                st.markdown('**Matrix alle Pritschen**')
                st.dataframe(bsd_matrix_df, use_container_width=True, hide_index=True)

    with tab5:
        if warnings_plan_df.empty:
            st.success('Keine Warnungen gefunden.')
        else:
            severe_count = len(warnings_plan_df)
            st.warning(f'{severe_count} Warnung(en) gefunden.')
            st.dataframe(warnings_plan_df, use_container_width=True, hide_index=True)

    with tab6:
        if platforms_used_df.empty:
            st.warning('Keine Pritsche für die Ansicht vorhanden.')
        else:
            selected_platform = st.selectbox('Fuhre / Pritsche für Ansicht', platforms_used_df['Pritsche'].tolist())
            info_row = edited_summary_df[edited_summary_df['Pritsche'] == selected_platform]
            if not info_row.empty:
                row = info_row.iloc[0]
                c1, c2, c3, c4 = st.columns(4)
                c1.metric('Länge genutzt', f"{row['Länge genutzt_mm']:.0f} mm")
                c2.metric('Breite genutzt', f"{row['Breite genutzt_mm']:.0f} mm")
                c3.metric('Höhe genutzt', f"{row['Höhe genutzt_mm']:.0f} mm")
                c4.metric('Gesamtgewicht', f"{safe_number(row.get('Gesamtgewicht inkl. Pritsche_kg')):.0f} kg")

            col1, col2 = st.columns(2)
            with col1:
                st.plotly_chart(draw_loading_view(display_placements_df, platforms_used_df, selected_platform, 'side'), use_container_width=True, key=f'ansicht_side_{selected_platform}')
            with col2:
                st.plotly_chart(draw_loading_view(display_placements_df, platforms_used_df, selected_platform, 'back'), use_container_width=True, key=f'ansicht_back_{selected_platform}')
            col3, col4 = st.columns(2)
            with col3:
                st.plotly_chart(draw_loading_view(display_placements_df, platforms_used_df, selected_platform, 'top'), use_container_width=True, key=f'ansicht_top_{selected_platform}')
            with col4:
                st.plotly_chart(draw_loading_view(display_placements_df, platforms_used_df, selected_platform, 'front'), use_container_width=True, key=f'ansicht_front_{selected_platform}')

    with tab7:
        # Falls der Ladeplan-BSD-Tab nicht angezeigt wurde, trotzdem für den Export erstellen.
        if 'bsd_header_df' not in locals() or 'bsd_matrix_df' not in locals():
            bsd_header_df, bsd_matrix_df = create_all_bsd_matrices(display_placements_df, platforms_used_df, edited_summary_df, warnings_plan_df, project_meta=project_meta)

        st.subheader('A3-Pritschenplan PDF')
        st.info(
            'Hauptausgabe ist wieder der A3-Pritschenplan als PDF. Die PDF enthält die grosse grafische Darstellung '
            'mit Seitenansicht, Rückansicht, Vorderansicht, Draufsicht und den kleinen Ladeplan-BSD-Seiten je Pritsche. '
            'Excel bleibt als Begleitdatei mit den kleinen BSD-Zetteln erhalten.'
        )

        try:
            pdf_data = create_loading_pdf(
                display_placements_df,
                platforms_used_df,
                edited_summary_df,
                warnings_plan_df,
                project_name=source_name,
                project_meta=project_meta,
                logo_bytes=logo_bytes,
            )
            st.download_button(
                label='A3-Pritschenplan als PDF herunterladen',
                data=pdf_data,
                file_name=f"{source_name.replace('.bvx', '').replace('.BVX', '').replace('.xlsx', '')}_pritschenplan_a3.pdf",
                mime='application/pdf',
                type='primary',
            )
        except RuntimeError as exc:
            st.warning(str(exc))

        st.divider()
        st.subheader('Excel-Daten / kleine BSD-Zettel')
        st.caption('Die Excel-Datei enthält die kleinen BSD-/Pritschenzettel und die Zusatzregister Bauteile, Fuhrenoptionen, Ladeplan_BSD_Kopf und Projektkopf.')

        excel_data = create_loading_excel(
            sorted_parts,
            units_df,
            display_placements_df,
            platforms_used_df,
            edited_summary_df,
            options_df=options_edit,
            fuhren_log_df=fuhren_log_df,
            warnings_df=warnings_plan_df,
            bsd_header_df=bsd_header_df,
            bsd_matrix_df=bsd_matrix_df,
            project_meta=project_meta,
            logo_bytes=logo_bytes,
        )
        st.download_button(
            label='Excel-Begleitdatei herunterladen',
            data=excel_data,
            file_name=f"{source_name.replace('.bvx', '').replace('.BVX', '').replace('.xlsx', '')}_verladeplan_bsd.xlsx",
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        st.markdown('''
        **Hinweis:** Diese Version erstellt den A3-Pritschenplan als Hauptausgabe im PDF.
        Die Excel-Datei enthält kleine BSD-Zettel. Manuelles Umplatzieren erfolgt über Excel-artige Tabelle, fachliche Vergabe und Live-Skizze.
        Drag-and-drop ist nicht enthalten.
        ''')


def main():
    st.set_page_config(
        page_title='BVX Auswertung / Verladeplanung',
        page_icon='🔧',
        layout='wide',
        initial_sidebar_state='expanded'
    )

    st.markdown('''
        <style>
        .main-header {
            font-size: 2.2rem;
            font-weight: bold;
            color: #1a1a2e;
            margin-bottom: 1rem;
        }
        </style>
    ''', unsafe_allow_html=True)

    st.markdown('<p class="main-header">BVX Auswertung / Verladeplanung</p>', unsafe_allow_html=True)

    with st.sidebar:
        try:
            st.image(get_embedded_default_logo(), width=70)
            st.caption('App-Logo')
        except Exception:
            pass
        st.header('Modul')
        module = st.radio('Bereich wählen', ['BVX Auswertung', 'Verladeplanung'], index=0)

        st.divider()
        if module == 'BVX Auswertung':
            st.subheader('BVX Analyse Import')
            analysis_file = st.file_uploader(
                'BVX-Datei für Analyse hochladen',
                type=['bvx', 'BVX'],
                key='analysis_upload',
                help='Normale BVX-Auswertung mit Operationen und Export.'
            )
            loading_file = None
            parts_excel_file = None
            logo_file = None
            if analysis_file:
                st.success(f'{analysis_file.name}')
        else:
            st.subheader('Verlade Import')
            loading_file = st.file_uploader(
                'BVX-Datei für Verladung hochladen',
                type=['bvx', 'BVX'],
                key='loading_upload',
                help='Eigenständiger Import für Verladeplanung.'
            )
            parts_excel_file = st.file_uploader(
                'Bauteile-Excel laden, falls keine BVX vorhanden ist',
                type=['xlsx'],
                key='parts_excel_upload',
                help='Excel mit Blatt Bauteile: Länge_mm, Breite_mm, Höhe_mm usw.'
            )
            transport_excel_file = st.file_uploader(
                'Excel Pritschen/Fuhren laden',
                type=['xlsx'],
                key='transport_excel_upload',
                help='Stammdaten mit Fuhrenoptionen, Pritschen und Standards.'
            )
            logo_file = st.file_uploader(
                'Logo für Pläne ersetzen (optional)',
                type=['png', 'jpg', 'jpeg'],
                key='logo_upload',
                help='Ohne Upload verwendet die App das Standardlogo. Upload ersetzt nur das Logo in den A3-PDF-Plänen.'
            )
            analysis_file = None
            if loading_file:
                st.success(f'BVX: {loading_file.name}')
            if parts_excel_file and not loading_file:
                st.success(f'Bauteile-Excel: {parts_excel_file.name}')
            if transport_excel_file:
                st.success(f'Excel: {transport_excel_file.name}')
            if logo_file:
                st.success(f'Logo: {logo_file.name}')

    if module == 'BVX Auswertung':
        render_analysis_module(analysis_file)
    else:
        render_loading_module(loading_file, transport_excel_file, logo_file, parts_excel_file)


if __name__ == '__main__':
    main()
